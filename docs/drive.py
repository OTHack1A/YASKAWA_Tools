"""YASKAWA DriveWizard Industrial (.YDWIProj) parser and report generator."""

import os
import re

try:
    from docs.ga500_params import GA500_PARAMS, get_param_name
except ImportError:
    GA500_PARAMS = {}
    def get_param_name(code, lang='EN'):
        return GA500_PARAMS.get(code, code)

_CAT_ORDER = ['A', 'b', 'C', 'd', 'E', 'F', 'H', 'L', 'n', 'o', 'q', 'r']

_CATEGORIES = {
    'A': 'A — Initialization',
    'b': 'b — Application',
    'C': 'C — Tuning',
    'd': 'd — Reference',
    'E': 'E — Motor Parameters',
    'F': 'F — Options / Comms',
    'H': 'H — Terminals',
    'L': 'L — Protection',
    'n': 'n — Advanced Adjustment',
    'o': 'o — Operator',
    'q': 'q — DWEZ',
    'r': 'r — DWEZ Advanced',
}


def _cat_label(cat):
    return _CATEGORIES.get(cat, cat)


def _xl_safe(val):
    """Prevent openpyxl from treating strings as Excel formulas.
    Strings starting with =, +, @ are auto-detected as formulas; prefix with
    a space so Excel stores them as plain text."""
    if isinstance(val, str) and val and val[0] in ('=', '+', '@'):
        return ' ' + val
    return val


# ── Parser ─────────────────────────────────────────────────────────────────────

def parse_ydwiproj(filepath):
    """Parse a .YDWIProj or .YDWGProj file.
    Returns (info, params_by_cat) or raises RuntimeError.
    YDWGProj files have an extra flat-header block before [SECTION] markers.
    """
    try:
        with open(filepath, 'r', encoding='latin-1', errors='replace') as f:
            content = f.read()
    except OSError as exc:
        raise RuntimeError(f'Cannot open: {exc}') from exc

    info = {}

    # ── YDWGProj flat header (key : value lines before first [SECTION]) ────────
    first_section = re.search(r'\n\[', content)
    header_block  = content[:first_section.start()] if first_section else ''
    for line in header_block.splitlines():
        line = line.strip()
        if not line:
            continue
        m = re.match(r'^([A-Z][A-Z0-9 ]*?)\s*:\s*(.*)$', line)
        if m:
            info[m.group(1).strip()] = m.group(2).strip()

    # ── Standard [SECTION] fields ─────────────────────────────────────────────
    m = re.search(r'\[FILE INFO\](.*?)(?=\n\[|\Z)', content, re.DOTALL)
    info['file_info'] = m.group(1).strip() if m else ''

    for key in ('PROJECT NAME', 'USER NAME', 'DBINFO', 'DBNAME',
                'DRIVE SERIES', 'SOFTWARE VERSION', 'DRIVE MODEL',
                'CONTROL METHOD', 'INITIALIZATION MODE'):
        m = re.search(r'\[' + re.escape(key) + r'\]\s*\n(.*?)(?=\n\[|\Z)',
                      content, re.DOTALL)
        info[key] = m.group(1).strip() if m else ''

    params_by_cat = {}
    for line in content.splitlines():
        line = line.strip()
        m = re.match(r'^([A-Za-z]\d+-\d+)=([^;]*);(.*)$', line)
        if not m:
            continue
        code = m.group(1)
        val  = m.group(2).strip()
        meta = {}
        for part in m.group(3).split(';'):
            if ':' in part:
                k, v = part.split(':', 1)
                meta[k.strip()] = v.strip()
        cat = code[0]
        params_by_cat.setdefault(cat, []).append((
            code, val,
            meta.get('DEF', ''), meta.get('MIN', ''),
            meta.get('MAX', ''), meta.get('PRV', ''),
        ))

    return info, params_by_cat


# ── Excel generator ────────────────────────────────────────────────────────────

def generate_drive_excel(info, params_by_cat, output_path,
                         col_param='Param', col_name='Name',
                         col_value='Value', col_default='Default',
                         col_min='Min', col_max='Max', col_prev='Previous',
                         lang='EN', log_fn=None):
    """Generate Excel: Info tab + one tab per parameter category. Returns True/False."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        if log_fn:
            log_fn('log_error_generic', f'openpyxl: {exc}')
        return False

    HDR_FILL     = PatternFill('solid', fgColor='D97757')
    HDR_FNT  = Font(color='FFFFFF', bold=True, name='Calibri', size=10)
    CENTER   = Alignment(horizontal='center', vertical='center')
    GRAY     = PatternFill('solid', fgColor='EFEFEF')
    CODE_FNT = Font(name='Courier New', size=9, color='003366')
    CODE_ALN = Alignment(horizontal='center')
    R_ALN    = Alignment(horizontal='right')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Info tab ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet('Info')
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 50

    c = ws.cell(row=1, column=1, value=_xl_safe('YASKAWA DriveWizard — Drive Info'))
    ws.merge_cells('A1:B1')
    c.fill = HDR_FILL; c.font = HDR_FNT; c.alignment = CENTER
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = 'A2'

    row = 2
    for label, key in [
        ('Project Name',        'PROJECT NAME'),
        ('User',                'USER NAME'),
        ('Drive Series',        'DRIVE SERIES'),
        ('Drive Model',         'DRIVE MODEL'),
        ('Software Version',    'SOFTWARE VERSION'),
        ('Control Method',      'CONTROL METHOD'),
        ('Init. Mode',          'INITIALIZATION MODE'),
        ('DB Info',             'DBINFO'),
        ('DB Name',             'DBNAME'),
    ]:
        val = info.get(key, '').strip()
        if not val:
            continue
        ca = ws.cell(row=row, column=1, value=_xl_safe(label))
        ca.font = Font(bold=True, name='Calibri', size=10); ca.fill = GRAY
        ws.cell(row=row, column=2, value=_xl_safe(val)).font = Font(name='Calibri', size=10)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='— File Info —').font = Font(
        bold=True, name='Calibri', size=10)
    row += 1
    for fi_line in info.get('file_info', '').splitlines():
        clean = fi_line.strip()
        if clean:
            c = ws.cell(row=row, column=1, value=_xl_safe(clean))
            c.font = Font(name='Courier New', size=9)
            ws.merge_cells(f'A{row}:B{row}')
            row += 1

    # ── Category tabs ─────────────────────────────────────────────────────────
    headers = [(col_param, 14), (col_name, 36), (col_value, 14),
               (col_default, 12), (col_min, 10), (col_max, 10), (col_prev, 12)]
    NAME_FNT = Font(name='Calibri', size=9)
    L_ALN    = Alignment(horizontal='left')

    for cat in _CAT_ORDER:
        if cat not in params_by_cat:
            continue
        rows = params_by_cat[cat]
        ws = wb.create_sheet(title=cat)

        for ci, (hdr_text, width) in enumerate(headers, start=1):
            c = ws.cell(row=1, column=ci, value=hdr_text)
            c.fill = HDR_FILL; c.font = HDR_FNT; c.alignment = CENTER
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 18
        ws.freeze_panes = 'A2'

        for ri, (code, val, dflt, mn, mx, prv) in enumerate(rows, start=2):
            c0 = ws.cell(row=ri, column=1, value=_xl_safe(code))
            c0.fill = GRAY; c0.font = CODE_FNT; c0.alignment = CODE_ALN
            cn = ws.cell(row=ri, column=2, value=_xl_safe(get_param_name(code, lang)))
            cn.font = NAME_FNT; cn.alignment = L_ALN
            for ci, v in enumerate([val, dflt, mn, mx, prv], start=3):
                ws.cell(row=ri, column=ci, value=_xl_safe(v)).alignment = R_ALN

        if log_fn:
            log_fn('log_drive_cat', cat, len(rows))

    try:
        wb.save(output_path)
        if log_fn:
            log_fn('log_drive_excel_saved', output_path)
        return True
    except Exception as exc:
        if log_fn:
            log_fn('log_error_generic', f'Drive Excel: {exc}')
        return False


# ── PDF generator ──────────────────────────────────────────────────────────────

def generate_drive_pdf(info, params_by_cat, output_path,
                       col_param='Param', col_name='Name',
                       col_value='Value', col_default='Default',
                       col_min='Min', col_max='Max', col_prev='Previous',
                       toc_title='Table of Contents',
                       info_section_title='Drive Information',
                       lang='EN', log_fn=None):
    """Generate A4 PDF: info page → TOC → parameter tables. Returns True/False."""
    try:
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         PageBreak, Paragraph, Spacer, Flowable)
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor, white, black
    except ImportError as exc:
        if log_fn:
            log_fn('log_error_generic', f'reportlab: {exc}')
        return False

    from docs.utils import pdf_font, xml_escape
    fn   = pdf_font(lang)
    fn_b = pdf_font(lang, bold=True)

    accent = HexColor('#D97757')
    row_gray  = HexColor('#EFEFEF')

    styles = getSampleStyleSheet()

    title_s = ParagraphStyle('DrTitle', parent=styles['Normal'],
        fontSize=14, fontName=fn_b,
        textColor=HexColor('#A85C42'), spaceAfter=4*mm)

    section_s = ParagraphStyle('DrSection', parent=styles['Normal'],
        fontSize=11, fontName=fn_b,
        textColor=accent, spaceAfter=2*mm, spaceBefore=3*mm)

    code_s = ParagraphStyle('DrCode', parent=styles['Normal'],
        fontSize=8, fontName='Courier', spaceAfter=0.5*mm)

    toc_s = ParagraphStyle('DrToc', parent=styles['Normal'],
        fontSize=10, fontName=fn, spaceAfter=2*mm)

    cat_text_s = ParagraphStyle('DrCatText', parent=styles['Normal'],
        fontSize=12, fontName=fn_b, textColor=black,
        leftIndent=4*mm)

    # ── Category tracker (zero-size flowable) ─────────────────────────────────
    tracker = {'cat': ''}

    class _SetCat(Flowable):
        def __init__(self, label):
            Flowable.__init__(self)
            self._label = label
            self.width = self.height = 0
        def wrap(self, *a): return 0, 0
        def draw(self): tracker['cat'] = self._label

    # ── Page callback: company header + footer ────────────────────────────────
    def draw_page(canvas, doc):
        from docs.pdf_header import draw_page_header
        draw_page_header(canvas, doc)
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(HexColor('#888888'))
        pn = canvas.getPageNumber()
        canvas.drawCentredString(A4[0] / 2, 8 * mm, f'- {pn} -')
        cat = tracker.get('cat', '')
        if cat:
            canvas.drawRightString(A4[0] - 15 * mm, 8 * mm, cat)
        canvas.restoreState()

    # ── Column widths (total = 180mm) ─────────────────────────────────────────
    # Code(18) Name(46) Value(18) Default(18) Min(18) Max(18) Previous(44)
    COL_W   = [18*mm, 46*mm, 18*mm, 18*mm, 18*mm, 18*mm, 44*mm]
    HDR_ROW = [col_param, col_name, col_value, col_default, col_min, col_max, col_prev]
    PAGE_W  = A4[0] - 30*mm  # 180mm

    name_s = ParagraphStyle('DrName', parent=styles['Normal'],
        fontSize=6, fontName=fn, leading=7)

    def _name_cell(code):
        n = get_param_name(code, lang)
        return Paragraph(xml_escape(n), name_s) if n else ''

    def cat_table_style(n):
        ts = TableStyle([
            ('BACKGROUND',  (0, 0), (-1, 0), accent),
            ('TEXTCOLOR',   (0, 0), (-1, 0), black),
            ('FONTNAME',    (0, 0), (-1, 0), fn_b),
            ('FONTSIZE',    (0, 0), (-1, 0), 7),
            ('ALIGN',       (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME',    (0, 1), (0, -1), 'Courier'),
            ('FONTSIZE',    (0, 1), (0, -1), 7),
            ('FONTNAME',    (2, 1), (-1, -1), 'Courier'),
            ('FONTSIZE',    (2, 1), (-1, -1), 7),
            ('ROWHEIGHT',   (0, 0), (0, 0), 12),
            ('ROWHEIGHT',   (0, 1), (-1, -1), 10),
            ('GRID',        (0, 0), (-1, -1), 0.25, HexColor('#CCCCCC')),
            ('ALIGN',       (0, 1), (0, -1), 'CENTER'),
            ('ALIGN',       (1, 1), (1, -1), 'LEFT'),
            ('ALIGN',       (2, 1), (-1, -1), 'RIGHT'),
            ('LEFTPADDING',  (1, 1), (1, -1), 2),
            ('RIGHTPADDING', (1, 1), (1, -1), 2),
        ])
        for r in range(1, n + 1):
            if r % 2 == 0:
                ts.add('BACKGROUND', (0, r), (-1, r), row_gray)
        return ts

    story = []

    # ── Page 1: Drive Info ────────────────────────────────────────────────────
    story.append(Paragraph('YASKAWA GA500 — Drive Info', title_s))
    story.append(Spacer(1, 3*mm))

    meta_rows = []
    for label, key in [
        ('Project Name',        'PROJECT NAME'),
        ('User',                'USER NAME'),
        ('Drive Series',        'DRIVE SERIES'),
        ('Drive Model',         'DRIVE MODEL'),
        ('Software Version',    'SOFTWARE VERSION'),
        ('Control Method',      'CONTROL METHOD'),
        ('Init. Mode',          'INITIALIZATION MODE'),
    ]:
        val = info.get(key, '').strip()
        if val:
            meta_rows.append([label, val])

    if meta_rows:
        info_tbl = Table(meta_rows, colWidths=[55*mm, 125*mm])
        info_tbl.setStyle(TableStyle([
            ('FONTNAME',        (0, 0), (0, -1), fn_b),
            ('FONTNAME',        (1, 0), (1, -1), fn),
            ('FONTSIZE',        (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS',  (0, 0), (-1, -1), [row_gray, white]),
            ('GRID',            (0, 0), (-1, -1), 0.25, HexColor('#CCCCCC')),
            ('VALIGN',          (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',      (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING',   (0, 0), (-1, -1), 3),
        ]))
        story.append(info_tbl)

    # ── YDWGProj extra header info (PROGRAM, VERSION, DATE, DRIVE SERIAL NO) ──
    _YDWG_KEYS = ['PROGRAM', 'VERSION', 'DATE', 'DRIVE SERIAL NO',
                  'DEVICE START', 'DBINFO', 'DBNAME']
    extra_rows = [(k, info[k]) for k in _YDWG_KEYS
                  if k in info and info[k] and k not in
                  ('PROJECT NAME', 'USER NAME', 'DRIVE SERIES', 'DRIVE MODEL',
                   'SOFTWARE VERSION', 'CONTROL METHOD', 'INITIALIZATION MODE')]
    if extra_rows:
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(info_section_title, section_s))
        extra_tbl = Table(extra_rows, colWidths=[55*mm, 125*mm])
        extra_tbl.setStyle(TableStyle([
            ('FONTNAME',        (0, 0), (0, -1), fn_b),
            ('FONTNAME',        (1, 0), (1, -1), fn),
            ('FONTSIZE',        (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS',  (0, 0), (-1, -1), [row_gray, white]),
            ('GRID',            (0, 0), (-1, -1), 0.25, HexColor('#CCCCCC')),
            ('VALIGN',          (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',      (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING',   (0, 0), (-1, -1), 3),
        ]))
        story.append(extra_tbl)

    story.append(Spacer(1, 4*mm))
    story.append(Paragraph('File Info', section_s))
    for line in info.get('file_info', '').splitlines():
        clean = line.strip()
        if clean:
            story.append(Paragraph(xml_escape(clean), code_s))

    story.append(PageBreak())

    # ── Page 2: TOC ───────────────────────────────────────────────────────────
    story.append(Paragraph(toc_title, title_s))
    story.append(Spacer(1, 3*mm))

    toc_rows = []
    for cat in _CAT_ORDER:
        if cat not in params_by_cat:
            continue
        count = len(params_by_cat[cat])
        label = _cat_label(cat)
        link_p = Paragraph(
            f'<link href="#cat_{cat}" color="#0044AA">{xml_escape(label)}</link>',
            toc_s)
        cnt_p = Paragraph(
            f'<font color="#666666">{count}</font>',
            ParagraphStyle('DrTocCnt', parent=styles['Normal'],
                           fontSize=10, alignment=2))
        toc_rows.append([link_p, cnt_p])

    if toc_rows:
        toc_tbl = Table(toc_rows, colWidths=[155*mm, 25*mm])
        toc_tbl.setStyle(TableStyle([
            ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',   (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 2),
            ('LINEBELOW',    (0, 0), (-1, -1), 0.25, HexColor('#DDDDDD')),
        ]))
        story.append(toc_tbl)

    # reportlab calls draw_page at page BEGIN (beforeDrawPage), so _SetCat must be
    # placed BEFORE the PageBreak that ends the previous section — this pre-loads
    # tracker['cat'] with the correct value for the next page's footer.
    cats = [c for c in _CAT_ORDER if c in params_by_cat]
    if cats:
        story.append(_SetCat(_cat_label(cats[0])))
    story.append(PageBreak())

    # ── Category pages ────────────────────────────────────────────────────────
    for idx, cat in enumerate(cats):
        rows = params_by_cat[cat]
        label = _cat_label(cat)

        # Colored heading with anchor for TOC links
        head_tbl = Table(
            [[Paragraph(f'<a name="cat_{cat}"/>{xml_escape(label)}', cat_text_s)]],
            colWidths=[PAGE_W])
        head_tbl.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1, -1), accent),
            ('TOPPADDING',   (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 5),
            ('LEFTPADDING',  (0, 0), (-1, -1), 6),
        ]))
        story.append(head_tbl)
        story.append(Spacer(1, 2*mm))

        tbl_data = [HDR_ROW] + [
            [code, _name_cell(code), val, dflt, mn, mx, prv]
            for code, val, dflt, mn, mx, prv in rows
        ]
        tbl = Table(tbl_data, colWidths=COL_W, repeatRows=1)
        tbl.setStyle(cat_table_style(len(rows)))
        story.append(tbl)

        # Pre-set tracker for the NEXT category before this PageBreak
        if idx + 1 < len(cats):
            story.append(_SetCat(_cat_label(cats[idx + 1])))
        story.append(PageBreak())

    try:
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=22*mm, bottomMargin=20*mm,
        )
        doc.build(story, onFirstPage=draw_page, onLaterPages=draw_page)
        if log_fn:
            log_fn('log_drive_pdf_saved', output_path)
        return True
    except Exception as exc:
        if log_fn:
            log_fn('log_error_generic', f'Drive PDF: {exc}')
        return False
