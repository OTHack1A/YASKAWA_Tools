import os
import re

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

TOOL_FILE   = "TOOL.CND"
UFRAME_FILE = "UFRAME.CND"

# ── Low-level helpers ─────────────────────────────────────────────────────────

def _safe_float(s):
    """Parse a value as float (accepting comma decimals), or 0.0 on failure."""
    try:
        return float(str(s).strip().replace(',', '.'))
    except (ValueError, TypeError):
        return 0.0

def _safe_int(s):
    """Parse a value as int, or 0 on failure."""
    try:
        return int(str(s).strip())
    except (ValueError, TypeError):
        return 0

def _fmt(v):
    """Format a number with 3 decimals."""
    return f"{v:.3f}"

def _is_nonzero(tool):
    """True if any of the tool's x/y/z/rx/ry/rz components is non-zero."""
    return any(abs(tool[k]) > 1e-9 for k in ('x', 'y', 'z', 'rx', 'ry', 'rz'))

# ── TOOL.CND parsers ──────────────────────────────────────────────────────────

def _parse_tool_robot(lines):
    """Robot-native TOOL.CND: //TOOL NO :n, //USER NAME :name, //X(mm) :v …"""
    tools, cur = [], None
    for line in lines:
        stripped = line.lstrip('/')
        if not stripped or ':' not in stripped:
            continue
        ci  = stripped.index(':')
        key = stripped[:ci].strip().upper()
        val = stripped[ci + 1:].strip()

        if re.match(r'TOOL\s*NO', key):
            if cur is not None:
                tools.append(cur)
            cur = {'num': _safe_int(val), 'name': '',
                   'x': 0., 'y': 0., 'z': 0., 'rx': 0., 'ry': 0., 'rz': 0.}
        elif cur is not None:
            if re.match(r'USER\s*NAME|^NAME$', key):
                cur['name'] = val
            elif re.match(r'X\s*\(MM\)', key):
                cur['x'] = _safe_float(val)
            elif re.match(r'Y\s*\(MM\)', key):
                cur['y'] = _safe_float(val)
            elif re.match(r'Z\s*\(MM\)', key):
                cur['z'] = _safe_float(val)
            elif re.match(r'RX\s*\(DEG\)', key):
                cur['rx'] = _safe_float(val)
            elif re.match(r'RY\s*\(DEG\)', key):
                cur['ry'] = _safe_float(val)
            elif re.match(r'RZ\s*\(DEG\)', key):
                cur['rz'] = _safe_float(val)
    if cur is not None:
        tools.append(cur)
    return tools

def _parse_tool_our(lines):
    """Our export format: //TOOL n / ///NAME … / x,y,z,rx,ry,rz / …"""
    tools = []
    i = 0
    while i < len(lines):
        m = re.match(r'^//TOOL\s+(\d+)', lines[i])
        if m:
            num  = int(m.group(1))
            name = ''
            x = y = z = rx = ry = rz = 0.0
            j = i + 1
            if j < len(lines):
                nm = re.match(r'^///NAME\s+(.*)', lines[j])
                if nm:
                    name = nm.group(1).strip()
                    j += 1
            if j < len(lines):
                parts = lines[j].split(',')
                if len(parts) >= 6:
                    x, y, z  = (_safe_float(parts[0]),
                                 _safe_float(parts[1]),
                                 _safe_float(parts[2]))
                    rx, ry, rz = (_safe_float(parts[3]),
                                  _safe_float(parts[4]),
                                  _safe_float(parts[5]))
                    j += 1
            tools.append({'num': num, 'name': name,
                          'x': x, 'y': y, 'z': z,
                          'rx': rx, 'ry': ry, 'rz': rz})
            i = j
        else:
            i += 1
    return tools

def parse_tool_cnd(filepath):
    """Parse TOOL.CND and return the list of tool-frame entries."""
    with open(filepath, 'r', encoding='latin-1', errors='replace') as fh:
        lines = [ln.rstrip() for ln in fh]
    is_robot = any(re.match(r'^//TOOL\s*NO', ln, re.IGNORECASE) for ln in lines)
    return _parse_tool_robot(lines) if is_robot else _parse_tool_our(lines)

# ── UFRAME.CND parser ─────────────────────────────────────────────────────────

def _try_csv6(line):
    """Return (x,y,z,rx,ry,rz) if line has ≥6 comma-separated floats, else None."""
    parts = line.split(',')
    if len(parts) < 6:
        return None
    floats = []
    for p in parts[:6]:
        try:
            floats.append(float(p.strip()))
        except ValueError:
            return None
    return tuple(floats) if len(floats) == 6 else None


def _parse_uframe_compact(lines):
    """
    Format:
        //UFRAME N
        ///NAME [name]          ← name may be empty
        ///BUSER                ← section with base-frame coordinates (used)
        x,y,z,rx,ry,rz
        [flag line]
        ///RUSER / ///TUSER … ← other sections (skipped)

    Only the CSV line immediately after ///BUSER is taken as coordinates.
    Falls back to the first CSV-6 block in the frame if ///BUSER is absent.
    """
    frames = []
    i = 0
    while i < len(lines):
        m = re.match(r'^//UFRAME\s+(\d+)', lines[i])
        if m:
            num  = int(m.group(1))
            name = ''
            x = y = z = rx = ry = rz = 0.0
            j = i + 1

            # ///NAME (may be empty)
            if j < len(lines) and re.match(r'^///NAME', lines[j], re.IGNORECASE):
                nm = re.match(r'^///NAME\s*(.*)', lines[j], re.IGNORECASE)
                name = nm.group(1).strip() if nm else ''
                j += 1

            # Scan lines until next //UFRAME block
            buser_coords = None
            first_csv    = None   # fallback if no ///BUSER marker
            in_buser     = False

            while j < len(lines) and not re.match(r'^//UFRAME\s+\d+', lines[j]):
                raw = lines[j].rstrip()

                # ////BUSER (any number of leading slashes) — inline coordinates
                if re.match(r'^/+BUSER', raw, re.IGNORECASE):
                    rest = re.sub(r'^/+BUSER\s*', '', raw, flags=re.IGNORECASE).strip()
                    csv  = _try_csv6(rest) if rest else None
                    if csv is not None and buser_coords is None:
                        buser_coords = csv
                    else:
                        in_buser = True   # coords expected on next line (rare)
                    j += 1
                    continue

                # Any other //-prefixed section marker (///X, ////X …) — skip, reset flag
                if re.match(r'^//', raw):
                    in_buser = False
                    j += 1
                    continue

                # Plain CSV line (no leading slashes) — use only if inside BUSER block
                csv = _try_csv6(raw)
                if csv is not None:
                    if in_buser and buser_coords is None:
                        buser_coords = csv
                        in_buser     = False
                    if first_csv is None:
                        first_csv = csv

                j += 1

            # Use BUSER coords preferentially, else first CSV found
            coords = buser_coords if buser_coords is not None else first_csv
            if coords:
                x, y, z, rx, ry, rz = coords

            frames.append({'num': num, 'name': name,
                           'x': x, 'y': y, 'z': z,
                           'rx': rx, 'ry': ry, 'rz': rz})
            i = j
        else:
            i += 1
    return frames


def _parse_uframe_kv(lines):
    """Fallback: key-value pairs, frame boundary on NO / UFRAME NO / USER FRAME NO."""
    frames, cur = [], None
    for line in lines:
        stripped = line.lstrip('/')
        if not stripped or ':' not in stripped:
            continue
        ci  = stripped.index(':')
        key = stripped[:ci].strip().upper()
        val = stripped[ci + 1:].strip()
        if re.match(r'^(?:USER\s+FRAME\s+|UFRAME\s+)?NO$', key):
            if cur is not None:
                frames.append(cur)
            cur = {'num': _safe_int(val), 'name': '',
                   'x': 0., 'y': 0., 'z': 0., 'rx': 0., 'ry': 0., 'rz': 0.}
        elif cur is not None:
            if key == 'NAME':
                cur['name'] = val
            elif re.match(r'X\s*\(MM\)', key):
                cur['x'] = _safe_float(val)
            elif re.match(r'Y\s*\(MM\)', key):
                cur['y'] = _safe_float(val)
            elif re.match(r'Z\s*\(MM\)', key):
                cur['z'] = _safe_float(val)
            elif re.match(r'RX\s*\(DEG\)', key):
                cur['rx'] = _safe_float(val)
            elif re.match(r'RY\s*\(DEG\)', key):
                cur['ry'] = _safe_float(val)
            elif re.match(r'RZ\s*\(DEG\)', key):
                cur['rz'] = _safe_float(val)
    if cur is not None:
        frames.append(cur)
    return frames


def parse_uframe_cnd(filepath):
    """
    Parse UFRAME.CND.  Auto-detects format:
      - Compact (Yaskawa native backup): //UFRAME N / ///NAME … / csv coords
      - Key-value fallback: ///KEY :value groups
    """
    with open(filepath, 'r', encoding='latin-1', errors='replace') as fh:
        lines = [ln.rstrip() for ln in fh]
    if any(re.match(r'^//UFRAME\s+\d+', ln) for ln in lines):
        return _parse_uframe_compact(lines)
    return _parse_uframe_kv(lines)

# ── PDF helpers ───────────────────────────────────────────────────────────────

_HDR_COLOR  = colors.HexColor('#a85c42')
_ALT_COLOR  = colors.HexColor('#FDF0E8')

def _table_style(n_rows, has_name_col=True, fn='Helvetica', fn_b='Helvetica-Bold'):
    """Build TableStyle for data tables."""
    style = [
        ('BACKGROUND',    (0, 0), (-1, 0),  _HDR_COLOR),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.black),
        ('FONTNAME',      (0, 0), (-1, 0),  fn_b),
        ('FONTSIZE',      (0, 0), (-1, 0),  7),
        ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
        ('FONTNAME',      (0, 1), (-1, -1), fn),
        ('FONTSIZE',      (0, 1), (-1, -1), 7),
        ('ALIGN',         (0, 1), (0, -1),  'CENTER'),
        ('ALIGN',         (2, 1), (-1, -1), 'RIGHT'),
        ('GRID',          (0, 0), (-1, -1), 0.25, colors.HexColor('#aaaaaa')),
        ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
    ]
    if has_name_col:
        style.append(('ALIGN', (1, 1), (1, -1), 'LEFT'))
    for i in range(1, n_rows):
        bg = colors.white if i % 2 else _ALT_COLOR
        style.append(('BACKGROUND', (0, i), (-1, i), bg))
    return TableStyle(style)

def _make_draw_page_num(page_offset=0):
    """Return a ReportLab page callback that draws the header and page-number footer."""
    def _draw(canvas, doc):
        """ReportLab page callback: draw the shared header and the page-number footer."""
        from docs.pdf_header import draw_page_header
        draw_page_header(canvas, doc)
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.HexColor('#555555'))
        canvas.drawCentredString(A4[0] / 2, 8 * mm, str(canvas.getPageNumber() + page_offset))
        canvas.restoreState()
    return _draw

# ── Public entry point ────────────────────────────────────────────────────────

def generate_pdf(folder_path, output_path, lang="IT", log_fn=None, page_offset=0):
    """
    Parse TOOL.CND and UFRAME.CND from *folder_path*, write PDF to *output_path*.
    Returns (n_configured_tools, n_frames).
    """
    def _log(key, *args):
        """Forward a log message to the caller's log callback, ignoring any error."""
        if log_fn:
            try:
                log_fn(key, *args)
            except Exception:
                pass

    from translations import TRANSLATIONS
    tr = TRANSLATIONS.get(lang, TRANSLATIONS.get("IT", {}))
    from docs.utils import pdf_font
    fn   = pdf_font(lang)
    fn_b = pdf_font(lang, bold=True)

    # ── Parse ─────────────────────────────────────────────────────────────────
    all_tools, all_frames = [], []

    tool_path = os.path.join(folder_path, TOOL_FILE)
    if os.path.isfile(tool_path):
        try:
            all_tools = parse_tool_cnd(tool_path)
        except Exception as exc:
            _log("log_error_generic", str(exc))
    else:
        _log("log_file_not_found", TOOL_FILE)

    uf_path = os.path.join(folder_path, UFRAME_FILE)
    if os.path.isfile(uf_path):
        try:
            all_frames = parse_uframe_cnd(uf_path)
        except Exception as exc:
            _log("log_error_generic", str(exc))
    else:
        _log("log_file_not_found", UFRAME_FILE)

    cfg_tools = [t for t in all_tools if _is_nonzero(t)]

    n_tools  = len(cfg_tools)
    n_frames = len(all_frames)
    _log("log_uf_tools_read_tools",  n_tools)
    _log("log_uf_tools_read_frames", n_frames)

    # ── Build PDF ─────────────────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=22 * mm,  bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()
    h_style = ParagraphStyle(
        'SectionHdr',
        parent=styles['Normal'],
        fontSize=9, fontName=fn_b,
        textColor=colors.HexColor('#a85c42'),
        spaceBefore=4, spaceAfter=3,
    )

    story = []

    # ── Tools section ──────────────────────────────────────────────────────────
    story.append(Paragraph(
        tr.get("pdf_uf_title_tools", "Tool TCP — YASKAWA YRC1000"), h_style))

    name_hdr = tr.get("tool_col_name", "Nome")
    t_hdrs   = ["#", name_hdr,
                "X [mm]", "Y [mm]", "Z [mm]",
                "Rx [°]", "Ry [°]", "Rz [°]"]
    t_rows   = [t_hdrs]
    for tool in cfg_tools:
        t_rows.append([
            str(tool['num']),
            tool['name'] or '—',
            _fmt(tool['x']), _fmt(tool['y']), _fmt(tool['z']),
            _fmt(tool['rx']), _fmt(tool['ry']), _fmt(tool['rz']),
        ])

    if len(t_rows) == 1:
        story.append(Paragraph("—", styles['Normal']))
    else:
        # Inner page width ≈ 510pt (A4 595 - 2×15mm margins)
        cw_t = [22, 90, 66, 66, 66, 66, 66, 66]
        tbl  = Table(t_rows, colWidths=cw_t, repeatRows=1)
        tbl.setStyle(_table_style(len(t_rows), has_name_col=True, fn=fn, fn_b=fn_b))
        story.append(tbl)

    story.append(Spacer(1, 5 * mm))

    # ── User Frame section ─────────────────────────────────────────────────────
    story.append(Paragraph(
        tr.get("pdf_uf_title_frames", "User Frame — YASKAWA YRC1000"), h_style))

    uf_name_hdr = tr.get("tool_col_name", "Nome")
    uf_hdrs = ["UF#", uf_name_hdr,
               "X [mm]", "Y [mm]", "Z [mm]",
               "Rx [°]", "Ry [°]", "Rz [°]"]
    uf_rows = [uf_hdrs]
    for fr in all_frames:
        uf_rows.append([
            str(fr['num']),
            fr.get('name', '') or '—',
            _fmt(fr['x']), _fmt(fr['y']), _fmt(fr['z']),
            _fmt(fr['rx']), _fmt(fr['ry']), _fmt(fr['rz']),
        ])

    if len(uf_rows) == 1:
        story.append(Paragraph("—", styles['Normal']))
    else:
        cw_uf = [22, 80, 68, 68, 68, 68, 68, 68]
        tbl_uf = Table(uf_rows, colWidths=cw_uf, repeatRows=1)
        tbl_uf.setStyle(_table_style(len(uf_rows), has_name_col=True, fn=fn, fn_b=fn_b))
        story.append(tbl_uf)

    _page_fn = _make_draw_page_num(page_offset)
    doc.build(story, onFirstPage=_page_fn, onLaterPages=_page_fn)

    return n_tools, n_frames


def generate_uf_excel(folder_path, output_path, lang="IT", log_fn=None):
    """Generate Excel with Tool + UFrame sheets. Returns (n_tools, n_frames)."""
    def _log(key, *args):
        """Forward a log message to the caller's log callback, ignoring any error."""
        if log_fn:
            try:
                log_fn(key, *args)
            except Exception:
                pass

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError as exc:
        _log("log_error_generic", f"openpyxl: {exc}")
        return 0, 0

    from translations import TRANSLATIONS
    tr = TRANSLATIONS.get(lang, TRANSLATIONS.get("IT", {}))

    all_tools, all_frames = [], []
    tool_path = os.path.join(folder_path, TOOL_FILE)
    uf_path   = os.path.join(folder_path, UFRAME_FILE)
    if os.path.isfile(tool_path):
        try:
            all_tools = parse_tool_cnd(tool_path)
        except Exception as exc:
            _log("log_error_generic", str(exc))
    if os.path.isfile(uf_path):
        try:
            all_frames = parse_uframe_cnd(uf_path)
        except Exception as exc:
            _log("log_error_generic", str(exc))

    cfg_tools = [t for t in all_tools if _is_nonzero(t)]

    wb = openpyxl.Workbook()

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1B3A6B")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="AAAAAA")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    from docs.utils import excel_safe
    def _make_sheet(ws, headers, rows):
        """Append a styled header row and the data rows to an Excel worksheet."""
        ws.append(headers)
        for cell in ws[1]:
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = hdr_align
            cell.border = border
        for row in rows:
            ws.append([excel_safe(c) for c in row])
            for cell in ws[ws.max_row]:
                cell.border = border
        for col in ws.columns:
            w = max(len(str(cell.value or "")) for cell in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(w, 30)

    name_hdr = tr.get("tool_col_name", "Nome")

    # Tools sheet
    ws_t = wb.active
    ws_t.title = "Tool"
    t_hdrs = ["#", name_hdr, "X [mm]", "Y [mm]", "Z [mm]", "Rx [°]", "Ry [°]", "Rz [°]"]
    t_rows = [
        [t['num'], t['name'] or '', _fmt(t['x']), _fmt(t['y']), _fmt(t['z']),
         _fmt(t['rx']), _fmt(t['ry']), _fmt(t['rz'])]
        for t in cfg_tools
    ]
    _make_sheet(ws_t, t_hdrs, t_rows)

    # UFrame sheet
    ws_uf = wb.create_sheet("User Frame")
    uf_hdrs = ["UF#", name_hdr, "X [mm]", "Y [mm]", "Z [mm]", "Rx [°]", "Ry [°]", "Rz [°]"]
    uf_rows = [
        [f['num'], f.get('name', '') or '', _fmt(f['x']), _fmt(f['y']), _fmt(f['z']),
         _fmt(f['rx']), _fmt(f['ry']), _fmt(f['rz'])]
        for f in all_frames
    ]
    _make_sheet(ws_uf, uf_hdrs, uf_rows)

    try:
        wb.save(output_path)
        _log("log_uf_excel_saved", output_path)
        return len(cfg_tools), len(all_frames)
    except Exception as exc:
        _log("log_error_generic", str(exc))
        return 0, 0


def write_uframe_cnd(filepath, frames_by_num, src_path=None):
    """Write NAME and BUSER coords back to UFRAME.CND.

    frames_by_num: {num: {'name': str, 'x': float, 'y', 'z', 'rx', 'ry', 'rz'}}.
    The original file is used as the template; ``src_path`` is where it is read
    from and ``filepath`` is where the result is written. When ``src_path`` is
    omitted it defaults to ``filepath`` (in-place overwrite). This separation
    lets the user export to a brand-new destination that does not exist yet
    without hitting "No such file or directory".
    Returns (ok, error_msg)."""
    import re as _re
    src = src_path or filepath
    try:
        with open(src, 'r', encoding='latin-1', errors='replace') as f:
            lines = f.readlines()
    except OSError as exc:
        return False, str(exc)

    def _eol(ln):
        """Return the line ending (CRLF or LF) matching the given line."""
        return '\r\n' if '\r' in ln else '\n'

    out = []
    i   = 0
    while i < len(lines):
        raw = lines[i]
        s   = raw.rstrip()
        m   = _re.match(r'^//UFRAME\s+(\d+)', s)
        if m:
            num = int(m.group(1))
            out.append(raw)
            i += 1
            if num not in frames_by_num:
                # copy block as-is
                while i < len(lines):
                    if _re.match(r'^//UFRAME\s+\d+', lines[i].rstrip()):
                        break
                    out.append(lines[i])
                    i += 1
            else:
                upd = frames_by_num[num]
                coords = (f"{upd['x']:.3f},{upd['y']:.3f},{upd['z']:.3f},"
                          f"{upd['rx']:.4f},{upd['ry']:.4f},{upd['rz']:.4f}")
                buser_written = False
                while i < len(lines):
                    ln = lines[i]
                    s2 = ln.rstrip()
                    if _re.match(r'^//UFRAME\s+\d+', s2):
                        break
                    if _re.match(r'^///NAME', s2, _re.IGNORECASE):
                        out.append(f"///NAME {upd['name']}{_eol(ln)}")
                        i += 1
                        continue
                    bm = _re.match(r'^(/+BUSER)', s2, _re.IGNORECASE)
                    if bm:
                        prefix = bm.group(1)
                        out.append(f"{prefix} {coords}{_eol(ln)}")
                        buser_written = True
                        i += 1
                        # skip next line if it contains only coords (split format)
                        if i < len(lines):
                            nxt = lines[i].rstrip()
                            if nxt and not nxt.startswith('/') and ',' in nxt:
                                try:
                                    [float(v) for v in nxt.split(',')[:6]]
                                    i += 1  # skip old coord line
                                except ValueError:
                                    pass
                        continue
                    out.append(ln)
                    i += 1
                if not buser_written:
                    # append BUSER if not present in original
                    out.append(f"////BUSER {coords}\r\n")
        else:
            out.append(raw)
            i += 1

    try:
        dest_dir = os.path.dirname(os.path.abspath(filepath))
        if dest_dir and not os.path.isdir(dest_dir):
            os.makedirs(dest_dir, exist_ok=True)
        with open(filepath, 'w', encoding='latin-1', newline='') as f:
            f.writelines(out)
        return True, None
    except OSError as exc:
        return False, str(exc)
