import os
import re

VAR_FILE = "VAR.DAT"
_VAR_SECTIONS = ["B", "I", "D", "R", "S", "P"]


def _open_r(filepath):
    return open(filepath, "r", encoding="latin-1", errors="replace")


# ── VAR.DAT parser ────────────────────────────────────────────────────────────

def parse_var_dat(filepath):
    """Parse VAR.DAT.
    Returns {type: [value, ...]} for B/I/D (int), R (float), S/P (str).
    Raises RuntimeError on file open failure."""
    sections = {k: [] for k in _VAR_SECTIONS}
    current = None

    try:
        with _open_r(filepath) as f:
            lines = f.readlines()
    except OSError as exc:
        raise RuntimeError(f"Cannot open VAR.DAT: {exc}") from exc

    for line in lines:
        s = line.rstrip("\r\n")
        if s.startswith("///SHARE") or s.startswith("///PFNUM"):
            continue
        if s.startswith("///"):
            tok = s[3:].strip().split()[0] if s[3:].strip() else ""
            current = tok if tok in sections else None
            continue
        if s.startswith("/") or current is None:
            continue

        if current in ("B", "I", "D"):
            for p in s.split(","):
                try:
                    sections[current].append(int(p.strip()))
                except ValueError:
                    sections[current].append(0)
        elif current == "R":
            for p in s.split(","):
                try:
                    sections[current].append(float(p.strip()))
                except ValueError:
                    sections[current].append(0.0)
        elif current in ("S", "P"):
            sections[current].append(s)

    return sections


# ── Value formatters ──────────────────────────────────────────────────────────

def _format_position(raw_str):
    """Format a VAR.DAT position string into a readable representation."""
    val = raw_str.strip()
    if not val or '"UNUSED"' in val:
        return ""
    m = re.match(r'^"([^"]+)"(.*)', val)
    if not m:
        return val
    ptype = m.group(1)
    rest = m.group(2).strip(",")
    parts = rest.split(",")
    try:
        coords = parts[6:12]
        if len(coords) < 6:
            return f"{ptype}: {rest}"
        if ptype == "RECTAN":
            x, y, z, rx, ry, rz = (float(c) for c in coords)
            return (f"X={x:.3f} Y={y:.3f} Z={z:.3f} "
                    f"Rx={rx:.4f} Ry={ry:.4f} Rz={rz:.4f}")
        if ptype == "PULSE":
            j1, j2, j3, j4, j5, j6 = (int(float(c)) for c in coords)
            return f"J1={j1} J2={j2} J3={j3} J4={j4} J5={j5} J6={j6}"
        return f"{ptype}: {','.join(coords)}"
    except (ValueError, IndexError):
        return val


def _format_value(var_type, raw_val):
    """Convert a raw parsed value to a display/Excel-ready value."""
    if raw_val is None:
        return ""
    if var_type in ("B", "I", "D"):
        return raw_val
    if var_type == "R":
        if raw_val == 0.0:
            return 0
        if raw_val == int(raw_val):
            return int(raw_val)
        return round(raw_val, 4)
    if var_type == "S":
        return raw_val.strip()
    if var_type == "P":
        return _format_position(raw_val)
    return str(raw_val)


# ── In-memory loader ──────────────────────────────────────────────────────────

def load_backup_data(template_path, robot_folder, log_fn=None):
    """Read template Excel + VAR.DAT.

    Returns (tab_data, sections) where tab_data = {tab: [(var_id, name, value), ...]}
    (only named variables) and sections = raw VAR.DAT dict.
    Returns (None, None) on failure."""
    import openpyxl
    from docs.names import _VAR_TABS

    var_dat_path = os.path.join(robot_folder, VAR_FILE)
    if not os.path.isfile(var_dat_path):
        if log_fn:
            log_fn("log_backup_no_var_dat", robot_folder)
        return None, None

    try:
        sections = parse_var_dat(var_dat_path)
        total = sum(len(v) for v in sections.values())
        if log_fn:
            log_fn("log_backup_var_read", total)
    except Exception as exc:
        if log_fn:
            log_fn("log_error_generic", f"VAR.DAT: {exc}")
        return None, None

    try:
        wb_tmpl = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    except Exception as exc:
        if log_fn:
            log_fn("log_error_generic", f"Template: {exc}")
        return None, None

    tab_data = {}

    for tab in _VAR_TABS:
        if tab not in wb_tmpl.sheetnames:
            continue

        try:
            ws_tmpl = wb_tmpl[tab]
            tmpl_rows = list(ws_tmpl.iter_rows(values_only=True))
        except Exception as exc:
            if log_fn:
                log_fn("log_error_generic", f"Tab {tab}: {exc}")
            continue

        if not tmpl_rows:
            continue

        tab_vals = sections.get(tab, [])
        rows = []

        for tmpl_row in tmpl_rows[1:]:
            if not tmpl_row:
                continue
            var_id = tmpl_row[0] if len(tmpl_row) > 0 else None
            if not var_id:
                continue

            var_name_raw = tmpl_row[1] if len(tmpl_row) > 1 else None
            var_name = str(var_name_raw).strip() if var_name_raw else ""

            idx = -1
            m = re.search(r"(\d+)$", str(var_id))
            if m:
                idx = int(m.group(1))

            raw = tab_vals[idx] if 0 <= idx < len(tab_vals) else None
            cell_val = _format_value(tab, raw)

            rows.append((str(var_id), var_name, cell_val))

        if rows:
            tab_data[tab] = rows

    try:
        wb_tmpl.close()
    except Exception:
        pass

    total_vars = sum(len(v) for v in tab_data.values())
    if log_fn:
        log_fn("log_backup_data_loaded", total_vars)

    return tab_data, sections


# ── Excel writer ──────────────────────────────────────────────────────────────

def generate_backup_excel(tab_data, output_path,
                          col_id_label="ID", col_name_label="Nome",
                          col_value_label="Valore", log_fn=None):
    """Save in-memory tab_data as formatted Excel. Returns True/False."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    if not tab_data:
        if log_fn:
            log_fn("log_backup_no_data")
        return False

    HDR_FILL     = PatternFill("solid", fgColor="D97757")
    HDR_FNT  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    CENTER   = Alignment(horizontal="center", vertical="center")
    GRAY     = PatternFill("solid", fgColor="EFEFEF")
    ID_FONT  = Font(name="Courier New", size=9, color="003366")
    ID_ALIGN = Alignment(horizontal="center")
    R_ALIGN  = Alignment(horizontal="right")

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for tab, rows in tab_data.items():
        ws = wb_out.create_sheet(title=tab)

        for ci, (label, width) in enumerate(
                [(col_id_label, 12), (col_name_label, 30), (col_value_label, 22)], start=1):
            c = ws.cell(row=1, column=ci, value=label)
            c.fill = HDR_FILL
            c.font = HDR_FNT
            c.alignment = CENTER
            ws.column_dimensions[c.column_letter].width = width
        ws.row_dimensions[1].height = 18
        ws.freeze_panes = "A2"

        from docs.utils import excel_safe
        for out_row, (var_id, name, val) in enumerate(rows, start=2):
            c_id = ws.cell(row=out_row, column=1, value=excel_safe(var_id))
            c_id.fill = GRAY
            c_id.font = ID_FONT
            c_id.alignment = ID_ALIGN

            ws.cell(row=out_row, column=2, value=excel_safe(name))

            if val not in (None, ""):
                c_val = ws.cell(row=out_row, column=3, value=excel_safe(val))
                if tab not in ("S", "P"):
                    c_val.alignment = R_ALIGN

        if log_fn:
            log_fn("log_backup_tab", tab, len(rows))

    try:
        wb_out.save(output_path)
        if log_fn:
            log_fn("log_backup_done", output_path)
        return True
    except Exception as exc:
        if log_fn:
            log_fn("log_error_generic", f"Salvataggio backup: {exc}")
        return False


# ── Template-based Excel writer ───────────────────────────────────────────────

def generate_backup_excel_from_template(template_path, sections, output_path,
                                        col_value_label="Valore", log_fn=None):
    """Copy template Excel and append a value column from VAR.DAT sections.

    Preserves template formatting; adds one column at the end of each VAR sheet.
    Returns True on success, False on any error."""
    import shutil
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    HDR_FILL    = PatternFill("solid", fgColor="D97757")
    HDR_FNT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    CENTER  = Alignment(horizontal="center", vertical="center")
    R_ALIGN = Alignment(horizontal="right")
    NUMERIC = {"B", "I", "D", "R"}

    try:
        shutil.copy2(template_path, output_path)
    except Exception as exc:
        if log_fn:
            log_fn("log_error_generic", f"Copia template: {exc}")
        return False

    try:
        wb = openpyxl.load_workbook(output_path)
    except Exception as exc:
        if log_fn:
            log_fn("log_error_generic", f"Apertura workbook: {exc}")
        return False

    for sheet_name in _VAR_SECTIONS:
        if sheet_name not in wb.sheetnames:
            continue
        tab_vals = sections.get(sheet_name, [])
        ws = wb[sheet_name]

        hdr_col = ws.max_column + 1
        col_letter = get_column_letter(hdr_col)

        hdr_cell = ws.cell(row=1, column=hdr_col, value=col_value_label)
        hdr_cell.fill = HDR_FILL
        hdr_cell.font = HDR_FNT
        hdr_cell.alignment = CENTER
        ws.column_dimensions[col_letter].width = 22

        written = 0
        for row_idx in range(2, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1)
            var_id_str = str(cell_a.value) if cell_a.value is not None else ""
            if not var_id_str:
                continue
            m = re.search(r"(\d+)$", var_id_str)
            if not m:
                continue
            idx = int(m.group(1))
            raw = tab_vals[idx] if 0 <= idx < len(tab_vals) else None
            cell_val = _format_value(sheet_name, raw)
            if cell_val not in (None, ""):
                from docs.utils import excel_safe
                val_cell = ws.cell(row=row_idx, column=hdr_col)
                val_cell.value = excel_safe(cell_val)
                if sheet_name in NUMERIC:
                    val_cell.alignment = R_ALIGN
                written += 1

        if log_fn:
            log_fn("log_backup_tab", sheet_name, written)

    try:
        wb.save(output_path)
        if log_fn:
            log_fn("log_backup_done", output_path)
        return True
    except Exception as exc:
        if log_fn:
            log_fn("log_error_generic", f"Salvataggio: {exc}")
        return False


# ── PDF exporter ──────────────────────────────────────────────────────────────

def generate_backup_pdf(tab_data, output_path, folder_name="",
                        col_id_label="ID", col_name_label="Nome",
                        col_value_label="Valore", log_fn=None, lang="IT"):
    """Generate A4 PDF with 3 record-groups per row, new page per variable type.
    Returns True/False."""
    from docs.utils import pdf_font, xml_escape
    f_reg  = pdf_font(lang, bold=False)
    f_bold = pdf_font(lang, bold=True)
    try:
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        PageBreak, Paragraph, Spacer)
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor, white, black
    except ImportError as exc:
        if log_fn:
            log_fn("log_error_generic", f"reportlab: {exc}")
        return False

    if not tab_data:
        if log_fn:
            log_fn("log_backup_no_pdf_data")
        return False

    # 3 groups of [ID=13mm, Name=30mm, Value=17mm] = 180mm content width
    CW = [13*mm, 30*mm, 17*mm] * 3

    accent = HexColor("#D97757")
    row_gray  = HexColor("#EFEFEF")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "bkp_title",
        parent=styles["Normal"],
        fontSize=11, fontName=f_bold,
        spaceAfter=3*mm,
    )

    def make_header_row():
        return [col_id_label, col_name_label, col_value_label] * 3

    def make_table_style(n_data_rows):
        ts = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), accent),
            ("TEXTCOLOR", (0, 0), (-1, 0), black),
            ("FONTNAME", (0, 0), (-1, 0), f_bold),
            ("FONTSIZE", (0, 0), (-1, 0), 7),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            ("ROWHEIGHT", (0, 0), (0, 0), 12),
            ("FONTNAME", (0, 1), (-1, -1), f_reg),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
            ("ROWHEIGHT", (0, 1), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.25, HexColor("#CCCCCC")),
            # Right-align value columns
            ("ALIGN", (2, 1), (2, -1), "RIGHT"),
            ("ALIGN", (5, 1), (5, -1), "RIGHT"),
            ("ALIGN", (8, 1), (8, -1), "RIGHT"),
        ])
        for r in range(1, n_data_rows + 1):
            if r % 2 == 0:
                ts.add("BACKGROUND", (0, r), (-1, r), row_gray)
        return ts

    story = []
    tabs = list(tab_data.keys())

    for ti, tab in enumerate(tabs):
        rows_data = tab_data[tab]

        if ti > 0:
            story.append(PageBreak())

        title_text = f"{tab}"
        if folder_name:
            title_text += f" — {folder_name}"
        story.append(Paragraph(xml_escape(title_text), title_style))

        # Pack entries 3 per table row
        padded = list(rows_data)
        while len(padded) % 3 != 0:
            padded.append(("", "", ""))

        table_rows = [make_header_row()]
        for i in range(0, len(padded), 3):
            r0, r1, r2 = padded[i], padded[i+1], padded[i+2]
            table_rows.append([
                r0[0], r0[1], "" if r0[2] in (None, "") else str(r0[2]),
                r1[0], r1[1], "" if r1[2] in (None, "") else str(r1[2]),
                r2[0], r2[1], "" if r2[2] in (None, "") else str(r2[2]),
            ])

        tbl = Table(table_rows, colWidths=CW, repeatRows=1)
        tbl.setStyle(make_table_style(len(table_rows) - 1))
        story.append(tbl)

    def _draw_page_number(canvas, doc):
        from docs.pdf_header import draw_page_header
        draw_page_header(canvas, doc)
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(HexColor("#888888"))
        page_text = f"- {canvas.getPageNumber()} -"
        canvas.drawCentredString(A4[0] / 2, 8 * mm, page_text)
        canvas.restoreState()

    try:
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            leftMargin=10*mm, rightMargin=10*mm,
            topMargin=22*mm, bottomMargin=20*mm,
        )
        doc.build(story,
                  onFirstPage=_draw_page_number,
                  onLaterPages=_draw_page_number)
        if log_fn:
            log_fn("log_backup_pdf_saved", output_path)
        return True
    except Exception as exc:
        if log_fn:
            log_fn("log_error_generic", f"PDF backup: {exc}")
        return False
