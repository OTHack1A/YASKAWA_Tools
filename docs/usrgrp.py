import os
import re

GPIN_FILE = "USRGRPIN.DAT"
GPOT_FILE = "USRGRPOT.DAT"


def _open_r(filepath):
    return open(filepath, "r", encoding="latin-1", errors="replace")


def _parse_file(filepath, tag):
    """Parse USRGRPIN.DAT or USRGRPOT.DAT.
    Returns list of dicts: {num, name, gpin, bits, val3}.
    """
    groups = []
    try:
        with _open_r(filepath) as f:
            lines = [ln.rstrip("\r\n") for ln in f]
    except OSError:
        return groups

    pattern = re.compile(rf'^//{re.escape(tag)}\s+(\d+)', re.IGNORECASE)
    i = 0
    while i < len(lines):
        m = pattern.match(lines[i])
        if m:
            num = int(m.group(1))
            name = ""
            gpin = 0
            bits = 0
            val3 = 0
            j = i + 1
            if j < len(lines):
                nm = re.match(r'^///NAME\s*(.*)', lines[j], re.IGNORECASE)
                if nm:
                    name = nm.group(1).strip()
                    j += 1
            if j < len(lines) and not lines[j].startswith('/'):
                parts = lines[j].split(',')
                try:
                    gpin = int(parts[0].strip()) if parts else 0
                    bits = int(parts[1].strip()) if len(parts) > 1 else 0
                    val3 = int(parts[2].strip()) if len(parts) > 2 else 0
                except (ValueError, IndexError):
                    pass
                j += 1
            groups.append({'num': num, 'name': name,
                           'gpin': gpin, 'bits': bits, 'val3': val3})
            i = j
        else:
            i += 1
    return groups


def parse_gpin(folder):
    """Parse USRGRPIN.DAT. Returns list of group dicts."""
    return _parse_file(os.path.join(folder, GPIN_FILE), "USRGRPIN")


def parse_gpot(folder):
    """Parse USRGRPOT.DAT. Returns list of group dicts."""
    return _parse_file(os.path.join(folder, GPOT_FILE), "USRGRPOT")


def is_active(g):
    """A group is active if it has a non-empty name AND data != 0,0,0."""
    return bool(g.get('name', '').strip()) and (
        g.get('gpin', 0) != 0 or g.get('bits', 0) != 0)


# ── PDF generation ────────────────────────────────────────────────────────────

def generate_pdf(gpin_groups, gpot_groups, output_path, lang="IT", progress_fn=None, page_offset=0):
    """Generate PDF report for User Groups. Returns (n_gpin_active, n_gpot_active)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from docs.utils import pdf_font, make_footer
    from translations import TRANSLATIONS

    t = TRANSLATIONS.get(lang, TRANSLATIONS["IT"])
    font      = pdf_font(lang)
    font_bold = pdf_font(lang, bold=True)

    HDR_COLOR = colors.HexColor('#a85c42')
    ALT_COLOR  = colors.HexColor('#FDF0E8')

    if progress_fn:
        progress_fn(10)

    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=22 * mm,
    )
    page_w = A4[0] - 30 * mm

    title_style = ParagraphStyle(
        'UGTitle', fontName=font_bold, fontSize=10,
        textColor=HDR_COLOR, spaceAfter=4,
    )

    def _tbl_style(n_rows):
        s = [
            ('BACKGROUND',    (0, 0), (-1, 0),  HDR_COLOR),
            ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.black),
            ('FONTNAME',      (0, 0), (-1, 0),  font_bold),
            ('FONTSIZE',      (0, 0), (-1, 0),  7),
            ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
            ('FONTNAME',      (0, 1), (-1, -1), font),
            ('FONTSIZE',      (0, 1), (-1, -1), 7),
            ('ALIGN',         (0, 1), (0, -1),  'CENTER'),
            ('ALIGN',         (1, 1), (1, -1),  'LEFT'),
            ('ALIGN',         (2, 1), (-1, -1), 'CENTER'),
            ('GRID',          (0, 0), (-1, -1), 0.25, colors.HexColor('#aaaaaa')),
            ('TOPPADDING',    (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ]
        for i in range(1, n_rows):
            bg = colors.white if i % 2 else ALT_COLOR
            s.append(('BACKGROUND', (0, i), (-1, i), bg))
        return TableStyle(s)

    def _build_table(groups, has_swap):
        col_num  = t.get("usrgrp_col_num",  "#")
        col_name = t.get("usrgrp_col_name", "Name")
        col_gpin = t.get("usrgrp_col_gpin", "GPIN")
        col_bits = t.get("usrgrp_col_bits", "Bit")
        col_swap = t.get("usrgrp_col_swap", "SWAP")

        if has_swap:
            hdrs = [col_num, col_name, col_gpin, col_bits, col_swap]
            col_w = [12 * mm, None, 22 * mm, 14 * mm, 16 * mm]
        else:
            hdrs = [col_num, col_name, col_gpin, col_bits]
            col_w = [12 * mm, None, 22 * mm, 14 * mm]

        used = sum(x for x in col_w if x is not None)
        col_w[1] = page_w - used

        active = [g for g in groups if is_active(g)]
        if not active:
            return None

        data = [hdrs]
        for g in active:
            row = [str(g['num']), g['name'], str(g['gpin']), str(g['bits'])]
            if has_swap:
                row.append(str(g['val3']))
            data.append(row)

        tbl = Table(data, colWidths=col_w)
        tbl.setStyle(_tbl_style(len(data)))
        return tbl

    story = []
    gpin_active = [g for g in gpin_groups if is_active(g)]
    gpot_active = [g for g in gpot_groups if is_active(g)]

    if progress_fn:
        progress_fn(30)

    if gpin_active:
        story.append(Paragraph(
            t.get("usrgrp_section_gpin", "INPUT GROUPS (USRGRPIN)"), title_style))
        tbl = _build_table(gpin_groups, has_swap=False)
        if tbl:
            story.append(tbl)
        story.append(Spacer(1, 8 * mm))

    if progress_fn:
        progress_fn(60)

    if gpot_active:
        story.append(Paragraph(
            t.get("usrgrp_section_gpot", "OUTPUT GROUPS (USRGRPOT)"), title_style))
        tbl = _build_table(gpot_groups, has_swap=False)
        if tbl:
            story.append(tbl)

    if not gpin_active and not gpot_active:
        story.append(Paragraph(
            t.get("usrgrp_no_active", "No active groups"), title_style))

    if progress_fn:
        progress_fn(80)

    footer = make_footer("USER GROUP", page_offset)
    doc.build(story, onFirstPage=footer, onLaterPages=footer)

    if progress_fn:
        progress_fn(100)

    return len(gpin_active), len(gpot_active)


# ── Excel generation ──────────────────────────────────────────────────────────

def generate_excel(gpin_groups, gpot_groups, output_path, lang="IT"):
    """Generate Excel with two sheets: GPIN and GPOT."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from translations import TRANSLATIONS

    t = TRANSLATIONS.get(lang, TRANSLATIONS["IT"])

    HDR_FILL = PatternFill("solid", fgColor="1B3A6B")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
    ALT_FILL = PatternFill("solid", fgColor="EEF2FF")
    DATA_FONT = Font(size=9)
    CENTER    = Alignment(horizontal="center", vertical="center")

    col_num  = t.get("usrgrp_col_num",  "#")
    col_name = t.get("usrgrp_col_name", "Name")
    col_gpin = t.get("usrgrp_col_gpin", "GPIN")
    col_bits = t.get("usrgrp_col_bits", "Bit")
    col_swap = t.get("usrgrp_col_swap", "SWAP")

    def _fill_ws(ws, groups, has_swap, tab_name):
        ws.title = tab_name[:31]
        if has_swap:
            headers = [col_num, col_name, col_gpin, col_bits, col_swap]
            widths   = [6, 25, 10, 8, 8]
        else:
            headers = [col_num, col_name, col_gpin, col_bits]
            widths   = [6, 25, 10, 8]

        ws.append(headers)
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        for cell in ws[1]:
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = CENTER

        from docs.utils import excel_safe
        for ri, g in enumerate([g for g in groups if is_active(g)], 2):
            row = [g['num'], excel_safe(g['name']), g['gpin'], g['bits']]
            if has_swap:
                row.append(g['val3'])
            ws.append(row)
            fill = ALT_FILL if ri % 2 == 0 else None
            for ci in range(1, len(headers) + 1):
                c = ws.cell(row=ri, column=ci)
                c.font = DATA_FONT
                if fill:
                    c.fill = fill

    wb = openpyxl.Workbook()
    ws1 = wb.active
    _fill_ws(ws1, gpin_groups, has_swap=False,
             tab_name=t.get("usrgrp_tab_gpin", "USRGRPIN"))
    ws2 = wb.create_sheet()
    _fill_ws(ws2, gpot_groups, has_swap=False,
             tab_name=t.get("usrgrp_tab_gpot", "USRGRPOT"))
    wb.save(output_path)


# ── File writers ──────────────────────────────────────────────────────────────

def write_gpin(groups, filepath):
    """Write all groups to USRGRPIN.DAT in Yaskawa native format."""
    lines = []
    for g in groups:
        lines.append(f"//USRGRPIN {g['num']}")
        name = g.get('name', '').strip()
        lines.append(f"///NAME {name}" if name else "///NAME")
        lines.append(f"{g.get('gpin', 0)},{g.get('bits', 0)},{g.get('val3', 0)}")
    content = "\n".join(lines) + "\n"
    with open(filepath, "w", encoding="latin-1", newline="") as f:
        f.write(content)


def write_gpot(groups, filepath):
    """Write all groups to USRGRPOT.DAT in Yaskawa native format."""
    lines = []
    for g in groups:
        lines.append(f"//USRGRPOT {g['num']}")
        name = g.get('name', '').strip()
        lines.append(f"///NAME {name}" if name else "///NAME")
        lines.append(f"{g.get('gpin', 0)},{g.get('bits', 0)},{g.get('val3', 0)}")
    content = "\n".join(lines) + "\n"
    with open(filepath, "w", encoding="latin-1", newline="") as f:
        f.write(content)
