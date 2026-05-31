"""Parser for YASKAWA YRC1000 network configuration files (IPNETCFG.DAT + IPNETEX.DAT)."""

import os
import re


def _read_lines(filepath):
    try:
        with open(filepath, 'r', encoding='latin-1', errors='replace') as f:
            return f.readlines()
    except OSError:
        return []


def _strip(line):
    return line.rstrip('\r\n').strip()


def _parse_ip_quad(value):
    """Return (ip, mask, gw) from a comma-separated IP config line, or raw string."""
    parts = value.split(',')
    if len(parts) >= 3:
        return parts[1].strip(), parts[2].strip(), (parts[3].strip() if len(parts) > 3 else '')
    return value, '', ''


def parse_ipnetcfg(filepath):
    """Parse IPNETCFG.DAT → list of (section, param, value) tuples."""
    lines = _read_lines(filepath)
    rows = []
    section = ''
    section_lines = []
    ip_index = 0
    host_index = 0

    i = 0
    while i < len(lines):
        line = _strip(lines[i])
        i += 1

        if not line or line.startswith('//IPNETCFG'):
            continue

        if line.startswith('///'):
            tag = line[3:].strip()
            section = tag
            ip_index = 0
            host_index = 0
            section_lines = []
        else:
            section_lines.append(line)
            if section == 'IP':
                ip_index += 1
                idx = (ip_index + 1) // 2
                sub = ip_index % 2
                if sub == 1:
                    ip, mask, gw = _parse_ip_quad(line)
                    rows.append((f'IP[{idx}]', 'Indirizzo IP', ip))
                    rows.append((f'IP[{idx}]', 'Subnet Mask', mask))
                    if gw:
                        rows.append((f'IP[{idx}]', 'Gateway', gw))
                # even lines: MAC + stats, skip detailed parsing
            elif section == 'HOST':
                host_index += 1
                if host_index == 1:
                    rows.append(('HOST', 'Modalità host', line))
                elif host_index == 2:
                    rows.append(('HOST', 'Nome host', line))
            elif section == 'DOMAIN':
                if len(section_lines) == 1:
                    rows.append(('DOMAIN', 'Modalità dominio', line))
                elif len(section_lines) == 2:
                    rows.append(('DOMAIN', 'Nome dominio', line))
            elif section == 'DGW':
                if len(section_lines) == 2:
                    rows.append(('DGW', 'Gateway default', line))
            elif section == 'DNSC':
                if len(section_lines) == 1:
                    rows.append(('DNSC', 'DNS primario', line))
                elif len(section_lines) == 2:
                    rows.append(('DNSC', 'DNS secondario', line))
            elif section == 'SNTPC':
                if len(section_lines) == 1:
                    rows.append(('NTP', 'Server NTP', line))
                elif len(section_lines) == 2:
                    rows.append(('NTP', 'Intervallo NTP (s)', line))
            elif section == 'FTPC':
                if len(section_lines) == 1:
                    rows.append(('FTP', 'Server FTP', line))
                elif len(section_lines) == 2:
                    rows.append(('FTP', 'Porta FTP', line))

    return rows


def parse_ipnetex(filepath):
    """Parse IPNETEX.DAT → list of (section, param, value) tuples."""
    lines = _read_lines(filepath)
    rows = []
    section = ''
    ip_index = 0

    for raw in lines:
        line = _strip(raw)
        if not line or line.startswith('//IPNETEX'):
            continue
        if line.startswith('///'):
            section = line[3:].strip()
            ip_index = 0
        elif section == 'IP':
            ip_index += 1
            sub = ip_index % 2
            idx = (ip_index + 1) // 2
            if sub == 1:
                ip, mask, gw = _parse_ip_quad(line)
                rows.append((f'EXT IP[{idx}]', 'Indirizzo IP', ip))
                rows.append((f'EXT IP[{idx}]', 'Subnet Mask', mask))
                if gw:
                    rows.append((f'EXT IP[{idx}]', 'Gateway', gw))

    return rows


def generate_ipnet_pdf(rows, output_path, folder_name="", lang="IT", log_fn=None):
    """Generate A4 PDF from IP network config rows. Returns True/False."""
    try:
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor, white, black
    except ImportError as exc:
        if log_fn:
            try:
                log_fn("log_error_generic", f"reportlab: {exc}")
            except Exception:
                pass
        return False

    from translations import TRANSLATIONS
    tr = TRANSLATIONS.get(lang, TRANSLATIONS.get("IT", {}))
    from docs.utils import pdf_font, xml_escape
    fn   = pdf_font(lang)
    fn_b = pdf_font(lang, bold=True)

    hdr_blue = HexColor('#A85C42')
    alt_gray = HexColor('#FDF0E8')
    styles   = getSampleStyleSheet()

    h_style = ParagraphStyle('IpHdr', parent=styles['Normal'],
        fontSize=9, fontName=fn_b,
        textColor=HexColor('#A85C42'), spaceBefore=4, spaceAfter=3)
    c_style = ParagraphStyle('IpCell', parent=styles['Normal'],
        fontSize=8, fontName=fn, leading=10)

    def _draw(canvas, doc):
        from docs.pdf_header import draw_page_header
        draw_page_header(canvas, doc)
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(HexColor('#888888'))
        canvas.drawCentredString(A4[0] / 2, 8 * mm, str(canvas.getPageNumber()))
        canvas.restoreState()

    col_s = tr.get("ipnet_pdf_section", "Sezione")
    col_p = tr.get("ipnet_pdf_param",   "Parametro")
    col_v = tr.get("ipnet_pdf_value",   "Valore")

    W = A4[0] - 30 * mm
    col_widths = [40 * mm, 70 * mm, W - 110 * mm]

    tbl_data = [[col_s, col_p, col_v]]
    for section, param, value in rows:
        tbl_data.append([
            Paragraph(xml_escape(section), c_style),
            Paragraph(xml_escape(param),   c_style),
            Paragraph(xml_escape(value),   c_style),
        ])

    ts = TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  hdr_blue),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  black),
        ('FONTNAME',      (0, 0), (-1, 0),  fn_b),
        ('FONTSIZE',      (0, 0), (-1, 0),  8),
        ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
        ('FONTNAME',      (0, 1), (-1, -1), fn),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('GRID',          (0, 0), (-1, -1), 0.3, HexColor('#BBBBBB')),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
    ])
    for i in range(1, len(tbl_data)):
        if i % 2 == 0:
            ts.add('BACKGROUND', (0, i), (-1, i), alt_gray)

    story = []
    view_title = tr.get("ipnet_view_title", "Network Configuration — YASKAWA YRC1000")
    if folder_name:
        view_title += f"  –  {folder_name}"
    story.append(Paragraph(xml_escape(view_title), h_style))
    story.append(Spacer(1, 3 * mm))
    if tbl_data[1:]:
        tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(ts)
        story.append(tbl)
    else:
        story.append(Paragraph(tr.get("ipnet_no_data", "No data"), styles['Normal']))

    try:
        doc = SimpleDocTemplate(output_path, pagesize=A4,
            leftMargin=15 * mm, rightMargin=15 * mm,
            topMargin=22 * mm, bottomMargin=16 * mm)
        doc.build(story, onFirstPage=_draw, onLaterPages=_draw)
        if log_fn:
            try:
                log_fn("log_ipnet_pdf_saved", output_path)
            except Exception:
                pass
        return True
    except Exception as exc:
        if log_fn:
            try:
                log_fn("log_error_generic", str(exc))
            except Exception:
                pass
        return False


def generate_ipnet_excel(rows, output_path, folder_name="", lang="IT", log_fn=None):
    """Generate Excel from IP network config rows. Returns True/False."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError as exc:
        if log_fn:
            try:
                log_fn("log_error_generic", f"openpyxl: {exc}")
            except Exception:
                pass
        return False

    from translations import TRANSLATIONS
    from docs.utils import excel_safe
    tr = TRANSLATIONS.get(lang, TRANSLATIONS.get("IT", {}))

    col_s = tr.get("ipnet_pdf_section", "Sezione")
    col_p = tr.get("ipnet_pdf_param",   "Parametro")
    col_v = tr.get("ipnet_pdf_value",   "Valore")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IP Net"

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1B3A6B")
    thin      = Side(style="thin", color="AAAAAA")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append([col_s, col_p, col_v])
    for cell in ws[1]:
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for section, param, value in rows:
        ws.append([excel_safe(section), excel_safe(param), excel_safe(value)])
        for cell in ws[ws.max_row]:
            cell.border = border

    for col in ws.columns:
        w = max(len(str(cell.value or "")) for cell in col) + 4
        ws.column_dimensions[col[0].column_letter].width = min(w, 40)

    try:
        wb.save(output_path)
        if log_fn:
            try:
                log_fn("log_ipnet_excel_saved", output_path)
            except Exception:
                pass
        return True
    except Exception as exc:
        if log_fn:
            try:
                log_fn("log_error_generic", str(exc))
            except Exception:
                pass
        return False


def load_network_config(folder):
    """Load IPNETCFG.DAT and IPNETEX.DAT from folder.
    Returns list of (section, param, value) or [].
    Silently skips missing files."""
    rows = []
    cfg_path = os.path.join(folder, 'IPNETCFG.DAT')
    ext_path = os.path.join(folder, 'IPNETEX.DAT')

    if os.path.isfile(cfg_path):
        rows.extend(parse_ipnetcfg(cfg_path))

    if os.path.isfile(ext_path):
        ext = parse_ipnetex(ext_path)
        # deduplicate: skip EXT entries already present in CFG
        existing = {(s, p, v) for s, p, v in rows}
        for row in ext:
            if row not in existing:
                rows.append(row)
                existing.add(row)

    # Remove rows with empty or zero-only IP addresses
    rows = [
        (s, p, v) for s, p, v in rows
        if v and v not in ('0', '0.0.0.0', '', '00:00:00:00:00:00')
    ]

    return rows
