import os
import sys

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.platypus import (BaseDocTemplate, Frame, PageTemplate,
                                Paragraph, Spacer, PageBreak,
                                Table, TableStyle, HRFlowable, Image)
from reportlab.platypus.tableofcontents import TableOfContents
from reportlab.platypus.flowables import Flowable
from docs.pdf_header import draw_page_header as _draw_header

PARAM_FILE = "ALL.PRM"

_state = {"section": "PARAMETRI"}


# ── Dynamic footer tracker ────────────────────────────────────────────────────

class _SetSection(Flowable):
    def __init__(self, name):
        super().__init__()
        self.name = name
        self.width = 0
        self.height = 0

    def draw(self):
        _state["section"] = self.name


# ── Bookmarkable heading ──────────────────────────────────────────────────────

class _Heading(Paragraph):
    def __init__(self, text, key, style, heading_text=None):
        super().__init__(text, style)
        self.key = key
        self.heading_text = heading_text if heading_text is not None else text

    def drawOn(self, canvas, x, y, _sW=0):
        canvas.bookmarkPage(self.key)
        super().drawOn(canvas, x, y, _sW)


# ── Custom doc template with TOC + onPageEnd footer ──────────────────────────

class _ParamsDocTemplate(BaseDocTemplate):
    def __init__(self, output_path, footer_fn, **kw):
        super().__init__(output_path, **kw)
        W, H = A4
        frame = Frame(
            self.leftMargin, self.bottomMargin,
            W - self.leftMargin - self.rightMargin,
            H - self.topMargin - self.bottomMargin,
            id="normal", leftPadding=0, rightPadding=0,
            topPadding=0, bottomPadding=0,
        )
        self.addPageTemplates([
            PageTemplate(id="main", frames=[frame],
                         onPage=_draw_header, onPageEnd=footer_fn)
        ])

    def afterFlowable(self, flowable):
        if isinstance(flowable, _Heading):
            self.notify("TOCEntry", (0, flowable.heading_text, self.page, flowable.key))
            # bookmarkPage called in _Heading.drawOn for correct top-of-heading placement


# ── Helpers ───────────────────────────────────────────────────────────────────

def _logo_path():
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, "logo-home.bmp")
    root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    return os.path.join(root, "logo-home.bmp")


def _make_footer(page_offset=0):
    GRAY = colors.HexColor("#aaaaaa")

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(GRAY)
        W = doc.pagesize[0]
        canvas.drawCentredString(W / 2.0, 10 * mm, str(doc.page + page_offset))
        canvas.drawRightString(W - 20 * mm, 10 * mm, _state["section"])
        canvas.restoreState()

    return _footer


# ── Parser ────────────────────────────────────────────────────────────────────

def parse_all_prm(filepath):
    """
    Parse ALL.PRM.
    Returns list of (section_name, [(param_index, value), ...]).
    Only non-zero values are included.
    Last line of file (checksum) is excluded.
    """
    sections = []
    current_section = None
    current_data = []
    param_idx = 0

    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
    except OSError as e:
        raise RuntimeError(f"Cannot read {filepath}: {e}") from e

    # Exclude last line (checksum)
    if lines:
        lines = lines[:-1]

    for line in lines:
        s = line.strip()
        if not s:
            continue
        if s.startswith("///"):
            if current_section is not None:
                sections.append((current_section, current_data))
            current_section = s[3:].strip()
            current_data = []
            param_idx = 0
        elif s.startswith("/"):
            continue  # skip file/group headers
        else:
            if current_section is None:
                continue
            for part in s.split(","):
                try:
                    val = int(part.strip())
                    if val != 0:
                        current_data.append((param_idx, val))
                except ValueError:
                    pass
                param_idx += 1

    if current_section is not None:
        sections.append((current_section, current_data))

    return sections


# ── Table builder ─────────────────────────────────────────────────────────────

def _build_section_table(section_name, param_data, s_hdr, s_name, s_val, s_note,
                         col_param="Parametro", col_value="Valore", col_note="Note"):
    """
    2-column table for one section's non-zero parameters.
    Columns: [name1 | value1 | note1 | name2 | value2 | note2]
    Alternating row colours; header row in the accent colour.
    """
    ACCENT   = colors.HexColor("#D97757")
    WHITE  = colors.white
    LACCENT  = colors.HexColor("#FDF0E8")   # even data rows
    CGRID  = colors.HexColor("#dddddd")
    MSEP   = colors.HexColor("#bbbbbb")   # mid-page vertical separator

    COL_W = [24*mm, 20*mm, 38*mm, 24*mm, 20*mm, 38*mm]  # 164mm — fits in 170mm

    from docs.utils import xml_escape

    # Build pairs of (index, value) for 2-column layout
    pairs = []
    for i in range(0, len(param_data), 2):
        left  = param_data[i]
        right = param_data[i + 1] if i + 1 < len(param_data) else None
        pairs.append((left, right))

    # Header row
    hdr_cell = lambda t: Paragraph(t, s_hdr)
    rows = [[hdr_cell(col_param), hdr_cell(col_value), hdr_cell(col_note),
             hdr_cell(col_param), hdr_cell(col_value), hdr_cell(col_note)]]

    for left, right in pairs:
        l_name = f"{section_name}{left[0]:04d}"
        l_val  = str(left[1])
        r_name = f"{section_name}{right[0]:04d}" if right else ""
        r_val  = str(right[1])                   if right else ""

        rows.append([
            Paragraph(xml_escape(l_name), s_name),
            Paragraph(xml_escape(l_val),  s_val),
            Paragraph("",                 s_note),
            Paragraph(xml_escape(r_name), s_name),
            Paragraph(xml_escape(r_val),  s_val),
            Paragraph("",                 s_note),
        ])

    t = Table(rows, colWidths=COL_W, repeatRows=1)
    t.setStyle(TableStyle([
        # Header
        ("BACKGROUND",    (0, 0), (-1, 0),  ACCENT),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.black),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  8),
        # Alternating data rows
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [WHITE, LACCENT]),
        # Grid
        ("GRID",          (0, 0), (-1, -1), 0.3, CGRID),
        # Vertical separator between the two halves
        ("LINEAFTER",     (2, 0), (2, -1),  1.0, MSEP),
        # Padding
        ("TOPPADDING",    (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    return t


# ── Excel export ─────────────────────────────────────────────────────────────

def export_excel(folder_path, output_path, lang="IT"):
    """
    Export ALL.PRM non-zero parameters to Excel.
    Single sheet with columns: Sezione, Parametro, Valore.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from translations import TRANSLATIONS
    tr = TRANSLATIONS.get(lang, TRANSLATIONS["IT"])

    filepath = os.path.join(folder_path, PARAM_FILE)
    if not os.path.isfile(filepath):
        raise FileNotFoundError(filepath)

    all_sections = parse_all_prm(filepath)
    sections = [(name, data) for name, data in all_sections if data]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tr.get("params_col_param", "Parametri")[:31]

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="D97757")
    hdr_align = Alignment(horizontal="center")

    col_sec   = tr.get("targ_section_info", "Sezione")
    col_param = tr.get("params_col_param",  "Parametro")
    col_val   = tr.get("params_col_value",  "Valore")

    ws.append([col_sec, col_param, col_val])
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    from docs.utils import excel_safe
    alt_fill = PatternFill("solid", fgColor="EEF2FF")
    row_idx = 2
    for section_name, param_data in sections:
        for param_idx, value in param_data:
            param_label = f"{section_name}{param_idx:04d}"
            row = [excel_safe(section_name), param_label, excel_safe(value)]
            ws.append(row)
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill
            row_idx += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14

    wb.save(output_path)


# ── PDF generation ────────────────────────────────────────────────────────────

def generate_pdf(folder_path, output_path, log_fn=None, lang="IT", page_offset=0):
    from translations import TRANSLATIONS
    from docs.utils import pdf_font, xml_escape
    tr = TRANSLATIONS.get(lang, TRANSLATIONS["IT"])

    filepath = os.path.join(folder_path, PARAM_FILE)
    if not os.path.isfile(filepath):
        if log_fn:
            log_fn("log_file_not_found", PARAM_FILE)
        return

    try:
        all_sections = parse_all_prm(filepath)
    except Exception as exc:
        if log_fn:
            log_fn("log_error_generic", str(exc))
        return

    # Log every section (including empty ones)
    if log_fn:
        for name, data in all_sections:
            log_fn("log_param_section", name, len(data))

    # Only sections with non-zero params appear in the PDF
    sections = [(name, data) for name, data in all_sections if data]

    # ── Styles ────────────────────────────────────────────────────────────────
    ACCENT  = colors.HexColor("#D97757")
    WHITE = colors.white
    GRAY  = colors.HexColor("#888888")

    W = A4[0] - 40 * mm
    folder_name = os.path.basename(os.path.normpath(folder_path))
    f_reg  = pdf_font(lang, bold=False)
    f_bold = pdf_font(lang, bold=True)

    def ps(name, font=None, size=9, color=colors.black,
           align=TA_LEFT, **kw):
        fn = font if font is not None else f_reg
        return ParagraphStyle(name, fontName=fn, fontSize=size,
                              textColor=color, alignment=align,
                              leading=size * 1.4, **kw)

    s_head  = ps("hd", f_bold, 15, WHITE,           TA_CENTER)
    s_sec   = ps("sc", f_bold,  9, ACCENT,            spaceBefore=4, spaceAfter=2)
    s_job   = ps("jb", f_bold, 10, ACCENT,            spaceBefore=4, spaceAfter=1)
    s_miss  = ps("ms", f_reg,   8, GRAY)
    s_hdr   = ps("th", f_bold,  8, WHITE)
    s_name  = ps("pn", f_bold,  8, colors.HexColor("#A85C42"))
    s_val   = ps("pv", f_reg,   8, colors.black,    TA_RIGHT)
    s_note  = ps("nt", f_reg,   8, colors.HexColor("#aaaaaa"))

    def make_header():
        hdr = Table(
            [[Paragraph(f"PARAMETRI<br/><font size='10'>{xml_escape(folder_name)}</font>",
                        s_head)]],
            colWidths=[W],
        )
        hdr.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), ACCENT),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ]))
        return hdr

    # ── TOC ───────────────────────────────────────────────────────────────────
    toc = TableOfContents()
    toc.levelStyles = [
        ParagraphStyle("toc0", fontName=f_reg, fontSize=9,
                       leading=14, leftIndent=0, textColor=ACCENT)
    ]

    # ── Document ──────────────────────────────────────────────────────────────
    _state["section"] = "PARAMETRI"
    footer_fn = _make_footer(page_offset)
    doc = _ParamsDocTemplate(
        output_path, footer_fn,
        pagesize=A4,
        leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=22 * mm, bottomMargin=22 * mm,
    )

    CGRID = colors.HexColor("#cccccc")
    story = []

    # First flowable resets state at the start of every multiBuild pass
    story.append(_SetSection("PARAMETRI"))

    lbl_sommario = tr.get("jobs_sommario", "SOMMARIO")

    # ── SOMMARIO page ─────────────────────────────────────────────────────────
    story.append(make_header())
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(lbl_sommario, s_sec))
    story.append(Spacer(1, 2 * mm))

    if not sections:
        story.append(Paragraph(tr.get("params_no_params", "Nessun parametro non-zero trovato in ALL.PRM."), s_miss))
    else:
        story.append(toc)

    story.append(PageBreak())

    col_param = tr.get("params_col_param", "Parametro")
    col_value = tr.get("params_col_value", "Valore")
    col_note  = tr.get("params_col_note",  "Note")
    lbl_count = tr.get("params_params_count", "{} parametri")

    # ── Section pages ─────────────────────────────────────────────────────────
    for section_name, param_data in sections:
        key           = f"prm_{section_name}"
        count_nonzero = len(param_data)
        heading_label = f"{section_name}  ({lbl_count.format(count_nonzero)})"

        story.append(_SetSection(section_name))
        story.append(_Heading(heading_label, key, s_job,
                              heading_text=heading_label))
        story.append(HRFlowable(width="100%", thickness=0.5,
                                color=CGRID, spaceAfter=3))
        story.append(Spacer(1, 1 * mm))

        try:
            tbl = _build_section_table(section_name, param_data,
                                       s_hdr, s_name, s_val, s_note,
                                       col_param, col_value, col_note)
            story.append(tbl)
        except Exception as exc:
            if log_fn:
                log_fn("log_error_generic", f"{section_name}: {exc}")
            story.append(Paragraph(f"[Errore generazione tabella: {exc}]", s_miss))

        story.append(PageBreak())

    doc.multiBuild(story)
