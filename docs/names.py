import os
import re

VARNAME_FILE  = "VARNAME.DAT"
IONAME_FILE   = "IONAME.DAT"
EXIONAME_FILE = "EXIONAME.DAT"

_SHARE_DEFAULT = "2000,2000,2000,2000,2000,2000,10000,10000"
_VAR_ORDER     = ["B", "I", "D", "R", "S", "P", "BP", "EX"]
_VAR_COUNTS    = {"B": 2000, "I": 2000, "D": 2000, "R": 2000,
                  "S": 2000, "P": 2000, "BP": 10000, "EX": 10000}
_IO_LINES      = 1024
_IO_PER_LINE   = 4
_IO_COUNT      = _IO_LINES * _IO_PER_LINE   # 4096
_MAX_NAME      = 16

# Excel tabs for variables (BP/EX omitted – rarely named)
_VAR_TABS = ["B", "I", "D", "R", "S", "P"]
_VAR_PFX  = {"B": "B", "I": "I", "D": "D", "R": "R", "S": "S", "P": "P"}

# (excel_tab, file_key, signal_display_prefix)
_IO_TABS = [
    ("IN",    "IN",    "IN#"),
    ("OUT",   "OUT",   "OUT#"),
    ("EXIN",  "EXIN",  "EIN#"),
    ("EXOUT", "EXOUT", "EOUT#"),
]


# ── Low-level file helpers ────────────────────────────────────────────────────

def _open_r(filepath):
    return open(filepath, "r", encoding="latin-1", errors="replace")

def _open_w(filepath):
    return open(filepath, "w", encoding="latin-1", newline="\r\n")


# ── Parsers ───────────────────────────────────────────────────────────────────

def parse_varname(filepath):
    """Parse VARNAME.DAT.  Returns (data_dict, share_line_str).
    data_dict: {type: [name_or_'' × count]}"""
    data = {k: [""] * v for k, v in _VAR_COUNTS.items()}
    share = _SHARE_DEFAULT
    try:
        with _open_r(filepath) as f:
            lines = f.readlines()
    except OSError:
        return data, share

    current = None
    for line in lines:
        s = line.rstrip("\r\n")
        if s.startswith("///SHARE"):
            share = s[len("///SHARE"):].strip()
            continue
        if s.startswith("///"):
            tok = s[3:].strip().split()[0] if s[3:].strip() else ""
            current = tok if tok in _VAR_COUNTS else None
            continue
        if s.startswith("/") or current is None:
            continue
        if not s.strip():
            continue
        m = re.match(r"^(\d{1,5})\s+\d+,\d+,(.*)$", s.strip())
        if m:
            idx, name = int(m.group(1)), m.group(2)
            if 0 <= idx < _VAR_COUNTS.get(current, 0):
                data[current][idx] = name
    return data, share


def _fill_io_sections(lines, result, section_map):
    current = None
    row = 0
    for line in lines:
        s = line.rstrip("\r\n")
        if s in section_map:
            current = section_map[s]
            row = 0
            continue
        if s.startswith("/"):
            continue
        if current is None:
            continue
        # pad to 4 fields
        fields = (s + ",,,").split(",")[:_IO_PER_LINE]
        base = row * _IO_PER_LINE
        for i in range(_IO_PER_LINE):
            idx = base + i
            if idx < _IO_COUNT:
                result[current][idx] = fields[i]
        row += 1


def parse_ioname(filepath):
    """Parse IONAME.DAT → {'IN': [...4096], 'OUT': [...4096]}."""
    result = {"IN": [""] * _IO_COUNT, "OUT": [""] * _IO_COUNT}
    try:
        with _open_r(filepath) as f:
            lines = f.readlines()
    except OSError:
        return result
    _fill_io_sections(lines, result, {"//IN": "IN", "//OUT": "OUT"})
    return result


def parse_exioname(filepath):
    """Parse EXIONAME.DAT → {'EXIN': [...], 'EXOUT': [...]}."""
    result = {"EXIN": [""] * _IO_COUNT, "EXOUT": [""] * _IO_COUNT}
    try:
        with _open_r(filepath) as f:
            lines = f.readlines()
    except OSError:
        return result
    _fill_io_sections(lines, result, {"//EXIN": "EXIN", "//EXOUT": "EXOUT"})
    return result


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _xl_init_sheet(ws, col_specs):
    """Header row + column widths.  col_specs = [(header, width), ...]"""
    from openpyxl.styles import PatternFill, Font, Alignment
    BLUE   = PatternFill("solid", fgColor="0072BC")
    WHITE  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    CENTER = Alignment(horizontal="center", vertical="center")
    for ci, (hdr, w) in enumerate(col_specs, start=1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.fill      = BLUE
        c.font      = WHITE
        c.alignment = CENTER
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def _xl_write_data(ws, ids, names):
    """Write id (gray, monospace) + name rows starting at row 2."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from docs.utils import excel_safe
    GRAY     = PatternFill("solid", fgColor="EFEFEF")
    ID_FONT  = Font(name="Courier New", size=9, color="003366")
    ID_ALIGN = Alignment(horizontal="center")
    for row_i, (vid, name) in enumerate(zip(ids, names), start=2):
        ca = ws.cell(row=row_i, column=1, value=excel_safe(vid))
        ca.fill      = GRAY
        ca.font      = ID_FONT
        ca.alignment = ID_ALIGN
        if name:
            ws.cell(row=row_i, column=2, value=excel_safe(name))


def _xl_add_var_tab(wb, tab, prefix, names):
    ws = wb.create_sheet(title=tab)
    _xl_init_sheet(ws, [
        ("ID", 12),
        (f"Descrizione (max {_MAX_NAME} car.)", 26),
        ("Note", 42),
    ])
    ids = [f"{prefix}{i:04d}" for i in range(len(names))]
    _xl_write_data(ws, ids, names)


def _xl_add_io_tab(wb, tab, prefix, names):
    ws = wb.create_sheet(title=tab)
    _xl_init_sheet(ws, [
        ("Segnale", 12),
        (f"Descrizione (max {_MAX_NAME} car.)", 26),
        ("Note", 42),
    ])
    ids = [f"{prefix}{i:04d}" for i in range(len(names))]
    _xl_write_data(ws, ids, names)


# ── Template generator ────────────────────────────────────────────────────────

def generate_template(folder_path, output_path, log_fn=None):
    """Read DAT files from folder_path, write Excel template to output_path."""
    import openpyxl

    vn_path = os.path.join(folder_path, VARNAME_FILE)
    io_path = os.path.join(folder_path, IONAME_FILE)
    ex_path = os.path.join(folder_path, EXIONAME_FILE)

    for fname, fpath in [(VARNAME_FILE, vn_path),
                         (IONAME_FILE,  io_path),
                         (EXIONAME_FILE, ex_path)]:
        if os.path.isfile(fpath):
            try:
                with _open_r(fpath) as f:
                    n = sum(1 for _ in f)
                if log_fn:
                    log_fn("log_file_read", fname, n)
            except Exception:
                if log_fn:
                    log_fn("log_file_not_found", fname)
        else:
            if log_fn:
                log_fn("log_file_not_found", fname)

    if os.path.isfile(vn_path):
        var_data, share = parse_varname(vn_path)
    else:
        var_data, share = {k: [""] * v for k, v in _VAR_COUNTS.items()}, _SHARE_DEFAULT

    io_data  = parse_ioname(io_path)   if os.path.isfile(io_path)  else \
               {"IN": [""] * _IO_COUNT, "OUT": [""] * _IO_COUNT}
    ex_data  = parse_exioname(ex_path) if os.path.isfile(ex_path)  else \
               {"EXIN": [""] * _IO_COUNT, "EXOUT": [""] * _IO_COUNT}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for tab in _VAR_TABS:
        names = var_data.get(tab, [""] * _VAR_COUNTS.get(tab, 0))
        _xl_add_var_tab(wb, tab, _VAR_PFX[tab], names)
        if log_fn:
            log_fn("log_template_tab", tab, sum(1 for n in names if n))

    io_map = {
        "IN":    (io_data.get("IN",    [""] * _IO_COUNT), "IN#"),
        "OUT":   (io_data.get("OUT",   [""] * _IO_COUNT), "OUT#"),
        "EXIN":  (ex_data.get("EXIN",  [""] * _IO_COUNT), "EIN#"),
        "EXOUT": (ex_data.get("EXOUT", [""] * _IO_COUNT), "EOUT#"),
    }
    for tab, file_key, _ in _IO_TABS:
        names, pfx = io_map[tab]
        _xl_add_io_tab(wb, tab, pfx, names)
        if log_fn:
            log_fn("log_template_tab", tab, sum(1 for n in names if n))

    wb.save(output_path)
    if log_fn:
        log_fn("log_template_saved", output_path)


# ── DAT writers ───────────────────────────────────────────────────────────────

def write_varname(output_path, var_data, share_line=None):
    """Write VARNAME.DAT from var_data dict."""
    if share_line is None:
        share_line = _SHARE_DEFAULT
    with _open_w(output_path) as f:
        f.write("//VARNAME\n")
        f.write(f"///SHARE {share_line}\n")
        for vtype in _VAR_ORDER:
            count = _VAR_COUNTS[vtype]
            names = var_data.get(vtype, [""] * count)
            f.write(f"///{vtype}\n")
            named = sorted(
                (i, n[:_MAX_NAME]) for i, n in enumerate(names) if n
            )
            for idx, name in named:
                f.write(f"{idx:04d} 1,0,{name}\n")
            for _ in range(count - len(named)):
                f.write("\n")


def write_ioname(output_path, io_data):
    """Write IONAME.DAT from {'IN': [...], 'OUT': [...]}."""
    with _open_w(output_path) as f:
        f.write("/IONAME\n")
        for sec in ("IN", "OUT"):
            f.write(f"//{sec}\n")
            names = io_data.get(sec, [""] * _IO_COUNT)
            for row in range(_IO_LINES):
                base = row * _IO_PER_LINE
                fields = [names[base + i] if base + i < len(names) else ""
                          for i in range(_IO_PER_LINE)]
                f.write(",".join(fields) + "\n")


def write_exioname(output_path, ex_data):
    """Write EXIONAME.DAT from {'EXIN': [...], 'EXOUT': [...]}."""
    with _open_w(output_path) as f:
        f.write("/EXIONAME\n")
        for sec in ("EXIN", "EXOUT"):
            f.write(f"//{sec}\n")
            names = ex_data.get(sec, [""] * _IO_COUNT)
            for row in range(_IO_LINES):
                base = row * _IO_PER_LINE
                fields = [names[base + i] if base + i < len(names) else ""
                          for i in range(_IO_PER_LINE)]
                f.write(",".join(fields) + "\n")


# ── Generate DAT files from Excel ─────────────────────────────────────────────

def _read_col_b(ws):
    """Read column B from row 2 onward; truncate to _MAX_NAME chars."""
    result = []
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        v = row[0]
        result.append("" if v is None else str(v).strip()[:_MAX_NAME])
    return result


def generate_dat_files(excel_path, output_folder, log_fn=None):
    """Read Excel template, write VARNAME.DAT + IONAME.DAT + EXIONAME.DAT."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as exc:
        if log_fn:
            log_fn("log_error_generic", str(exc))
        return

    var_data = {k: [""] * v for k, v in _VAR_COUNTS.items()}
    io_data  = {"IN":   [""] * _IO_COUNT, "OUT":  [""] * _IO_COUNT}
    ex_data  = {"EXIN": [""] * _IO_COUNT, "EXOUT":[""] * _IO_COUNT}

    for tab in _VAR_TABS:
        if tab not in wb.sheetnames:
            if log_fn:
                log_fn("log_tab_missing", tab)
            continue
        names = _read_col_b(wb[tab])
        count = _VAR_COUNTS[tab]
        for i, n in enumerate(names[:count]):
            var_data[tab][i] = n
        if log_fn:
            log_fn("log_names_var", tab, sum(1 for n in var_data[tab] if n))

    for tab_name, file_key, _ in _IO_TABS:
        if tab_name not in wb.sheetnames:
            if log_fn:
                log_fn("log_tab_missing", tab_name)
            continue
        names = _read_col_b(wb[tab_name])
        target = io_data if file_key in ("IN", "OUT") else ex_data
        for i, n in enumerate(names[:_IO_COUNT]):
            target[file_key][i] = n
        if log_fn:
            log_fn("log_names_io", file_key,
                   sum(1 for n in target[file_key] if n))

    wb.close()

    os.makedirs(output_folder, exist_ok=True)

    for fname, writer in [
        (VARNAME_FILE,  lambda p: write_varname(p, var_data)),
        (IONAME_FILE,   lambda p: write_ioname(p, io_data)),
        (EXIONAME_FILE, lambda p: write_exioname(p, ex_data)),
    ]:
        try:
            path = os.path.join(output_folder, fname)
            writer(path)
            if log_fn:
                log_fn("log_dat_saved", path)
        except Exception as exc:
            if log_fn:
                log_fn("log_error_generic", f"{fname}: {exc}")
