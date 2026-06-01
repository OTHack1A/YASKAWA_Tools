import os
import secure_paths
import re
import sys
import socket
import getpass
import platform
import psutil
from PySide6.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                               QMenuBar, QMenu, QTextEdit, QApplication, QSplitter, QLabel,
                               QDialog, QGroupBox, QTabWidget, QTableWidget,
                               QTableWidgetItem, QPushButton, QHeaderView, QFrame,
                               QStyledItemDelegate, QProgressBar)
from PySide6.QtGui import QAction, QPixmap, QTextCursor, QRegularExpressionValidator
from PySide6.QtCore import Qt, QRegularExpression, QObject

from gui.top_bar import TopBar
from translations import TRANSLATIONS
import logger

class _NumericCommaDelegate(QStyledItemDelegate):
    """Restricts cell editing to numeric values (integer or decimal with comma)."""
    _EDIT_RX  = re.compile(r'^[+\-]?\d*([,]\d*)?$')
    _FINAL_RX = re.compile(r'^[+\-]?\d+([,]\d+)?$|^[+\-]?\d+$')

    def createEditor(self, parent, option, index):
        from PySide6.QtWidgets import QLineEdit
        editor = QLineEdit(parent)
        rx = QRegularExpression(r'^[+\-]?\d*([,]\d*)?$')
        editor.setValidator(QRegularExpressionValidator(rx, editor))
        editor.inputRejected.connect(
            lambda: logger.warning("log_invalid_input", editor.text() or "—"))
        return editor

    def setModelData(self, editor, model, index):
        text = editor.text().strip()
        if not text or self._FINAL_RX.match(text):
            model.setData(index, text)
        else:
            logger.warning("log_invalid_input", text)


_RE_NOME = re.compile(r'^[A-Za-z0-9 ,_\-\.]*$')

class _NomeColumnDelegate(QStyledItemDelegate):
    """Column Nome: max 16 chars, [A-Za-z0-9 ,_-.] only. Logs warning on violation."""
    _MAX = 16

    def createEditor(self, parent, option, index):
        from PySide6.QtWidgets import QLineEdit
        editor = QLineEdit(parent)
        editor.setMaxLength(self._MAX)
        rx = QRegularExpression(r'[A-Za-z0-9 ,_\-\.]*')
        editor.setValidator(QRegularExpressionValidator(rx, editor))
        editor.inputRejected.connect(
            lambda: logger.warning("log_invalid_input", editor.text() or "—"))
        return editor

    def setModelData(self, editor, model, index):
        text = editor.text().strip()
        if not _RE_NOME.match(text):
            logger.warning("log_invalid_input", text)
            return
        if len(text) > self._MAX:
            logger.warning("log_invalid_input", text)
            return
        model.setData(index, text)


class _ExcelPreview(QWidget):
    """In-window Excel preview panel: shows named rows per expected tab."""

    def __init__(self, excel_path, sheets_data, app_state,
                 on_close_cb, on_generate_cb, editable=False, nomi_mode=False):
        super().__init__()
        self._excel_path   = excel_path
        self._app_state    = app_state
        self._on_close     = on_close_cb
        self._on_generate  = on_generate_cb
        self._editable     = editable
        self._nomi_mode    = nomi_mode
        self._sheet_tables = {}   # sheet_name → QTableWidget (for save_edits)
        self._build_ui(sheets_data)
        self._apply_theme()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self, sheets_data):
        t = TRANSLATIONS[self._app_state.language]
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 6, 8, 6)
        root.setSpacing(4)

        # Toolbar row
        toolbar = QHBoxLayout()
        toolbar.setSpacing(6)

        lbl = QLabel(os.path.basename(self._excel_path))
        lbl.setStyleSheet("font-weight: bold; font-size: 10pt;")
        toolbar.addWidget(lbl)
        toolbar.addStretch()

        if self._on_generate is not None:
            gen_key = "btn_genera" if (self._editable or self._nomi_mode) else "preview_generate"
            self._btn_gen = QPushButton(t.get(gen_key, "Genera"))
            self._btn_gen.setFixedHeight(26)
            self._btn_gen.clicked.connect(self._on_generate)
            toolbar.addWidget(self._btn_gen)

        self._btn_close = QPushButton(t["preview_close"])
        self._btn_close.setFixedHeight(26)
        self._btn_close.clicked.connect(self._on_close)
        toolbar.addWidget(self._btn_close)

        root.addLayout(toolbar)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Sunken)
        root.addWidget(sep)

        # Tab widget
        self._tab_widget = QTabWidget()
        self._tab_widget.setDocumentMode(True)

        for sheet_name, headers, named_rows, total in sheets_data:
            self._tab_widget.addTab(
                self._make_sheet_tab(t, headers, named_rows, total, sheet_name),
                sheet_name
            )

        root.addWidget(self._tab_widget, 1)

    def save_edits(self, out_path):
        """Write current table edits back to an Excel file (nomi_mode only)."""
        import openpyxl
        wb = openpyxl.load_workbook(self._excel_path)
        for sheet_name, tbl in self._sheet_tables.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows_iter = list(ws.iter_rows(min_row=2))
            for r in range(tbl.rowCount()):
                if r >= len(rows_iter):
                    break
                for c in range(tbl.columnCount()):
                    it = tbl.item(r, c)
                    val = it.text() if it else ""
                    rows_iter[r][c].value = val if val else None
        wb.save(out_path)

    # ── Language update ───────────────────────────────────────────────────────

    def update_language(self, lang):
        t = TRANSLATIONS[lang]
        if hasattr(self, '_btn_gen'):
            gen_key = "btn_genera" if (self._editable or self._nomi_mode) else "preview_generate"
            self._btn_gen.setText(t.get(gen_key, "Genera"))
        self._btn_close.setText(t.get("preview_close", "Chiudi"))

    def _make_sheet_tab(self, t, headers, named_rows, total, sheet_name=""):
        container = QWidget()
        lay = QVBoxLayout(container)
        lay.setContentsMargins(2, 4, 2, 2)
        lay.setSpacing(2)

        if self._nomi_mode:
            info_text = t.get("preview_rows", "{} / {}").format(len(named_rows), total)
        elif self._editable:
            info_text = t.get("preview_compila_info", "{} righe").format(len(named_rows))
        else:
            info_text = t["preview_rows"].format(len(named_rows), total)
        info = QLabel(info_text)
        info.setStyleSheet("color: gray; font-size: 8pt;")
        lay.addWidget(info)

        # Rename "Descrizione" column to "Nome" in nomi_mode
        disp_headers = list(headers)
        if self._nomi_mode and len(disp_headers) > 1:
            disp_headers[1] = t.get("compila_col_name", "Nome")

        n_cols = len(disp_headers)
        tbl = QTableWidget(len(named_rows), n_cols)
        tbl.setHorizontalHeaderLabels(disp_headers)
        if self._editable or self._nomi_mode:
            tbl.setEditTriggers(
                QTableWidget.DoubleClicked | QTableWidget.AnyKeyPressed)
        else:
            tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        tbl.setAlternatingRowColors(True)
        tbl.setSelectionBehavior(QTableWidget.SelectRows)
        tbl.verticalHeader().setVisible(False)
        tbl.horizontalHeader().setStretchLastSection(True)

        for r, row_data in enumerate(named_rows):
            row_list = list(row_data)
            for c in range(n_cols):
                val = row_list[c] if c < len(row_list) else None
                item = QTableWidgetItem("" if val is None else str(val))
                if (self._editable or self._nomi_mode) and c == 0:
                    item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                tbl.setItem(r, c, item)

        if self._nomi_mode and n_cols > 1:
            tbl.setItemDelegateForColumn(1, _NomeColumnDelegate(tbl))
        elif self._editable and n_cols > 2:
            tbl.setItemDelegateForColumn(2, _NumericCommaDelegate(tbl))

        tbl.resizeColumnToContents(0)
        if n_cols > 1:
            tbl.resizeColumnToContents(1)

        if sheet_name:
            self._sheet_tables[sheet_name] = tbl

        lay.addWidget(tbl, 1)
        return container

    # ── Theme ─────────────────────────────────────────────────────────────────

    def _apply_theme(self):
        if self._app_state.is_dark_mode:
            self.setStyleSheet("""
                QWidget          { background:#231811; color:white; }
                QTabWidget::pane { border:1px solid #5C4938; }
                QTabBar::tab     { background:#3A2D26; color:white;
                                   padding:4px 10px; margin-right:1px; }
                QTabBar::tab:selected { background:#D97757; }
                QTableWidget     { background:#231811; color:white;
                                   gridline-color:#5C4938;
                                   alternate-background-color:#3A2D26; }
                QTableWidget::item:selected, QAbstractItemView::item:selected {
                                   background:#FF9248; color:black; }
                QHeaderView::section { background:#8A4533; color:white;
                                       border:1px solid #5C4938; padding:2px; }
                QPushButton      { background:#3A2D26; color:white;
                                   border:1px solid #5C4938;
                                   padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; }
                QLabel           { color:white; }
                QFrame[frameShape="4"] { color:#5C4938; }
            """)
        else:
            self.setStyleSheet("""
                QWidget          { background:#f5f5f5; color:black; }
                QTabWidget::pane { border:1px solid #cccccc; }
                QTabBar::tab     { background:#e0e0e0; color:black;
                                   padding:4px 10px; margin-right:1px; }
                QTabBar::tab:selected { background:#D97757; color:white; }
                QTableWidget     { background:white; color:black;
                                   gridline-color:#dddddd;
                                   alternate-background-color:#FFF7F1; }
                QTableWidget::item:selected, QAbstractItemView::item:selected {
                                   background:#FF9248; color:black; }
                QHeaderView::section { background:#D97757; color:white;
                                       border:1px solid #cccccc;
                                       font-weight:bold; padding:2px; }
                QPushButton      { background:white; color:#A85C42;
                                   border:1px solid #aaaaaa;
                                   padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; color:white; }
                QLabel           { color:black; }
                QFrame[frameShape="4"] { color:#cccccc; }
            """)


class _CompletaDialog(QDialog):
    """Dialog for selecting PDF attachments before generating the complete documentation."""

    def __init__(self, parent, app_state, folder):
        super().__init__(parent)
        self._app_state = app_state
        self._folder    = folder
        self.attachments = []
        t = TRANSLATIONS[app_state.language]
        self.setWindowTitle(t.get("completa_dialog_title", "Documentazione Completa"))
        self.setMinimumWidth(520)
        self._build_ui(t)

    def _build_ui(self, t):
        from PySide6.QtWidgets import (QVBoxLayout, QHBoxLayout, QLabel,
                                        QListWidget, QPushButton)
        root = QVBoxLayout(self)
        root.setSpacing(8)
        root.setContentsMargins(12, 12, 12, 12)

        desc = QLabel(t.get("completa_dialog_desc",
                             "Seleziona i documenti PDF da allegare alla documentazione completa."))
        desc.setWordWrap(True)
        root.addWidget(desc)

        self._list = QListWidget()
        self._list.setMinimumHeight(150)
        root.addWidget(self._list)

        att_row = QHBoxLayout()
        btn_add = QPushButton(t.get("completa_btn_add", "Aggiungi allegato..."))
        btn_add.clicked.connect(self._add_attachment)
        att_row.addWidget(btn_add)
        btn_rem = QPushButton(t.get("completa_btn_remove", "Rimuovi"))
        btn_rem.clicked.connect(self._remove_attachment)
        att_row.addWidget(btn_rem)
        att_row.addStretch()
        root.addLayout(att_row)

        action_row = QHBoxLayout()
        action_row.addStretch()
        btn_gen = QPushButton(t.get("completa_btn_generate", "Genera"))
        btn_gen.setDefault(True)
        btn_gen.clicked.connect(self.accept)
        action_row.addWidget(btn_gen)
        btn_cancel = QPushButton(t.get("completa_btn_cancel", "Annulla"))
        btn_cancel.clicked.connect(self.reject)
        action_row.addWidget(btn_cancel)
        root.addLayout(action_row)

    def _add_attachment(self):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self._app_state.language]
        paths, _ = QFileDialog.getOpenFileNames(
            self, t.get("completa_btn_add", "Aggiungi allegato"),
            self._folder, "PDF Files (*.pdf)")
        for p in paths:
            if p not in self.attachments:
                self.attachments.append(p)
                self._list.addItem(os.path.basename(p))

    def _remove_attachment(self):
        row = self._list.currentRow()
        if row >= 0:
            self._list.takeItem(row)
            self.attachments.pop(row)


def _extract_pdf_links(pdf_path):
    """Return list of (src_page, x0, y0, x1, y1, target_page) in PDF points."""
    links = []
    try:
        from pypdf import PdfReader
        reader = PdfReader(pdf_path)
        page_id_map = {}
        for i, p in enumerate(reader.pages):
            if p.indirect_reference:
                page_id_map[p.indirect_reference.idnum] = i

        # Named destination → page index (handles ReportLab <a href="#key"> links)
        named_dest_pages = {}
        try:
            for name, dest_obj in reader.named_destinations.items():
                page_ref = getattr(dest_obj, 'page', None)
                if page_ref is not None and hasattr(page_ref, 'idnum'):
                    named_dest_pages[str(name)] = page_id_map.get(page_ref.idnum)
        except Exception:
            pass

        for src_idx, page in enumerate(reader.pages):
            annots = page.get('/Annots')
            if not annots:
                continue
            for ar in annots:
                try:
                    ann = reader.get_object(ar)
                    if ann.get('/Subtype') != '/Link':
                        continue
                    dest = ann.get('/Dest')
                    if dest is None:
                        act = ann.get('/A')
                        if act:
                            act = reader.get_object(act) if hasattr(act, 'idnum') else act
                            dest = act.get('/D') if act else None
                    if dest is None:
                        continue
                    if hasattr(dest, 'idnum'):
                        dest = reader.get_object(dest)

                    target_page = None
                    if isinstance(dest, (list, tuple)) and dest:
                        page_ref = dest[0]
                        if hasattr(page_ref, 'idnum'):
                            # Indirect object reference (ReportLab style)
                            target_page = page_id_map.get(page_ref.idnum)
                        else:
                            # Direct page index number (pypdf Link annotation style)
                            try:
                                target_page = int(page_ref)
                            except Exception:
                                pass
                    else:
                        # Named destination string (e.g. from <a href="#key">)
                        dest_str = str(dest).strip()
                        target_page = named_dest_pages.get(dest_str)

                    if target_page is None:
                        continue
                    rect = ann.get('/Rect')
                    if not rect:
                        continue
                    x0, y0, x1, y1 = float(rect[0]), float(rect[1]), float(rect[2]), float(rect[3])
                    if x0 > x1:
                        x0, x1 = x1, x0
                    if y0 > y1:
                        y0, y1 = y1, y0
                    links.append((src_idx, x0, y0, x1, y1, target_page))
                except Exception:
                    pass
    except Exception:
        pass
    return links


class _PdfLinkNavigator(QObject):
    """Event filter that intercepts mouse clicks on PDF link annotations in QPdfView."""

    _MARGIN = 6  # px gap between pages in MultiPage mode

    def __init__(self, view, doc, links):
        super().__init__(view)
        self._view = view
        self._doc = doc
        self._links = links
        view.viewport().installEventFilter(self)

    def eventFilter(self, obj, event):
        from PySide6.QtCore import QEvent
        t = event.type()
        if t == QEvent.Type.MouseButtonRelease:
            if self._handle_click(event.pos()):
                return True
        elif t == QEvent.Type.MouseMove:
            self._update_cursor(event.pos())
        return False

    def _abs_pos(self, screen_pos):
        v = self._view
        return (screen_pos.x() + v.horizontalScrollBar().value(),
                screen_pos.y() + v.verticalScrollBar().value())

    def _hit_test(self, abs_x, abs_y):
        doc = self._doc
        view = self._view
        zoom = view.zoomFactor()
        m = self._MARGIN
        vport_w = view.viewport().width()
        page_sizes = [doc.pagePointSize(i) for i in range(doc.pageCount())]
        max_w_px = max((ps.width() * zoom for ps in page_sizes), default=0)
        content_width = max_w_px + 2 * m
        layout_offset_x = max(0.0, (vport_w - content_width) / 2)
        x_layout = abs_x - layout_offset_x
        cumulative_y = m
        for page_idx, ps in enumerate(page_sizes):
            h_px = ps.height() * zoom
            w_px = ps.width() * zoom
            if abs_y < cumulative_y:
                return None
            if abs_y < cumulative_y + h_px:
                page_x = m + (max_w_px - w_px) / 2
                x_in_page = x_layout - page_x
                if x_in_page < 0 or x_in_page > w_px:
                    return None
                x_pts = x_in_page / zoom
                y_pts = ps.height() - (abs_y - cumulative_y) / zoom
                return (page_idx, x_pts, y_pts)
            cumulative_y += h_px + m
        return None

    def _find_link(self, page_idx, x_pts, y_pts, tol=4.0):
        for lp, x0, y0, x1, y1, target in self._links:
            if lp == page_idx:
                if (x0 - tol) <= x_pts <= (x1 + tol) and (y0 - tol) <= y_pts <= (y1 + tol):
                    return target
        return None

    def _handle_click(self, screen_pos):
        abs_x, abs_y = self._abs_pos(screen_pos)
        result = self._hit_test(abs_x, abs_y)
        if result is None:
            return False
        page_idx, x_pts, y_pts = result
        target = self._find_link(page_idx, x_pts, y_pts)
        if target is None:
            return False
        from PySide6.QtCore import QPointF
        self._view.pageNavigator().jump(target, QPointF(0, 0))
        return True

    def _update_cursor(self, screen_pos):
        from PySide6.QtCore import Qt as _Qt
        abs_x, abs_y = self._abs_pos(screen_pos)
        result = self._hit_test(abs_x, abs_y)
        if result and self._find_link(*result) is not None:
            self._view.viewport().setCursor(_Qt.CursorShape.PointingHandCursor)
        else:
            self._view.viewport().setCursor(_Qt.CursorShape.ArrowCursor)


class _PdfPreview(QWidget):
    """In-window PDF preview panel."""

    def __init__(self, pdf_path, app_state, on_close_cb, on_save_cb=None, on_excel_cb=None,
                 nav_items=None, regen_fn=None):
        super().__init__()
        self._pdf_path = pdf_path
        self._app_state = app_state
        self._on_close = on_close_cb
        self._on_save  = on_save_cb
        self._on_excel = on_excel_cb
        self._nav_items = nav_items or []
        self._regen_fn = regen_fn
        self._saved_folder = None
        self._pdf_view = None
        self._doc = None
        self._zoom_level = 1.0
        self._build_ui()
        self._apply_theme()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        t = TRANSLATIONS[self._app_state.language]
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 6, 8, 6)
        root.setSpacing(4)

        # Toolbar row
        toolbar = QHBoxLayout()
        toolbar.setSpacing(6)

        lbl = QLabel(os.path.basename(self._pdf_path))
        lbl.setStyleSheet("font-weight: bold; font-size: 10pt;")
        toolbar.addWidget(lbl)
        toolbar.addStretch()

        if self._on_save is not None:
            self._btn_save = QPushButton(t.get("btn_generate_pdf", "Genera PDF"))
            self._btn_save.setFixedHeight(26)
            self._btn_save.clicked.connect(self._on_save)
            toolbar.addWidget(self._btn_save)

        if self._on_excel is not None:
            self._btn_excel = QPushButton(t.get("btn_export_excel", "Esporta Excel"))
            self._btn_excel.setFixedHeight(26)
            self._btn_excel.clicked.connect(self._on_excel)
            toolbar.addWidget(self._btn_excel)

        self._btn_zoom_in = QPushButton(t.get("zoom_in", "Zoom +"))
        self._btn_zoom_in.setFixedHeight(26)
        self._btn_zoom_in.setFixedWidth(70)
        self._btn_zoom_in.clicked.connect(self._zoom_in)
        toolbar.addWidget(self._btn_zoom_in)

        self._btn_zoom_out = QPushButton(t.get("zoom_out", "Zoom −"))
        self._btn_zoom_out.setFixedHeight(26)
        self._btn_zoom_out.setFixedWidth(70)
        self._btn_zoom_out.clicked.connect(self._zoom_out)
        toolbar.addWidget(self._btn_zoom_out)

        self._btn_open_file = QPushButton(t.get("preview_open_file", "Apri PDF"))
        self._btn_open_file.setFixedHeight(26)
        self._btn_open_file.setToolTip(t.get("preview_open_file", "Apri PDF"))
        self._btn_open_file.clicked.connect(self._open_file)
        toolbar.addWidget(self._btn_open_file)

        self._btn_folder = QPushButton(t.get("preview_open_folder", "..."))
        self._btn_folder.setFixedHeight(26)
        self._btn_folder.setVisible(False)
        self._btn_folder.clicked.connect(self._open_folder)
        toolbar.addWidget(self._btn_folder)

        self._btn_close = QPushButton(t["preview_close"])
        self._btn_close.setFixedHeight(26)
        self._btn_close.clicked.connect(self._on_close)
        toolbar.addWidget(self._btn_close)

        root.addLayout(toolbar)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Sunken)
        root.addWidget(sep)

        if self._nav_items:
            from PySide6.QtWidgets import QSplitter, QListWidget, QListWidgetItem
            splitter = QSplitter(Qt.Horizontal)
            nav = QListWidget()
            nav.setMaximumWidth(220)
            nav.setMinimumWidth(150)
            for text, page_num in self._nav_items:
                item = QListWidgetItem(f"{text}  [{page_num}]")
                item.setData(Qt.UserRole, page_num)
                nav.addItem(item)
            self._nav_list = nav
            nav.itemClicked.connect(self._on_nav_clicked)
            splitter.addWidget(nav)
            splitter.addWidget(self._make_viewer(t))
            splitter.setStretchFactor(0, 0)
            splitter.setStretchFactor(1, 1)
            root.addWidget(splitter, 1)
        else:
            root.addWidget(self._make_viewer(t), 1)

    def _make_viewer(self, t):
        try:
            from PySide6.QtPdfWidgets import QPdfView
            from PySide6.QtPdf import QPdfDocument
            doc = QPdfDocument(self)
            doc.load(self._pdf_path)
            if doc.pageCount() <= 0:
                raise RuntimeError("PDF non leggibile o vuoto")
            self._doc = doc
            view = QPdfView(self)
            view.setDocument(doc)
            try:
                view.setPageMode(QPdfView.PageMode.MultiPage)
            except AttributeError:
                pass
            self._pdf_view = view
            try:
                links = _extract_pdf_links(self._pdf_path)
                if links:
                    self._link_navigator = _PdfLinkNavigator(view, doc, links)
            except Exception:
                pass
            return view
        except ImportError:
            logger.info("log_error_generic", "PySide6.QtPdfWidgets non disponibile — modalità fallback")
            return self._fallback_widget(t)
        except Exception as exc:
            logger.warning("log_error_generic", str(exc))
            return self._fallback_widget(t)

    def _fallback_widget(self, t):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setAlignment(Qt.AlignCenter)
        try:
            kb = os.path.getsize(self._pdf_path) / 1024
            size_str = f"{kb:.1f} KB"
        except Exception:
            size_str = ""
        info = t.get("preview_pdf_info", "PDF")
        msg = f"{info}\n{os.path.basename(self._pdf_path)}"
        if size_str:
            msg += f"  ({size_str})"
        lbl = QLabel(msg)
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setStyleSheet("font-size: 11pt;")
        lay.addWidget(lbl)
        btn_open = QPushButton(t.get("preview_open_file", "Apri PDF"))
        btn_open.setFixedHeight(32)
        btn_open.setFixedWidth(160)
        btn_open.clicked.connect(self._open_file)
        lay.addWidget(btn_open, alignment=Qt.AlignCenter)
        return w

    # ── Language update ───────────────────────────────────────────────────────

    def update_language(self, lang):
        t = TRANSLATIONS.get(lang, TRANSLATIONS["IT"])
        if self._on_save is not None and hasattr(self, '_btn_save'):
            self._btn_save.setText(t.get("btn_generate_pdf", "Genera PDF"))
        if self._on_excel is not None and hasattr(self, '_btn_excel'):
            self._btn_excel.setText(t.get("btn_export_excel", "Esporta Excel"))
        self._btn_zoom_in.setText(t.get("zoom_in", "Zoom +"))
        self._btn_zoom_out.setText(t.get("zoom_out", "Zoom −"))
        self._btn_open_file.setText(t.get("preview_open_file", "Apri PDF"))
        self._btn_folder.setText(t.get("preview_open_folder", "..."))
        self._btn_close.setText(t.get("preview_close", "Chiudi"))

    def _zoom_in(self):
        if self._pdf_view is not None:
            try:
                self._zoom_level = min(getattr(self, '_zoom_level', 1.0) + 0.25, 4.0)
                self._pdf_view.setZoomFactor(self._zoom_level)
            except Exception:
                pass

    def _zoom_out(self):
        if self._pdf_view is not None:
            try:
                self._zoom_level = max(getattr(self, '_zoom_level', 1.0) - 0.25, 0.25)
                self._pdf_view.setZoomFactor(self._zoom_level)
            except Exception:
                pass

    # ── Actions ───────────────────────────────────────────────────────────────

    def _on_nav_clicked(self, item):
        page_num = item.data(Qt.UserRole)
        if self._pdf_view is not None:
            try:
                from PySide6.QtCore import QPointF
                self._pdf_view.pageNavigator().jump(page_num - 1, QPointF(), 1.0)
            except Exception:
                pass

    def on_pdf_saved(self, folder_path):
        self._saved_folder = folder_path
        self._btn_folder.setVisible(True)
        # Open the destination folder automatically every time a PDF is saved,
        # so the user immediately sees the file they just generated.
        if folder_path:
            try:
                os.startfile(folder_path)
            except Exception as exc:
                logger.error("log_error_generic", str(exc))

    def regen(self):
        """Regenerate the PDF with the current language and reload the viewer."""
        if self._regen_fn is None:
            return
        try:
            result = self._regen_fn()
            if self._doc is not None:
                self._doc.close()
                self._doc.load(self._pdf_path)
            if result is not None and hasattr(self, '_nav_list'):
                self._nav_list.clear()
                for text, page_num in (result or []):
                    from PySide6.QtWidgets import QListWidgetItem
                    item = QListWidgetItem(f"{text}  [{page_num}]")
                    item.setData(Qt.UserRole, page_num)
                    self._nav_list.addItem(item)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def _open_folder(self):
        folder = self._saved_folder or os.path.dirname(os.path.abspath(self._pdf_path))
        try:
            os.startfile(folder)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def _open_file(self):
        try:
            os.startfile(self._pdf_path)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    # ── Theme ─────────────────────────────────────────────────────────────────

    def _apply_theme(self):
        if self._app_state.is_dark_mode:
            self.setStyleSheet("""
                QWidget          { background:#231811; color:white; }
                QPushButton      { background:#3A2D26; color:white;
                                   border:1px solid #5C4938;
                                   padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; }
                QLabel           { color:white; }
                QFrame[frameShape="4"] { color:#5C4938; }
            """)
        else:
            self.setStyleSheet("""
                QWidget          { background:#f5f5f5; color:black; }
                QPushButton      { background:white; color:#A85C42;
                                   border:1px solid #aaaaaa;
                                   padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; color:white; }
                QLabel           { color:black; }
                QFrame[frameShape="4"] { color:#cccccc; }
            """)


class _BackupView(QWidget):
    """In-window Backup Variables panel: tabs per variable type, Generate/Export buttons."""

    def __init__(self, tab_data, folder_name, app_state,
                 on_close_cb, on_generate_excel_cb, on_export_pdf_cb):
        super().__init__()
        self._tab_data = tab_data
        self._folder_name = folder_name
        self._app_state = app_state
        self._on_close = on_close_cb
        self._on_generate_excel = on_generate_excel_cb
        self._on_export_pdf = on_export_pdf_cb
        self._build_ui()
        self._apply_theme()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        t = TRANSLATIONS[self._app_state.language]
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 6, 8, 6)
        root.setSpacing(4)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(6)

        lbl = QLabel(self._folder_name or "Backup Variabili")
        lbl.setStyleSheet("font-weight: bold; font-size: 10pt;")
        toolbar.addWidget(lbl)
        toolbar.addStretch()

        self._btn_excel = QPushButton(t.get("btn_generate_excel", "Genera Excel"))
        self._btn_excel.setFixedHeight(26)
        self._btn_excel.clicked.connect(self._on_generate_excel)
        toolbar.addWidget(self._btn_excel)

        self._btn_pdf = QPushButton(t.get("btn_export_pdf", "Esporta PDF"))
        self._btn_pdf.setFixedHeight(26)
        self._btn_pdf.clicked.connect(self._on_export_pdf)
        toolbar.addWidget(self._btn_pdf)

        self._btn_close = QPushButton(t["preview_close"])
        self._btn_close.setFixedHeight(26)
        self._btn_close.clicked.connect(self._on_close)
        toolbar.addWidget(self._btn_close)

        root.addLayout(toolbar)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Sunken)
        root.addWidget(sep)

        self._tab_widget = QTabWidget()
        self._tab_widget.setDocumentMode(True)
        self._var_tbls = []

        total_vars = 0
        for tab_name, rows in self._tab_data.items():
            total_vars += len(rows)
            self._tab_widget.addTab(
                self._make_tab(t, rows),
                f"{tab_name} ({len(rows)})"
            )

        if not self._tab_data:
            empty_lbl = QLabel(t.get("backup_no_named_vars",
                                     "Nessuna variabile con nome trovata nel template.\n"
                                     "Clicca 'Genera Excel' per esportare tutti i valori."))
            empty_lbl.setAlignment(Qt.AlignCenter)
            empty_lbl.setStyleSheet("color: gray; font-size: 10pt; padding: 20px;")
            self._tab_widget.addTab(empty_lbl, "—")

        root.addWidget(self._tab_widget, 1)

        self._total_vars = total_vars
        self._lbl_footer = QLabel(
            t.get("backup_footer_vars", "Totale: {} variabili").format(total_vars))
        self._lbl_footer.setStyleSheet("color: gray; font-size: 8pt;")
        root.addWidget(self._lbl_footer)

    def _make_tab(self, t, rows):
        container = QWidget()
        lay = QVBoxLayout(container)
        lay.setContentsMargins(2, 4, 2, 2)
        lay.setSpacing(2)

        headers = ["ID",
                   t.get("backup_col_name", "Nome"),
                   t.get("backup_col_value", "Valore")]
        tbl = QTableWidget(len(rows), 3)
        tbl.setHorizontalHeaderLabels(headers)
        tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        tbl.setAlternatingRowColors(True)
        tbl.setSelectionBehavior(QTableWidget.SelectRows)
        tbl.verticalHeader().setVisible(False)
        tbl.horizontalHeader().setStretchLastSection(True)

        for r, (var_id, name, val) in enumerate(rows):
            tbl.setItem(r, 0, QTableWidgetItem(str(var_id) if var_id else ""))
            tbl.setItem(r, 1, QTableWidgetItem(name))
            tbl.setItem(r, 2, QTableWidgetItem(
                "" if val in (None, "") else str(val)
            ))

        tbl.resizeColumnToContents(0)
        tbl.resizeColumnToContents(1)
        self._var_tbls.append(tbl)
        lay.addWidget(tbl, 1)
        return container

    # ── Language update ───────────────────────────────────────────────────────

    def update_language(self, lang):
        t = TRANSLATIONS[lang]
        self._btn_excel.setText(t.get("btn_generate_excel", "Genera Excel"))
        self._btn_pdf.setText(t.get("btn_export_pdf", "Esporta PDF"))
        self._btn_close.setText(t.get("preview_close", "Chiudi"))
        self._lbl_footer.setText(
            t.get("backup_footer_vars", "Totale: {} variabili").format(self._total_vars))
        headers = ["ID",
                   t.get("backup_col_name",  "Nome"),
                   t.get("backup_col_value", "Valore")]
        for tbl in self._var_tbls:
            tbl.setHorizontalHeaderLabels(headers)

    # ── Theme ─────────────────────────────────────────────────────────────────

    def _apply_theme(self):
        if self._app_state.is_dark_mode:
            self.setStyleSheet("""
                QWidget          { background:#231811; color:white; }
                QTabWidget::pane { border:1px solid #5C4938; }
                QTabBar::tab     { background:#3A2D26; color:white;
                                   padding:4px 10px; margin-right:1px; }
                QTabBar::tab:selected { background:#D97757; }
                QTableWidget     { background:#231811; color:white;
                                   gridline-color:#5C4938;
                                   alternate-background-color:#3A2D26; }
                QTableWidget::item:selected, QAbstractItemView::item:selected {
                                   background:#FF9248; color:black; }
                QHeaderView::section { background:#8A4533; color:white;
                                       border:1px solid #5C4938; padding:2px; }
                QPushButton      { background:#3A2D26; color:white;
                                   border:1px solid #5C4938;
                                   padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; }
                QLabel           { color:white; }
                QFrame[frameShape="4"] { color:#5C4938; }
            """)
        else:
            self.setStyleSheet("""
                QWidget          { background:#f5f5f5; color:black; }
                QTabWidget::pane { border:1px solid #cccccc; }
                QTabBar::tab     { background:#e0e0e0; color:black;
                                   padding:4px 10px; margin-right:1px; }
                QTabBar::tab:selected { background:#D97757; color:white; }
                QTableWidget     { background:white; color:black;
                                   gridline-color:#dddddd;
                                   alternate-background-color:#FFF7F1; }
                QTableWidget::item:selected, QAbstractItemView::item:selected {
                                   background:#FF9248; color:black; }
                QHeaderView::section { background:#D97757; color:white;
                                       border:1px solid #cccccc;
                                       font-weight:bold; padding:2px; }
                QPushButton      { background:white; color:#A85C42;
                                   border:1px solid #aaaaaa;
                                   padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; color:white; }
                QLabel           { color:black; }
                QFrame[frameShape="4"] { color:#cccccc; }
            """)


class _DriveView(QWidget):
    """In-window MotionDrive panel: Info tab + category tabs, Generate/Export/Close buttons."""

    def __init__(self, file_name, info, params_by_cat, app_state,
                 on_close_cb, on_generate_excel_cb, on_export_pdf_cb):
        super().__init__()
        self._file_name = file_name
        self._info = info
        self._params_by_cat = params_by_cat
        self._app_state = app_state
        self._on_close = on_close_cb
        self._on_generate_excel = on_generate_excel_cb
        self._on_export_pdf = on_export_pdf_cb
        self._build_ui()
        self._apply_theme()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        from docs.drive import _CAT_ORDER, _CATEGORIES
        t = TRANSLATIONS[self._app_state.language]
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 6, 8, 6)
        root.setSpacing(4)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(6)

        lbl = QLabel(self._file_name or "MotionDrive")
        lbl.setStyleSheet("font-weight: bold; font-size: 10pt;")
        toolbar.addWidget(lbl)
        toolbar.addStretch()

        self._btn_excel = QPushButton(t.get("btn_generate_excel", "Genera Excel"))
        self._btn_excel.setFixedHeight(26)
        self._btn_excel.clicked.connect(self._on_generate_excel)
        toolbar.addWidget(self._btn_excel)

        self._btn_pdf = QPushButton(t.get("btn_export_pdf", "Esporta PDF"))
        self._btn_pdf.setFixedHeight(26)
        self._btn_pdf.clicked.connect(self._on_export_pdf)
        toolbar.addWidget(self._btn_pdf)

        self._btn_close = QPushButton(t["preview_close"])
        self._btn_close.setFixedHeight(26)
        self._btn_close.clicked.connect(self._on_close)
        toolbar.addWidget(self._btn_close)

        root.addLayout(toolbar)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Sunken)
        root.addWidget(sep)

        self._tab_widget = QTabWidget()
        self._tab_widget.setDocumentMode(True)
        self._cat_tbls = []

        self._tab_widget.addTab(self._make_info_tab(t), t.get("drive_tab_info", "Info"))

        for cat in _CAT_ORDER:
            if cat not in self._params_by_cat:
                continue
            rows = self._params_by_cat[cat]
            label = _CATEGORIES.get(cat, cat)
            self._tab_widget.addTab(self._make_cat_tab(t, rows), f"{label} ({len(rows)})")

        root.addWidget(self._tab_widget, 1)

        self._total_params = sum(len(v) for v in self._params_by_cat.values())
        self._lbl_footer = QLabel(
            t.get("drive_footer_params", "Totale: {} parametri").format(self._total_params))
        self._lbl_footer.setStyleSheet("color: gray; font-size: 8pt;")
        root.addWidget(self._lbl_footer)

    def _make_info_tab(self, t):
        container = QWidget()
        lay = QVBoxLayout(container)
        lay.setContentsMargins(4, 4, 4, 4)
        lay.setSpacing(4)

        meta_fields = [
            (t.get("drive_lbl_project_name", 'Project Name'),    'PROJECT NAME'),
            (t.get("drive_lbl_user",          'User'),            'USER NAME'),
            (t.get("drive_lbl_drive_series",  'Drive Series'),    'DRIVE SERIES'),
            (t.get("drive_lbl_drive_model",   'Drive Model'),     'DRIVE MODEL'),
            (t.get("drive_lbl_sw_version",    'Software Version'),'SOFTWARE VERSION'),
            (t.get("drive_lbl_ctrl_method",   'Control Method'),  'CONTROL METHOD'),
            (t.get("drive_lbl_init_mode",     'Init. Mode'),      'INITIALIZATION MODE'),
            (t.get("drive_lbl_db_info",       'DB Info'),         'DBINFO'),
            (t.get("drive_lbl_db_name",       'DB Name'),         'DBNAME'),
        ]
        meta_rows = [(label, self._info.get(key, ''))
                     for label, key in meta_fields
                     if self._info.get(key, '').strip()]

        if meta_rows:
            tbl = QTableWidget(len(meta_rows), 2)
            tbl.horizontalHeader().setVisible(False)
            tbl.setEditTriggers(QTableWidget.NoEditTriggers)
            tbl.setAlternatingRowColors(True)
            tbl.setSelectionBehavior(QTableWidget.SelectRows)
            tbl.verticalHeader().setVisible(False)
            tbl.horizontalHeader().setStretchLastSection(True)
            for r, (label, val) in enumerate(meta_rows):
                item_l = QTableWidgetItem(label)
                item_l.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                tbl.setItem(r, 0, item_l)
                tbl.setItem(r, 1, QTableWidgetItem(val))
            tbl.resizeColumnToContents(0)
            lay.addWidget(tbl)

        file_info = self._info.get('file_info', '').strip()
        if file_info:
            fi_lbl = QLabel(t.get("drive_lbl_file_info", "File Info:"))
            fi_lbl.setStyleSheet("font-weight: bold;")
            lay.addWidget(fi_lbl)
            fi_text = QTextEdit()
            fi_text.setReadOnly(True)
            fi_text.setPlainText(file_info)
            fi_text.setMaximumHeight(120)
            lay.addWidget(fi_text)

        lay.addStretch()
        return container

    def _make_cat_tab(self, t, rows):
        container = QWidget()
        lay = QVBoxLayout(container)
        lay.setContentsMargins(2, 4, 2, 2)
        lay.setSpacing(2)

        from docs.ga500_params import GA500_PARAMS, get_param_name as _gpn
        lang = self._app_state.language
        headers = [
            t.get("drive_col_param",   "Param"),
            t.get("drive_col_name",    "Name"),
            t.get("drive_col_value",   "Value"),
            t.get("drive_col_default", "Default"),
            t.get("drive_col_min",     "Min"),
            t.get("drive_col_max",     "Max"),
            t.get("drive_col_prev",    "Previous"),
        ]
        tbl = QTableWidget(len(rows), 7)
        tbl.setHorizontalHeaderLabels(headers)
        tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        tbl.setAlternatingRowColors(True)
        tbl.setSelectionBehavior(QTableWidget.SelectRows)
        tbl.verticalHeader().setVisible(False)
        tbl.horizontalHeader().setStretchLastSection(True)

        for r, (code, val, dflt, mn, mx, prv) in enumerate(rows):
            tbl.setItem(r, 0, QTableWidgetItem(str(code)))
            tbl.setItem(r, 1, QTableWidgetItem(_gpn(code, lang)))
            tbl.setItem(r, 2, QTableWidgetItem(str(val)))
            tbl.setItem(r, 3, QTableWidgetItem(str(dflt)))
            tbl.setItem(r, 4, QTableWidgetItem(str(mn)))
            tbl.setItem(r, 5, QTableWidgetItem(str(mx)))
            tbl.setItem(r, 6, QTableWidgetItem(str(prv)))

        tbl.resizeColumnToContents(0)
        self._cat_tbls.append(tbl)
        lay.addWidget(tbl, 1)
        return container

    # ── Language update ───────────────────────────────────────────────────────

    def update_language(self, lang):
        t = TRANSLATIONS[lang]
        self._btn_excel.setText(t.get("btn_generate_excel", "Genera Excel"))
        self._btn_pdf.setText(t.get("btn_export_pdf", "Esporta PDF"))
        self._btn_close.setText(t.get("preview_close", "Chiudi"))
        self._lbl_footer.setText(
            t.get("drive_footer_params", "Totale: {} parametri").format(self._total_params))
        drive_headers = [
            t.get("drive_col_param",   "Param"),
            t.get("drive_col_name",    "Name"),
            t.get("drive_col_value",   "Value"),
            t.get("drive_col_default", "Default"),
            t.get("drive_col_min",     "Min"),
            t.get("drive_col_max",     "Max"),
            t.get("drive_col_prev",    "Previous"),
        ]
        for tbl in self._cat_tbls:
            tbl.setHorizontalHeaderLabels(drive_headers)

    # ── Theme ─────────────────────────────────────────────────────────────────

    def _apply_theme(self):
        if self._app_state.is_dark_mode:
            self.setStyleSheet("""
                QWidget          { background:#231811; color:white; }
                QTabWidget::pane { border:1px solid #5C4938; }
                QTabBar::tab     { background:#3A2D26; color:white;
                                   padding:4px 10px; margin-right:1px; }
                QTabBar::tab:selected { background:#D97757; }
                QTableWidget     { background:#231811; color:white;
                                   gridline-color:#5C4938;
                                   alternate-background-color:#3A2D26; }
                QTableWidget::item:selected, QAbstractItemView::item:selected {
                                   background:#FF9248; color:black; }
                QHeaderView::section { background:#8A4533; color:white;
                                       border:1px solid #5C4938; padding:2px; }
                QPushButton      { background:#3A2D26; color:white;
                                   border:1px solid #5C4938;
                                   padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; }
                QLabel           { color:white; }
                QFrame[frameShape="4"] { color:#5C4938; }
                QTextEdit        { background:#231811; color:white;
                                   border:1px solid #5C4938; }
            """)
        else:
            self.setStyleSheet("""
                QWidget          { background:#f5f5f5; color:black; }
                QTabWidget::pane { border:1px solid #cccccc; }
                QTabBar::tab     { background:#e0e0e0; color:black;
                                   padding:4px 10px; margin-right:1px; }
                QTabBar::tab:selected { background:#D97757; color:white; }
                QTableWidget     { background:white; color:black;
                                   gridline-color:#dddddd;
                                   alternate-background-color:#FFF7F1; }
                QTableWidget::item:selected, QAbstractItemView::item:selected {
                                   background:#FF9248; color:black; }
                QHeaderView::section { background:#D97757; color:white;
                                       border:1px solid #cccccc;
                                       font-weight:bold; padding:2px; }
                QPushButton      { background:white; color:#A85C42;
                                   border:1px solid #aaaaaa;
                                   padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; color:white; }
                QLabel           { color:black; }
                QFrame[frameShape="4"] { color:#cccccc; }
                QTextEdit        { background:white; color:black;
                                   border:1px solid #cccccc; }
            """)


class _IpNetView(QWidget):
    """In-window network configuration viewer (STEP 8)."""

    def __init__(self, rows, app_state, on_close_cb):
        super().__init__()
        self._app_state = app_state
        self._on_close = on_close_cb
        self._rows = rows
        self._build_ui()
        self._apply_theme()

    def _build_ui(self):
        t = TRANSLATIONS[self._app_state.language]
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 6, 8, 6)
        root.setSpacing(4)

        toolbar = QHBoxLayout()
        self._lbl_title = QLabel(t.get("ipnet_view_title", "Configurazione Rete — YASKAWA YRC1000"))
        self._lbl_title.setStyleSheet("font-weight: bold; font-size: 10pt;")
        toolbar.addWidget(self._lbl_title)
        toolbar.addStretch()
        self._btn_close = QPushButton(t.get("preview_close", "Chiudi"))
        self._btn_close.setFixedHeight(26)
        self._btn_close.clicked.connect(self._on_close)
        toolbar.addWidget(self._btn_close)
        root.addLayout(toolbar)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Sunken)
        root.addWidget(sep)

        self._tbl = QTableWidget(len(self._rows), 3)
        col_sec   = t.get("ipnet_col_section", "Sezione")
        col_param = t.get("ipnet_col_param", "Parametro")
        col_val   = t.get("ipnet_col_value", "Valore")
        self._tbl.setHorizontalHeaderLabels([col_sec, col_param, col_val])
        self._tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.horizontalHeader().setStretchLastSection(True)
        for r, (section, param, value) in enumerate(self._rows):
            self._tbl.setItem(r, 0, QTableWidgetItem(section))
            self._tbl.setItem(r, 1, QTableWidgetItem(param))
            self._tbl.setItem(r, 2, QTableWidgetItem(value))
        self._tbl.resizeColumnsToContents()
        root.addWidget(self._tbl, 1)

        if not self._rows:
            root.addWidget(QLabel(t.get("ipnet_no_data", "Nessun dato rete")))

    def update_language(self, lang):
        t = TRANSLATIONS.get(lang, TRANSLATIONS["IT"])
        self._lbl_title.setText(t.get("ipnet_view_title", "Configurazione Rete — YASKAWA YRC1000"))
        self._btn_close.setText(t.get("preview_close", "Chiudi"))
        col_s = t.get("ipnet_col_section", "Sezione")
        col_p = t.get("ipnet_col_param",   "Parametro")
        col_v = t.get("ipnet_col_value",   "Valore")
        self._tbl.setHorizontalHeaderItem(0, QTableWidgetItem(col_s))
        self._tbl.setHorizontalHeaderItem(1, QTableWidgetItem(col_p))
        self._tbl.setHorizontalHeaderItem(2, QTableWidgetItem(col_v))

    def _apply_theme(self):
        if self._app_state.is_dark_mode:
            self.setStyleSheet("""
                QWidget { background:#231811; color:white; }
                QTableWidget { background:#231811; color:white; gridline-color:#5C4938;
                               alternate-background-color:#3A2D26; }
                QTableWidget::item:selected, QAbstractItemView::item:selected { background:#FF9248; color:black; }
                QHeaderView::section { background:#8A4533; color:white;
                                       border:1px solid #5C4938; padding:2px; }
                QPushButton { background:#3A2D26; color:white; border:1px solid #5C4938;
                              padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; }
                QLabel { color:white; }
                QFrame[frameShape="4"] { color:#5C4938; }
            """)
        else:
            self.setStyleSheet("""
                QWidget { background:#f5f5f5; color:black; }
                QTableWidget { background:white; color:black; gridline-color:#dddddd;
                               alternate-background-color:#FFF7F1; }
                QTableWidget::item:selected, QAbstractItemView::item:selected { background:#FF9248; color:black; }
                QHeaderView::section { background:#D97757; color:white;
                                       border:1px solid #cccccc; font-weight:bold; padding:2px; }
                QPushButton { background:white; color:#A85C42; border:1px solid #aaaaaa;
                              padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; color:white; }
                QLabel { color:black; }
                QFrame[frameShape="4"] { color:#cccccc; }
            """)


class _CompilaView(QWidget):
    """Editable variable table from robot backup folder (STEP 5+6)."""

    _APPDATA_FILE = None

    @staticmethod
    def _get_appdata_path():
        appdata = os.environ.get('APPDATA', os.path.expanduser('~'))
        d = os.path.join(appdata, 'YaskawaTools')
        os.makedirs(d, exist_ok=True)
        return os.path.join(d, 'compila_data.json')

    def __init__(self, folder, app_state, on_close_cb):
        super().__init__()
        self._folder = folder
        self._app_state = app_state
        self._on_close = on_close_cb
        self._var_data = {}   # {tab_name: [(id, name, vtype, value), ...]}
        self._load_folder()
        self._build_ui()
        self._apply_theme()

    # ── Parse robot folder ────────────────────────────────────────────────────

    def _load_folder(self):
        import json
        # Load persisted values
        persisted = {}
        try:
            with open(self._get_appdata_path(), 'r', encoding='utf-8') as f:
                persisted = json.load(f)
        except Exception:
            pass

        folder = self._folder
        var_dat  = os.path.join(folder, 'VAR.DAT')
        varname  = os.path.join(folder, 'VARNAME.DAT')
        ioname   = os.path.join(folder, 'IONAME.DAT')
        exioname = os.path.join(folder, 'EXIONAME.DAT')

        counts = self._parse_var_counts(var_dat)
        var_values = self._parse_var_dat(var_dat)
        var_names  = self._parse_varname_dat(varname, counts)
        io_names, io_values = self._parse_ioname_dat(ioname)
        exio_names, exio_values = self._parse_ioname_dat(exioname)

        tabs = {}
        for vtype in ('B', 'I', 'D', 'R', 'S'):
            rows = []
            vals = var_values.get(vtype, [])
            names = var_names.get(vtype, [])
            n = counts.get(vtype, len(vals))
            for i in range(n):
                vid = i + 1
                vname = names[i] if i < len(names) else ''
                vval = persisted.get(f'{vtype}_{vid}',
                       vals[i] if i < len(vals) else '')
                rows.append((vid, vname, vtype, str(vval)))
            if rows:
                tabs[vtype] = rows

        for tag, label, src_names, src_vals, readonly in [
            ('IN',    'IN',    io_names.get('IN', []),    io_values.get('IN', []),    True),
            ('OUT',   'OUT',   io_names.get('OUT', []),   io_values.get('OUT', []),   False),
            ('EXIN',  'EXIN',  exio_names.get('IN', []),  exio_values.get('IN', []),  True),
            ('EXOUT', 'EXOUT', exio_names.get('OUT', []), exio_values.get('OUT', []), False),
        ]:
            rows = []
            for i, name in enumerate(src_names):
                vid = i + 1
                vval = persisted.get(f'{tag}_{vid}',
                       src_vals[i] if i < len(src_vals) else '0')
                rows.append((vid, name, tag, str(vval), readonly))
            if rows:
                tabs[label] = rows

        self._var_data = tabs

    @staticmethod
    def _parse_var_counts(filepath):
        counts = {}
        try:
            with open(filepath, 'r', encoding='latin-1', errors='replace') as f:
                for line in f:
                    line = line.strip()
                    if line.startswith('///SHARE'):
                        parts = line.split()
                        if len(parts) >= 2:
                            nums = parts[1].split(',')
                            keys = ['B', 'I', 'D', 'R', 'S']
                            for k, v in zip(keys, nums):
                                try:
                                    counts[k] = int(v)
                                except ValueError:
                                    pass
                        break
        except OSError:
            pass
        return counts

    @staticmethod
    def _parse_var_dat(filepath):
        result = {}
        try:
            with open(filepath, 'r', encoding='latin-1', errors='replace') as f:
                lines = f.readlines()
        except OSError:
            return result
        current = None
        for raw in lines:
            line = raw.strip()
            if line in ('///B', '///I', '///D', '///R', '///S'):
                current = line[3:]
                result[current] = []
            elif current and line and not line.startswith('//'):
                vals = line.split(',')
                for v in vals:
                    v = v.strip()
                    if v:
                        result[current].append(v)
        return result

    @staticmethod
    def _parse_varname_dat(filepath, counts):
        result = {}
        if not os.path.isfile(filepath):
            return result
        try:
            with open(filepath, 'r', encoding='latin-1', errors='replace') as f:
                lines = f.readlines()
        except OSError:
            return result
        current = None
        for raw in lines:
            line = raw.rstrip('\r\n')
            stripped = line.strip()
            if stripped.startswith('///') and not stripped.startswith('////'):
                current = stripped[3:].strip()
                if current not in result:
                    result[current] = []
            elif current and stripped and not stripped.startswith('//'):
                result[current].append(stripped)
        return result

    @staticmethod
    def _parse_ioname_dat(filepath):
        """Returns (names_dict, values_dict) keyed by 'IN'/'OUT'."""
        names = {}
        values = {}
        if not os.path.isfile(filepath):
            return names, values
        try:
            with open(filepath, 'r', encoding='latin-1', errors='replace') as f:
                lines = f.readlines()
        except OSError:
            return names, values
        current = None
        for raw in lines:
            line = raw.rstrip('\r\n')
            stripped = line.strip()
            if stripped in ('//IN', '//OUT', '//EXIN', '//EXOUT'):
                current = stripped[2:]
                names[current] = []
                values[current] = []
            elif current and stripped and not stripped.startswith('//'):
                parts = [p.strip() for p in stripped.split(',')]
                for i, p in enumerate(parts):
                    if i < 4:
                        names[current].append(p)
                        values[current].append('0')
        return names, values

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        t = TRANSLATIONS[self._app_state.language]
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 6, 8, 6)
        root.setSpacing(4)

        # Toolbar
        toolbar = QHBoxLayout()
        toolbar.setSpacing(6)
        self._lbl_title = QLabel(t.get("compila_view_title", "Compila — YASKAWA"))
        self._lbl_title.setStyleSheet("font-weight: bold; font-size: 10pt;")
        toolbar.addWidget(self._lbl_title)
        toolbar.addStretch()
        self._btn_clear = QPushButton(t.get("compila_btn_pulisci", "Pulisci"))
        self._btn_clear.setFixedHeight(26)
        self._btn_clear.clicked.connect(self._on_clear)
        toolbar.addWidget(self._btn_clear)
        self._btn_write = QPushButton(t.get("compila_btn_scrivi", "Scrivi DAT"))
        self._btn_write.setFixedHeight(26)
        self._btn_write.clicked.connect(self._on_write)
        toolbar.addWidget(self._btn_write)
        self._btn_close = QPushButton(t.get("preview_close", "Chiudi"))
        self._btn_close.setFixedHeight(26)
        self._btn_close.clicked.connect(self._on_close)
        toolbar.addWidget(self._btn_close)
        root.addLayout(toolbar)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Sunken)
        root.addWidget(sep)

        col_id    = t.get("compila_col_id",    "ID")
        col_name  = t.get("compila_col_name",  "Nome")
        col_val   = t.get("compila_col_value", "Valore")
        col_desc  = t.get("compila_col_desc",  "Descrizione")

        tabs = QTabWidget()
        tabs.setDocumentMode(True)
        self._tables = {}

        for tab_key, rows in self._var_data.items():
            tbl = QTableWidget(len(rows), 4)
            tbl.setHorizontalHeaderLabels([col_id, col_name, col_val, col_desc])
            tbl.setAlternatingRowColors(True)
            tbl.verticalHeader().setVisible(False)
            tbl.horizontalHeader().setStretchLastSection(True)
            tbl.setSelectionBehavior(QTableWidget.SelectRows)
            readonly_tab = tab_key in ('IN', 'EXIN')
            for r, row in enumerate(rows):
                if len(row) == 5:
                    vid, vname, vtype, vval, ro = row
                else:
                    vid, vname, vtype, vval = row
                    ro = readonly_tab
                # ID: type prefix + zero-padded index (0-based)
                id_label = f"{tab_key}{r:04d}"
                id_item = QTableWidgetItem(id_label)
                id_item.setFlags(id_item.flags() & ~Qt.ItemIsEditable)
                tbl.setItem(r, 0, id_item)
                name_item = QTableWidgetItem(str(vname))
                name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
                tbl.setItem(r, 1, name_item)
                val_item = QTableWidgetItem(str(vval))
                if ro:
                    val_item.setFlags(val_item.flags() & ~Qt.ItemIsEditable)
                tbl.setItem(r, 2, val_item)
                tbl.setItem(r, 3, QTableWidgetItem(""))  # Descrizione: free text
            tbl.resizeColumnToContents(0)
            tbl.resizeColumnToContents(1)
            tbl.setColumnWidth(2, 90)
            tabs.addTab(tbl, tab_key)
            self._tables[tab_key] = tbl

        root.addWidget(tabs, 1)

    def _collect_values(self):
        data = {}
        for tab_key, tbl in self._tables.items():
            for r in range(tbl.rowCount()):
                val_item = tbl.item(r, 2)   # Valore is column 2
                if val_item:
                    data[f'{tab_key}_{r + 1}'] = val_item.text()
        return data

    def _persist(self):
        import json
        try:
            with open(self._get_appdata_path(), 'w', encoding='utf-8') as f:
                json.dump(self._collect_values(), f)
        except Exception:
            pass

    def _on_clear(self):
        for tab_key, tbl in self._tables.items():
            ro_tab = tab_key in ('IN', 'EXIN')
            for r in range(tbl.rowCount()):
                it = tbl.item(r, 2)
                if it and not ro_tab:
                    it.setText('0' if tab_key in ('B', 'OUT', 'EXOUT') else '')
        self._persist()
        logger.info("log_compila_cleared")

    def _on_write(self):
        self._persist()
        var_path = os.path.join(self._folder, 'VAR.DAT')
        if not os.path.isfile(var_path):
            logger.warning("compila_no_var_dat")
            return
        ok, err = self._write_var_dat(var_path)
        if ok:
            logger.info("log_compila_saved", self._folder)
        else:
            logger.error("log_error_generic", str(err))

    def _write_var_dat(self, filepath):
        """Rebuild VAR.DAT with table values, preserving header and unknown sections."""
        try:
            with open(filepath, 'r', encoding='latin-1', errors='replace') as f:
                original = f.readlines()
        except OSError as exc:
            return False, str(exc)

        # Collect per-type values from the editable tables
        section_vals = {}
        for vtype in ('B', 'I', 'D', 'R', 'S'):
            tbl = self._tables.get(vtype)
            if tbl is None:
                continue
            vals = []
            for r in range(tbl.rowCount()):
                it = tbl.item(r, 2)   # Valore column is now index 2
                vals.append(it.text().strip() if it else '0')
            section_vals[vtype] = vals

        out_lines = []
        current_section = None
        in_handled = False

        for raw in original:
            line = raw.rstrip('\r\n')
            stripped = line.strip()

            if stripped.startswith('///'):
                tag = stripped[3:].strip()
                # tag may be "B", "I", "S", "SHARE 100,...", etc.
                simple_tag = tag.split()[0] if tag else ''
                current_section = simple_tag
                in_handled = simple_tag in section_vals
                out_lines.append(raw)
                if in_handled:
                    vals = section_vals[simple_tag]
                    if simple_tag in ('B', 'I', 'D', 'R'):
                        for i in range(0, len(vals), 10):
                            chunk = vals[i:i + 10]
                            out_lines.append(','.join(chunk) + '\r\n')
                    else:  # S
                        for v in vals:
                            out_lines.append(v + '\r\n')
            elif in_handled:
                pass  # skip original data — already written above
            else:
                out_lines.append(raw)

        try:
            with open(filepath, 'w', encoding='latin-1', newline='') as f:
                f.writelines(out_lines)
            return True, None
        except OSError as exc:
            return False, str(exc)

    def update_language(self, lang):
        t = TRANSLATIONS.get(lang, TRANSLATIONS["IT"])
        self._lbl_title.setText(t.get("compila_view_title", "Compila — YASKAWA"))
        self._btn_clear.setText(t.get("compila_btn_pulisci", "Pulisci"))
        self._btn_write.setText(t.get("compila_btn_scrivi", "Scrivi DAT"))
        self._btn_close.setText(t.get("preview_close", "Chiudi"))
        headers = [
            t.get("compila_col_id",    "ID"),
            t.get("compila_col_name",  "Nome"),
            t.get("compila_col_value", "Valore"),
            t.get("compila_col_desc",  "Descrizione"),
        ]
        for tbl in self._tables.values():
            tbl.setHorizontalHeaderLabels(headers)

    def _apply_theme(self):
        if self._app_state.is_dark_mode:
            self.setStyleSheet("""
                QWidget { background:#231811; color:white; }
                QTableWidget { background:#231811; color:white; gridline-color:#5C4938;
                               alternate-background-color:#3A2D26; }
                QTableWidget::item:selected, QAbstractItemView::item:selected { background:#FF9248; color:black; }
                QHeaderView::section { background:#8A4533; color:white;
                                       border:1px solid #5C4938; padding:2px; }
                QPushButton { background:#3A2D26; color:white; border:1px solid #5C4938;
                              padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; }
                QLabel { color:white; }
                QFrame[frameShape="4"] { color:#5C4938; }
                QTabWidget::pane { border:1px solid #5C4938; }
                QTabBar::tab { background:#3A2D26; color:white; padding:4px 12px; }
                QTabBar::tab:selected { background:#D97757; }
            """)
        else:
            self.setStyleSheet("""
                QWidget { background:#f5f5f5; color:black; }
                QTableWidget { background:white; color:black; gridline-color:#dddddd;
                               alternate-background-color:#FFF7F1; }
                QTableWidget::item:selected, QAbstractItemView::item:selected { background:#FF9248; color:black; }
                QHeaderView::section { background:#D97757; color:white;
                                       border:1px solid #cccccc; font-weight:bold; padding:2px; }
                QPushButton { background:white; color:#A85C42; border:1px solid #aaaaaa;
                              padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; color:white; }
                QLabel { color:black; }
                QFrame[frameShape="4"] { color:#cccccc; }
                QTabWidget::pane { border:1px solid #cccccc; }
                QTabBar::tab { background:#e8e8e8; color:#333333; padding:4px 12px; }
                QTabBar::tab:selected { background:#D97757; color:white; }
            """)


class _GA500ParamsView(QWidget):
    """In-window GA500 parameter reference: code+name table, printable to PDF."""

    def __init__(self, app_state, on_close_cb):
        super().__init__()
        self._app_state = app_state
        self._on_close = on_close_cb
        self._build_ui()
        self._apply_theme()

    def _build_ui(self):
        from docs.ga500_params import GA500_PARAMS
        t = TRANSLATIONS[self._app_state.language]
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 6, 8, 6)
        root.setSpacing(4)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(6)

        self._lbl_title = QLabel(t.get("ga500_view_title", "GA500 Parameters — YASKAWA"))
        self._lbl_title.setStyleSheet("font-weight: bold; font-size: 10pt;")
        toolbar.addWidget(self._lbl_title)
        toolbar.addStretch()

        self._btn_pdf = QPushButton(t.get("btn_generate_pdf", "Genera PDF"))
        self._btn_pdf.setFixedHeight(26)
        self._btn_pdf.clicked.connect(self._on_generate_pdf)
        toolbar.addWidget(self._btn_pdf)

        self._btn_close = QPushButton(t["preview_close"])
        self._btn_close.setFixedHeight(26)
        self._btn_close.clicked.connect(self._on_close)
        toolbar.addWidget(self._btn_close)

        root.addLayout(toolbar)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Sunken)
        root.addWidget(sep)

        col_code = t.get("ga500_col_code", "Code")
        col_name = t.get("ga500_col_name", "Name")

        self._tbl = QTableWidget(len(GA500_PARAMS), 2)
        self._tbl.setHorizontalHeaderLabels([col_code, col_name])
        self._tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.horizontalHeader().setStretchLastSection(True)

        lang = self._app_state.language
        from docs.ga500_params import get_param_name as _gpn
        for r, code in enumerate(sorted(GA500_PARAMS.keys())):
            self._tbl.setItem(r, 0, QTableWidgetItem(code))
            self._tbl.setItem(r, 1, QTableWidgetItem(_gpn(code, lang)))

        self._tbl.resizeColumnToContents(0)
        root.addWidget(self._tbl, 1)

        self._ga500_count = len(GA500_PARAMS)
        self._lbl_footer = QLabel(
            t.get("ga500_footer_params", "{} parametri").format(self._ga500_count))
        self._lbl_footer.setStyleSheet("color: gray; font-size: 8pt;")
        root.addWidget(self._lbl_footer)

    def _on_generate_pdf(self):
        import tempfile, shutil
        from PySide6.QtWidgets import QFileDialog
        from docs.ga500_params import GA500_PARAMS
        t = TRANSLATIONS[self._app_state.language]

        logger.info("log_btn_pressed", t.get("btn_generate_pdf", "Genera PDF"))

        save_path, _ = QFileDialog.getSaveFileName(
            self, t["dialog_save_pdf"],
            os.path.join(os.path.expanduser("~"), "Desktop", "GA500_Parameters.pdf"),
            "PDF Files (*.pdf)"
        )
        if not save_path:
            logger.info("log_cancelled", t.get("ga500_view_title", "GA500"))
            return

        tmp_path = secure_paths.temp_path("GA500_Parameters.pdf")
        try:
            ok = self._generate_ga500_pdf(GA500_PARAMS, tmp_path, t,
                                          lang=self._app_state.language)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        if not ok:
            return
        try:
            shutil.copy2(tmp_path, save_path)
            os.remove(tmp_path)
            logger.info("log_ga500_pdf_saved", save_path)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return
        try:
            os.startfile(os.path.dirname(os.path.abspath(save_path)))
        except Exception as exc:
            logger.warning("log_error_generic", str(exc))

    @staticmethod
    def _generate_ga500_pdf(params, output_path, t, lang='EN'):
        try:
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                             Paragraph, Spacer, PageBreak, Flowable)
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.lib.colors import HexColor, white
        except ImportError as exc:
            logger.error("log_error_generic", f"reportlab: {exc}")
            return False

        from docs.drive import _CAT_ORDER, _CATEGORIES

        accent = HexColor('#D97757')
        row_gray  = HexColor('#EFEFEF')
        styles    = getSampleStyleSheet()

        title_s = ParagraphStyle('G5Title', parent=styles['Normal'],
            fontSize=13, fontName='Helvetica-Bold', textColor=HexColor('#A85C42'),
            spaceAfter=4*mm)
        cat_s = ParagraphStyle('G5Cat', parent=styles['Normal'],
            fontSize=10, fontName='Helvetica-Bold', textColor=white, leftIndent=3*mm)

        PAGE_W = A4[0] - 30*mm  # 180mm
        COL_W  = [24*mm, 156*mm]

        col_code = t.get("ga500_col_code", "Code")
        col_name = t.get("ga500_col_name", "Name")
        HDR_ROW  = [col_code, col_name]

        # group params by category letter, with translated names
        from docs.ga500_params import get_param_name as _gpn
        by_cat = {}
        for code in params:
            by_cat.setdefault(code[0], []).append((code, _gpn(code, lang)))

        # page tracker for footer
        tracker = {'page': 0}

        def draw_page(canvas, doc):
            from docs.pdf_header import draw_page_header
            draw_page_header(canvas, doc)
            canvas.saveState()
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(HexColor('#888888'))
            pn = canvas.getPageNumber()
            canvas.drawCentredString(A4[0] / 2, 8 * mm, f'- {pn} -')
            canvas.restoreState()

        def row_style(n):
            ts = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), accent),
                ('TEXTCOLOR',  (0, 0), (-1, 0), HexColor('#000000')),
                ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE',   (0, 0), (-1, 0), 7),
                ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME',   (0, 1), (0, -1), 'Courier'),
                ('FONTSIZE',   (0, 1), (-1, -1), 7),
                ('ALIGN',      (0, 1), (0, -1), 'CENTER'),
                ('ALIGN',      (1, 1), (1, -1), 'LEFT'),
                ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWHEIGHT',  (0, 0), (0, 0), 12),
                ('ROWHEIGHT',  (0, 1), (-1, -1), 10),
                ('GRID',       (0, 0), (-1, -1), 0.25, HexColor('#CCCCCC')),
            ])
            for r in range(1, n + 1):
                if r % 2 == 0:
                    ts.add('BACKGROUND', (0, r), (-1, r), row_gray)
            return ts

        story = []
        story.append(Paragraph(t.get("ga500_view_title", "GA500 Parameters"), title_s))
        story.append(Spacer(1, 3*mm))

        cats = [c for c in _CAT_ORDER if c in by_cat]
        for idx, cat in enumerate(cats):
            label = _CATEGORIES.get(cat, cat)
            head_tbl = Table([[Paragraph(label, cat_s)]], colWidths=[PAGE_W])
            head_tbl.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0), (-1, -1), accent),
                ('TOPPADDING',    (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(head_tbl)
            story.append(Spacer(1, 1*mm))

            rows = sorted(by_cat[cat])
            tbl_data = [HDR_ROW] + [[code, name] for code, name in rows]
            tbl = Table(tbl_data, colWidths=COL_W, repeatRows=1)
            tbl.setStyle(row_style(len(rows)))
            story.append(tbl)

            if idx + 1 < len(cats):
                story.append(Spacer(1, 4*mm))

        try:
            doc = SimpleDocTemplate(
                output_path, pagesize=A4,
                leftMargin=15*mm, rightMargin=15*mm,
                topMargin=22*mm, bottomMargin=20*mm,
            )
            doc.build(story, onFirstPage=draw_page, onLaterPages=draw_page)
            return True
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return False

    def update_language(self, lang):
        t = TRANSLATIONS.get(lang, TRANSLATIONS["IT"])
        self._lbl_title.setText(t.get("ga500_view_title", "GA500 Parameters — YASKAWA"))
        self._btn_pdf.setText(t.get("btn_generate_pdf", "Genera PDF"))
        self._btn_close.setText(t.get("preview_close", "Chiudi"))
        self._lbl_footer.setText(
            t.get("ga500_footer_params", "{} parametri").format(self._ga500_count))
        from PySide6.QtWidgets import QTableWidgetItem as _TWI
        from docs.ga500_params import GA500_PARAMS, get_param_name as _gpn
        self._tbl.setHorizontalHeaderItem(0, _TWI(t.get("ga500_col_code", "Code")))
        self._tbl.setHorizontalHeaderItem(1, _TWI(t.get("ga500_col_name", "Name")))
        for r, code in enumerate(sorted(GA500_PARAMS.keys())):
            item = self._tbl.item(r, 1)
            if item is not None:
                item.setText(_gpn(code, lang))

    def _apply_theme(self):
        if self._app_state.is_dark_mode:
            self.setStyleSheet("""
                QWidget          { background:#231811; color:white; }
                QTableWidget     { background:#231811; color:white;
                                   gridline-color:#5C4938;
                                   alternate-background-color:#3A2D26; }
                QTableWidget::item:selected, QAbstractItemView::item:selected {
                                   background:#FF9248; color:black; }
                QHeaderView::section { background:#8A4533; color:white;
                                       border:1px solid #5C4938; padding:2px; }
                QPushButton      { background:#3A2D26; color:white;
                                   border:1px solid #5C4938;
                                   padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; }
                QLabel           { color:white; }
                QFrame[frameShape="4"] { color:#5C4938; }
            """)
        else:
            self.setStyleSheet("""
                QWidget          { background:#f5f5f5; color:black; }
                QTableWidget     { background:white; color:black;
                                   gridline-color:#dddddd;
                                   alternate-background-color:#FFF7F1; }
                QTableWidget::item:selected, QAbstractItemView::item:selected {
                                   background:#FF9248; color:black; }
                QHeaderView::section { background:#D97757; color:white;
                                       border:1px solid #cccccc;
                                       font-weight:bold; padding:2px; }
                QPushButton      { background:white; color:#A85C42;
                                   border:1px solid #aaaaaa;
                                   padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; color:white; }
                QLabel           { color:black; }
                QFrame[frameShape="4"] { color:#cccccc; }
            """)


class AboutDialog(QDialog):
    def __init__(self, parent=None, is_dark_mode=False, app_state=None):
        super().__init__(parent)
        self.app_state = app_state
        t = TRANSLATIONS[self.app_state.language]
        
        self.setWindowTitle(t["about_title"])
        self.setFixedSize(600, 750)
        
        layout = QVBoxLayout(self)
        
        # Section 1: Product Information
        group1 = QGroupBox(t["about_section1"])
        l1 = QVBoxLayout(group1)
        l1.setSpacing(5) # Consistent interlinea
        l1.addWidget(QLabel(t["about_app_name"]))
        l1.addWidget(QLabel(t["about_version"]))
        layout.addWidget(group1, 0)
        
        # Section 2: Additional Information
        group2 = QGroupBox(t["about_section2"])
        l2 = QVBoxLayout(group2)
        
        hostname = socket.gethostname()
        username = getpass.getuser()
        attempts = getattr(self.app_state, 'login_attempts', 1)
        app_path = sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(sys.argv[0])
        log_path = os.path.abspath(logger.LOG_FILE)
        processor = platform.processor()
        os_info = f"{platform.system()} {platform.release()}"
        
        ram_tot = psutil.virtual_memory().total / (1024**3)
        ram_app = psutil.Process(os.getpid()).memory_info().rss / (1024**2)
        
        l2.addWidget(QLabel(f"<b>OS:</b> {os_info}"))
        l2.addWidget(QLabel(f"<b>Hostname:</b> {hostname}"))
        l2.addWidget(QLabel(f"<b>User:</b> {username}"))
        l2.addWidget(QLabel(f"<b>Processor:</b> {processor}"))
        l2.addWidget(QLabel(f"<b>Total RAM:</b> {ram_tot:.2f} GB"))
        l2.addWidget(QLabel(f"<b>App RAM Usage:</b> {ram_app:.2f} MB"))
        l2.addWidget(QLabel(f"<b>Login Attempts:</b> {attempts}"))
        
        lbl_app_path = QLabel(f"<b>App Path:</b> {app_path}")
        lbl_app_path.setWordWrap(True)
        l2.addWidget(lbl_app_path)

        lbl_log_path = QLabel(f"<b>Log Path:</b> {log_path}")
        lbl_log_path.setWordWrap(True)
        l2.addWidget(lbl_log_path)

        l2.setSpacing(5)
        group2.setMaximumHeight(250)
        
        layout.addWidget(group2, 0) # Use 0 stretch to keep it compact
        
        # Section 3: Creator Information
        group3 = QGroupBox(t["about_section3"])
        l3 = QVBoxLayout(group3)
        l3.setAlignment(Qt.AlignCenter)
        l3.setSpacing(8)

        from gui.top_bar import get_resource_path
        photo_lbl = QLabel()
        photo_lbl.setAlignment(Qt.AlignCenter)
        photo_path = get_resource_path("Foto_profilo.jpg")
        photo_pix = QPixmap(photo_path)
        if not photo_pix.isNull():
            photo_pix = photo_pix.scaledToHeight(180, Qt.SmoothTransformation)
            photo_lbl.setPixmap(photo_pix)
        l3.addWidget(photo_lbl, alignment=Qt.AlignCenter)

        caption_lbl = QLabel("0THack1A")
        caption_lbl.setAlignment(Qt.AlignCenter)
        caption_lbl.setStyleSheet("font-size: 16px; font-weight: bold; margin-top: 6px;")
        l3.addWidget(caption_lbl, alignment=Qt.AlignCenter)

        layout.addWidget(group3, 1)

        if is_dark_mode:
            self.setStyleSheet("QDialog { background-color: #3A2D26; color: white; } QGroupBox { color: white; border: 1px solid #5C4938; margin-top: 10px; font-weight: bold; } QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 3px 0 3px; } QLabel { color: white; font-weight: normal; }")
        else:
            self.setStyleSheet("QDialog { background-color: white; color: black; } QGroupBox { color: black; border: 1px solid #ccc; margin-top: 10px; font-weight: bold; } QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 3px 0 3px; } QLabel { color: black; font-weight: normal; }")

    def closeEvent(self, event):
        logger.info("log_about_closed")
        super().closeEvent(event)

_AIUTO_CONTENT = {
    "IT": """\
<h2 style='color:#D97757;'>Guida rapida — YASKAWA Tools</h2>
<b>File</b><br>
&nbsp;• <b>Indirizzo IP</b> — Legge la configurazione di rete dal backup del robot.<br>
&nbsp;• <b>Yaskawa logData</b> — Importa e visualizza il file LOGDATA.DAT del robot.<br>
&nbsp;• <b>Compila</b> — Scrive valori nelle variabili del robot (VAR.DAT).<br><br>
<b>Genera</b><br>
&nbsp;• <b>Template</b> — Crea un template Excel per i nomi delle variabili e I/O.<br>
&nbsp;• <b>Nomi</b> — Importa nomi da Excel e genera i file DAT per il robot.<br>
&nbsp;• <b>Tool</b> — Configura ed esporta i dati dei tool TCP (fino a 64 tool).<br>
&nbsp;• <b>UF#()</b> — Visualizza e modifica i 63 User Frame (UFRAME.CND).<br>
&nbsp;• <b>IFPanel</b> — Legge e modifica il file IFPANEL.DAT per la configurazione del pannello operatore.<br><br>
<b>Documentazione</b><br>
&nbsp;• <b>Flowchart</b> — Visualizza i diagrammi di flusso dei file JBI con esportazione PDF e draw.io.<br>
&nbsp;• <b>Targhetta</b> — Genera PDF con informazioni di sistema dal SYSTEM.SYS.<br>
&nbsp;• <b>Panel</b> — Genera PDF del pannello con parametri e configurazione.<br>
&nbsp;• <b>JOBs</b> — Analizza i file JBI e genera PDF con sommario e albero di richiamo.<br>
&nbsp;• <b>Parametri</b> — Genera PDF dei parametri non-zero da ALL.PRM.<br>
&nbsp;• <b>UF#()/Tools</b> — Genera PDF con tool TCP e user frame configurati.<br>
&nbsp;• <b>User Group</b> — Visualizza e modifica gli User Group (GPIN/GPOT) con esportazione PDF/Excel.<br>
&nbsp;• <b>Completa</b> — Genera un PDF completo con tutti i documenti.<br><br>
<b>Backup</b><br>
&nbsp;• <b>Variabili</b> — Esporta le variabili del robot in Excel.<br>
&nbsp;• <b>Punti</b> — Esporta le posizioni dei programmi JBI in PDF.<br><br>
<b>Drive</b><br>
&nbsp;• <b>MotionDrive</b> — Legge file .YDWIProj o .YDWGProj e genera PDF/Excel dei parametri.<br><br>
<b>Help</b><br>
&nbsp;• <b>Registro</b> — Mostra/nasconde il pannello log.<br>
&nbsp;• <b>Parametri GA500</b> — Elenco parametri YASKAWA GA500 con ricerca.<br>
&nbsp;• <b>Parametri</b> — Elenco parametri YRC1000 con descrizioni.<br>
&nbsp;• <b>Param.Noti</b> — Parametri YRC1000 con valori consigliati noti.<br>
&nbsp;• <b>Informazioni</b> — Versione e contatti aziendali.<br>
&nbsp;• <b>Aiuto</b> — Questa guida.<br>
""",
    "EN": """\
<h2 style='color:#D97757;'>Quick Guide — YASKAWA Tools</h2>
<b>File</b><br>
&nbsp;• <b>IP Address</b> — Reads network configuration from robot backup.<br>
&nbsp;• <b>Yaskawa logData</b> — Imports and displays the robot's LOGDATA.DAT file.<br>
&nbsp;• <b>Compile</b> — Writes values to robot variables (VAR.DAT).<br><br>
<b>Generate</b><br>
&nbsp;• <b>Template</b> — Creates an Excel template for variable and I/O names.<br>
&nbsp;• <b>Names</b> — Imports names from Excel and generates DAT files for the robot.<br>
&nbsp;• <b>Tool</b> — Configures and exports TCP tool data (up to 64 tools).<br>
&nbsp;• <b>UF#()</b> — Views and edits the 63 User Frames (UFRAME.CND).<br>
&nbsp;• <b>IFPanel</b> — Reads and edits the IFPANEL.DAT file for operator panel configuration.<br><br>
<b>Documentation</b><br>
&nbsp;• <b>Flowchart</b> — Displays JBI flowcharts with PDF and draw.io export.<br>
&nbsp;• <b>Nameplate</b> — Generates PDF with system info from SYSTEM.SYS.<br>
&nbsp;• <b>Panel</b> — Generates PDF of the panel with parameters and configuration.<br>
&nbsp;• <b>JOBs</b> — Analyzes JBI files and generates PDF with summary and call tree.<br>
&nbsp;• <b>Parameters</b> — Generates PDF of non-zero parameters from ALL.PRM.<br>
&nbsp;• <b>UF#()/Tools</b> — Generates PDF with configured TCP tools and user frames.<br>
&nbsp;• <b>User Group</b> — Views and edits User Groups (GPIN/GPOT) with PDF/Excel export.<br>
&nbsp;• <b>Full</b> — Generates a full PDF with all documents.<br><br>
<b>Backup</b><br>
&nbsp;• <b>Variables</b> — Exports robot variables to Excel.<br>
&nbsp;• <b>Points</b> — Exports JBI program positions to PDF.<br><br>
<b>Drive</b><br>
&nbsp;• <b>MotionDrive</b> — Reads .YDWIProj or .YDWGProj file and generates PDF/Excel of parameters.<br><br>
<b>Help</b><br>
&nbsp;• <b>Log</b> — Shows/hides the log panel.<br>
&nbsp;• <b>GA500 Parameters</b> — YASKAWA GA500 parameter list with search.<br>
&nbsp;• <b>Parameters</b> — YRC1000 parameter list with descriptions.<br>
&nbsp;• <b>Known Params</b> — YRC1000 parameters with known recommended values.<br>
&nbsp;• <b>About</b> — Version and company contacts.<br>
&nbsp;• <b>Help Guide</b> — This guide.<br>
""",
    "FR": """\
<h2 style='color:#D97757;'>Guide rapide — YASKAWA Tools</h2>
<b>Fichier</b><br>
&nbsp;• <b>Adresse IP</b> — Lit la configuration réseau depuis la sauvegarde du robot.<br>
&nbsp;• <b>Yaskawa logData</b> — Importe et affiche le fichier LOGDATA.DAT du robot.<br>
&nbsp;• <b>Compiler</b> — Écrit des valeurs dans les variables du robot (VAR.DAT).<br><br>
<b>Générer</b><br>
&nbsp;• <b>Template</b> — Crée un modèle Excel pour les noms de variables et E/S.<br>
&nbsp;• <b>Noms</b> — Importe les noms depuis Excel et génère les fichiers DAT.<br>
&nbsp;• <b>Outil</b> — Configure et exporte les données des outils TCP (jusqu'à 64).<br>
&nbsp;• <b>UF#()</b> — Affiche et édite les 63 User Frames (UFRAME.CND).<br>
&nbsp;• <b>IFPanel</b> — Lit et édite le fichier IFPANEL.DAT pour la configuration du panneau opérateur.<br><br>
<b>Documentation</b><br>
&nbsp;• <b>Flowchart</b> — Affiche les organigrammes des fichiers JBI avec export PDF et draw.io.<br>
&nbsp;• <b>Plaque</b> — Génère un PDF avec les informations système depuis SYSTEM.SYS.<br>
&nbsp;• <b>Panel</b> — Génère un PDF du panneau avec paramètres et configuration.<br>
&nbsp;• <b>JOBs</b> — Analyse les fichiers JBI et génère un PDF avec sommaire et arbre d'appels.<br>
&nbsp;• <b>Paramètres</b> — Génère un PDF des paramètres non-zéro depuis ALL.PRM.<br>
&nbsp;• <b>UF#()/Outils</b> — Génère un PDF avec les outils TCP et frames utilisateur configurés.<br>
&nbsp;• <b>User Group</b> — Affiche et modifie les User Groups (GPIN/GPOT) avec export PDF/Excel.<br>
&nbsp;• <b>Complet</b> — Génère un PDF complet avec tous les documents.<br><br>
<b>Sauvegarde</b><br>
&nbsp;• <b>Variables</b> — Exporte les variables du robot vers Excel.<br>
&nbsp;• <b>Points</b> — Exporte les positions des programmes JBI en PDF.<br><br>
<b>Drive</b><br>
&nbsp;• <b>MotionDrive</b> — Lit les fichiers .YDWIProj ou .YDWGProj et génère PDF/Excel des paramètres.<br><br>
<b>Aide</b><br>
&nbsp;• <b>Journal</b> — Affiche/masque le panneau de log.<br>
&nbsp;• <b>Paramètres GA500</b> — Liste des paramètres YASKAWA GA500 avec recherche.<br>
&nbsp;• <b>Paramètres</b> — Liste des paramètres YRC1000 avec descriptions.<br>
&nbsp;• <b>Param. connus</b> — Paramètres YRC1000 avec valeurs recommandées connues.<br>
&nbsp;• <b>À propos</b> — Version et contacts entreprise.<br>
&nbsp;• <b>Guide</b> — Ce guide.<br>
""",
    "ES": """\
<h2 style='color:#D97757;'>Guía rápida — YASKAWA Tools</h2>
<b>Archivo</b><br>
&nbsp;• <b>Dirección IP</b> — Lee la configuración de red desde la copia de seguridad del robot.<br>
&nbsp;• <b>Yaskawa logData</b> — Importa y muestra el archivo LOGDATA.DAT del robot.<br>
&nbsp;• <b>Compilar</b> — Escribe valores en las variables del robot (VAR.DAT).<br><br>
<b>Generar</b><br>
&nbsp;• <b>Template</b> — Crea una plantilla Excel para nombres de variables y E/S.<br>
&nbsp;• <b>Nombres</b> — Importa nombres desde Excel y genera archivos DAT para el robot.<br>
&nbsp;• <b>Herramienta</b> — Configura y exporta datos de herramientas TCP (hasta 64).<br>
&nbsp;• <b>UF#()</b> — Visualiza y edita los 63 User Frames (UFRAME.CND).<br>
&nbsp;• <b>IFPanel</b> — Lee y edita el archivo IFPANEL.DAT para la configuración del panel de operador.<br><br>
<b>Documentación</b><br>
&nbsp;• <b>Flowchart</b> — Muestra los diagramas de flujo de los archivos JBI con exportación PDF y draw.io.<br>
&nbsp;• <b>Placa</b> — Genera PDF con información del sistema desde SYSTEM.SYS.<br>
&nbsp;• <b>Panel</b> — Genera PDF del panel con parámetros y configuración.<br>
&nbsp;• <b>JOBs</b> — Analiza archivos JBI y genera PDF con resumen y árbol de llamadas.<br>
&nbsp;• <b>Parámetros</b> — Genera PDF de parámetros no-cero desde ALL.PRM.<br>
&nbsp;• <b>UF#()/Herramientas</b> — Genera PDF con herramientas TCP y frames de usuario configurados.<br>
&nbsp;• <b>User Group</b> — Visualiza y edita los User Groups (GPIN/GPOT) con exportación PDF/Excel.<br>
&nbsp;• <b>Completo</b> — Genera un PDF completo con todos los documentos.<br><br>
<b>Copia de seguridad</b><br>
&nbsp;• <b>Variables</b> — Exporta variables del robot a Excel.<br>
&nbsp;• <b>Puntos</b> — Exporta posiciones de programas JBI en PDF.<br><br>
<b>Drive</b><br>
&nbsp;• <b>MotionDrive</b> — Lee archivos .YDWIProj o .YDWGProj y genera PDF/Excel de parámetros.<br><br>
<b>Ayuda</b><br>
&nbsp;• <b>Registro</b> — Muestra/oculta el panel de log.<br>
&nbsp;• <b>Parámetros GA500</b> — Lista de parámetros YASKAWA GA500 con búsqueda.<br>
&nbsp;• <b>Parámetros</b> — Lista de parámetros YRC1000 con descripciones.<br>
&nbsp;• <b>Param. conocidos</b> — Parámetros YRC1000 con valores recomendados conocidos.<br>
&nbsp;• <b>Acerca de</b> — Versión y contactos de la empresa.<br>
&nbsp;• <b>Guía</b> — Esta guía.<br>
""",
    "DE": """\
<h2 style='color:#D97757;'>Kurzanleitung — YASKAWA Tools</h2>
<b>Datei</b><br>
&nbsp;• <b>IP-Adresse</b> — Liest die Netzwerkkonfiguration aus der Roboter-Sicherung.<br>
&nbsp;• <b>Yaskawa logData</b> — Importiert und zeigt die LOGDATA.DAT-Datei des Roboters an.<br>
&nbsp;• <b>Kompilieren</b> — Schreibt Werte in die Robotervariablen (VAR.DAT).<br><br>
<b>Generieren</b><br>
&nbsp;• <b>Template</b> — Erstellt eine Excel-Vorlage für Variablen- und E/A-Namen.<br>
&nbsp;• <b>Namen</b> — Importiert Namen aus Excel und generiert DAT-Dateien für den Roboter.<br>
&nbsp;• <b>Werkzeug</b> — Konfiguriert und exportiert TCP-Werkzeugdaten (bis zu 64).<br>
&nbsp;• <b>UF#()</b> — Zeigt und bearbeitet die 63 User Frames (UFRAME.CND).<br>
&nbsp;• <b>IFPanel</b> — Liest und bearbeitet die Datei IFPANEL.DAT zur Konfiguration des Bedienpanels.<br><br>
<b>Dokumentation</b><br>
&nbsp;• <b>Flowchart</b> — Zeigt JBI-Flussdiagramme mit PDF- und draw.io-Export.<br>
&nbsp;• <b>Typenschild</b> — Generiert PDF mit Systeminformationen aus SYSTEM.SYS.<br>
&nbsp;• <b>Panel</b> — Generiert PDF des Panels mit Parametern und Konfiguration.<br>
&nbsp;• <b>JOBs</b> — Analysiert JBI-Dateien und generiert PDF mit Inhaltsverzeichnis und Aufrufbaum.<br>
&nbsp;• <b>Parameter</b> — Generiert PDF der Nicht-Null-Parameter aus ALL.PRM.<br>
&nbsp;• <b>UF#()/Werkzeuge</b> — Generiert PDF mit konfigurierten TCP-Werkzeugen und User Frames.<br>
&nbsp;• <b>User Group</b> — Zeigt und bearbeitet User Groups (GPIN/GPOT) mit PDF/Excel-Export.<br>
&nbsp;• <b>Vollständig</b> — Generiert ein vollständiges PDF mit allen Dokumenten.<br><br>
<b>Backup</b><br>
&nbsp;• <b>Variablen</b> — Exportiert Robotervariablen nach Excel.<br>
&nbsp;• <b>Punkte</b> — Exportiert JBI-Programmpositionen als PDF.<br><br>
<b>Drive</b><br>
&nbsp;• <b>MotionDrive</b> — Liest .YDWIProj- oder .YDWGProj-Dateien und generiert PDF/Excel der Parameter.<br><br>
<b>Hilfe</b><br>
&nbsp;• <b>Protokoll</b> — Zeigt/verbirgt das Log-Panel.<br>
&nbsp;• <b>GA500-Parameter</b> — YASKAWA GA500-Parameterliste mit Suche.<br>
&nbsp;• <b>Parameter</b> — YRC1000-Parameterliste mit Beschreibungen.<br>
&nbsp;• <b>Bekannte Param.</b> — YRC1000-Parameter mit bekannten empfohlenen Werten.<br>
&nbsp;• <b>Über</b> — Version und Firmenkontakte.<br>
&nbsp;• <b>Hilfe</b> — Diese Anleitung.<br>
""",
    "UA": """\
<h2 style='color:#D97757;'>Коротка довідка — YASKAWA Tools</h2>
<b>Файл</b><br>
&nbsp;• <b>IP-адреса</b> — Зчитує мережеву конфігурацію з резервної копії робота.<br>
&nbsp;• <b>Yaskawa logData</b> — Імпортує та відображає файл LOGDATA.DAT робота.<br>
&nbsp;• <b>Компілювати</b> — Записує значення у змінні робота (VAR.DAT).<br><br>
<b>Генерувати</b><br>
&nbsp;• <b>Template</b> — Створює шаблон Excel для імен змінних та вхідних/вихідних сигналів.<br>
&nbsp;• <b>Назви</b> — Імпортує назви з Excel і генерує DAT-файли для робота.<br>
&nbsp;• <b>Інструмент</b> — Налаштовує та експортує дані TCP інструментів (до 64).<br>
&nbsp;• <b>UF#()</b> — Переглядає та редагує 63 User Frame (UFRAME.CND).<br>
&nbsp;• <b>IFPanel</b> — Зчитує та редагує файл IFPANEL.DAT для конфігурації операторської панелі.<br><br>
<b>Документація</b><br>
&nbsp;• <b>Flowchart</b> — Відображає блок-схеми JBI-файлів з експортом у PDF і draw.io.<br>
&nbsp;• <b>Табличка</b> — Генерує PDF з системною інформацією з SYSTEM.SYS.<br>
&nbsp;• <b>Панель</b> — Генерує PDF панелі з параметрами та конфігурацією.<br>
&nbsp;• <b>JOBs</b> — Аналізує JBI-файли та генерує PDF з підсумком і деревом викликів.<br>
&nbsp;• <b>Параметри</b> — Генерує PDF ненульових параметрів з ALL.PRM.<br>
&nbsp;• <b>UF#()/Інструменти</b> — Генерує PDF з налаштованими TCP інструментами та User Frame.<br>
&nbsp;• <b>User Group</b> — Переглядає та редагує User Groups (GPIN/GPOT) з експортом PDF/Excel.<br>
&nbsp;• <b>Повний</b> — Генерує повний PDF з усіма документами.<br><br>
<b>Резервне копіювання</b><br>
&nbsp;• <b>Змінні</b> — Експортує змінні робота в Excel.<br>
&nbsp;• <b>Точки</b> — Експортує позиції програм JBI в PDF.<br><br>
<b>Drive</b><br>
&nbsp;• <b>MotionDrive</b> — Зчитує файли .YDWIProj або .YDWGProj і генерує PDF/Excel параметрів.<br><br>
<b>Допомога</b><br>
&nbsp;• <b>Журнал</b> — Показує/приховує панель журналу.<br>
&nbsp;• <b>Параметри GA500</b> — Список параметрів YASKAWA GA500 з пошуком.<br>
&nbsp;• <b>Параметри</b> — Список параметрів YRC1000 з описами.<br>
&nbsp;• <b>Відомі параметри</b> — Параметри YRC1000 з відомими рекомендованими значеннями.<br>
&nbsp;• <b>Про програму</b> — Версія та контакти компанії.<br>
&nbsp;• <b>Довідка</b> — Цей посібник.<br>
""",
    "JA": """\
<h2 style='color:#D97757;'>クイックガイド — YASKAWA Tools</h2>
<b>ファイル</b><br>
&nbsp;• <b>IPアドレス</b> — ロボットのバックアップからネットワーク設定を読み取ります。<br>
&nbsp;• <b>Yaskawa logData</b> — ロボットのLOGDATA.DATファイルをインポートして表示します。<br>
&nbsp;• <b>コンパイル</b> — ロボット変数（VAR.DAT）に値を書き込みます。<br><br>
<b>生成</b><br>
&nbsp;• <b>テンプレート</b> — 変数とI/O名のExcelテンプレートを作成します。<br>
&nbsp;• <b>名称</b> — ExcelからインポートしてロボットのDATファイルを生成します。<br>
&nbsp;• <b>ツール</b> — TCPツールデータを設定・エクスポートします（最大64ツール）。<br>
&nbsp;• <b>UF#()</b> — 63のUser Frame（UFRAME.CND）を表示・編集します。<br>
&nbsp;• <b>IFPanel</b> — オペレーターパネル設定用のIFPANEL.DATファイルを読み取り・編集します。<br><br>
<b>ドキュメント</b><br>
&nbsp;• <b>Flowchart</b> — JBIファイルのフローチャートを表示し、PDFとdraw.ioにエクスポートします。<br>
&nbsp;• <b>銘板</b> — SYSTEM.SYSのシステム情報からPDFを生成します。<br>
&nbsp;• <b>パネル</b> — パラメータと設定を含むパネルのPDFを生成します。<br>
&nbsp;• <b>JOBs</b> — JBIファイルを解析し、サマリーとコールツリーのPDFを生成します。<br>
&nbsp;• <b>パラメータ</b> — ALL.PRMの非ゼロパラメータのPDFを生成します。<br>
&nbsp;• <b>UF#()/ツール</b> — 設定済みTCPツールとUser FrameのPDFを生成します。<br>
&nbsp;• <b>User Group</b> — User Group（GPIN/GPOT）を表示・編集し、PDF/Excelにエクスポートします。<br>
&nbsp;• <b>完全版</b> — 全ドキュメントを含む完全なPDFを生成します。<br><br>
<b>バックアップ</b><br>
&nbsp;• <b>変数</b> — ロボット変数をExcelにエクスポートします。<br>
&nbsp;• <b>ポイント</b> — JBIプログラムの位置をPDFにエクスポートします。<br><br>
<b>ドライブ</b><br>
&nbsp;• <b>モーションドライブ</b> — .YDWIProjectまたは.YDWGProjファイルを読み取り、PDF/Excelを生成します。<br><br>
<b>ヘルプ</b><br>
&nbsp;• <b>ログ</b> — ログパネルの表示/非表示を切り替えます。<br>
&nbsp;• <b>GA500パラメータ</b> — 検索付きYASKAWA GA500パラメータリスト。<br>
&nbsp;• <b>パラメータ</b> — 説明付きYRC1000パラメータリスト。<br>
&nbsp;• <b>既知パラメータ</b> — 推奨値付きYRC1000パラメータ。<br>
&nbsp;• <b>バージョン情報</b> — バージョンと会社情報。<br>
&nbsp;• <b>ヘルプガイド</b> — このガイド。<br>
""",
}


class _AiutoView(QWidget):
    """Inline help guide panel shown in the content area."""

    def __init__(self, app_state, on_close_cb):
        super().__init__()
        self._app_state = app_state
        self._on_close  = on_close_cb
        self._build_ui()
        self._apply_theme()

    def _build_ui(self):
        from PySide6.QtWidgets import QScrollArea, QTextBrowser
        t    = TRANSLATIONS[self._app_state.language]
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 6, 8, 6)
        root.setSpacing(4)

        toolbar = QHBoxLayout()
        self._lbl = QLabel(t.get("menu_aiuto", "Aiuto"))
        self._lbl.setStyleSheet("font-weight: bold; font-size: 10pt;")
        toolbar.addWidget(self._lbl)
        toolbar.addStretch()
        self._btn_close = QPushButton(t.get("preview_close", "Chiudi"))
        self._btn_close.setFixedHeight(26)
        self._btn_close.clicked.connect(self._on_close)
        toolbar.addWidget(self._btn_close)
        root.addLayout(toolbar)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Sunken)
        root.addWidget(sep)

        self._text = QTextBrowser()
        self._text.setOpenExternalLinks(False)
        lang = self._app_state.language
        html = _AIUTO_CONTENT.get(lang, _AIUTO_CONTENT["IT"])
        self._text.setHtml(html)
        root.addWidget(self._text, 1)

    def update_language(self, lang):
        try:
            t = TRANSLATIONS.get(lang, TRANSLATIONS["IT"])
            self._lbl.setText(t.get("menu_aiuto", "Aiuto"))
            self._btn_close.setText(t.get("preview_close", "Chiudi"))
            html = _AIUTO_CONTENT.get(lang, _AIUTO_CONTENT["IT"])
            self._text.setHtml(html)
        except Exception:
            pass

    def _apply_theme(self):
        if self._app_state.is_dark_mode:
            self.setStyleSheet("""
                QWidget { background:#231811; color:white; }
                QTextBrowser { background:#231811; color:white; border:none; }
                QPushButton { background:#3A2D26; color:white;
                              border:1px solid #5C4938;
                              padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; }
                QLabel { color:white; }
            """)
        else:
            self.setStyleSheet("""
                QWidget { background:#f5f5f5; color:black; }
                QTextBrowser { background:white; color:black; border:none; }
                QPushButton { background:white; color:#A85C42;
                              border:1px solid #aaaaaa;
                              padding:2px 10px; border-radius:3px; }
                QPushButton:hover { background:#D97757; color:white; }
                QLabel { color:black; }
            """)


class MainWindow(QMainWindow):
    def __init__(self, app_state):
        super().__init__()
        self.app_state = app_state

        self.setWindowTitle(TRANSLATIONS[self.app_state.language]["title_main"])
        self.resize(1000, 700)

        self._pdf_preview    = None
        self._excel_preview  = None
        self._backup_view    = None
        self._drive_view     = None
        self._ga500_view     = None
        self._tool_panel     = None
        self._help_view      = None
        self._uf_tools_view  = None
        self._ipnet_view     = None
        self._compila_view   = None
        self._uframe_view    = None
        self._aiuto_view     = None
        self._ifpanel_view   = None
        self._flowchart_view = None
        self._usrgrp_view    = None
        self._work_folder    = self._load_work_folder()
        self.setup_ui()
        self.apply_theme()

        # Connect real-time logging
        logger.signals.log_emitted.connect(self.append_log)
        self.load_logs()

    def setup_ui(self):
        # Central widget
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        
        main_layout = QVBoxLayout(self.central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Top Bar
        self.top_bar = TopBar(self.app_state)
        self.top_bar.subtitle_label.setText("")
        self.top_bar.theme_toggled.connect(self.on_theme_toggled)
        self.top_bar.language_changed.connect(self.on_language_changed)
        main_layout.addWidget(self.top_bar)

        # ── Working folder path bar (STEP 11) ────────────────────────────────
        self._folder_bar = QWidget()
        self._folder_bar.setFixedHeight(32)
        fb_layout = QHBoxLayout(self._folder_bar)
        fb_layout.setContentsMargins(10, 2, 10, 2)
        fb_layout.setSpacing(6)
        t0 = TRANSLATIONS[self.app_state.language]
        self._folder_lbl_key = QLabel(t0.get("work_folder_label", "Cartella robot:"))
        self._folder_lbl_key.setStyleSheet("font-weight: bold; font-size: 9pt;")
        fb_layout.addWidget(self._folder_lbl_key)
        self._folder_lbl_path = QLabel(
            self._work_folder if self._work_folder
            else t0.get("work_folder_none", "Nessuna cartella selezionata")
        )
        self._folder_lbl_path.setStyleSheet("font-size: 9pt;")
        self._folder_lbl_path.setMinimumWidth(200)
        fb_layout.addWidget(self._folder_lbl_path, 1)
        self._folder_btn = QPushButton(t0.get("work_folder_btn", "Cartella robot"))
        self._folder_btn.setObjectName("work_folder_btn")
        self._folder_btn.setFixedHeight(24)
        self._folder_btn.clicked.connect(self._on_choose_work_folder)
        fb_layout.addWidget(self._folder_btn)
        main_layout.addWidget(self._folder_bar)

        # ── Progress bar (PDF generation indicator) ───────────────────────────
        self._progress_bar = QProgressBar()
        self._progress_bar.setRange(0, 100)
        self._progress_bar.setValue(0)
        self._progress_bar.setFixedHeight(4)
        self._progress_bar.setTextVisible(False)
        self._progress_bar.setStyleSheet(
            "QProgressBar { border: none; background: transparent; }"
            "QProgressBar::chunk { background: #22c55e; border-radius: 2px; }"
        )
        self._progress_bar.hide()
        main_layout.addWidget(self._progress_bar)

        # Splitter for content and logs
        self.splitter = QSplitter(Qt.Vertical)

        # Content area (Top part of splitter)
        self.content_area = QWidget()
        content_layout = QVBoxLayout(self.content_area)
        content_layout.setContentsMargins(0, 0, 0, 0)

        # ── Welcome label (STEP 4) ────────────────────────────────────────────
        self._welcome_lbl = QLabel(t0.get("welcome_message", ""))
        self._welcome_lbl.setAlignment(Qt.AlignCenter)
        self._welcome_lbl.setWordWrap(True)
        content_layout.addWidget(self._welcome_lbl)
        
        self.splitter.addWidget(self.content_area)
        
        # Log Panel (Bottom part of splitter)
        self.log_panel = QWidget()
        log_layout = QVBoxLayout(self.log_panel)
        log_layout.setContentsMargins(10, 10, 10, 10)
        
        self.log_header = QLabel(
            TRANSLATIONS[self.app_state.language].get("log_panel_title", "System Logs"))
        self.log_header.setStyleSheet("font-weight: bold; font-size: 14px;")
        log_layout.addWidget(self.log_header)
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        log_layout.addWidget(self.log_text)
        
        self.splitter.addWidget(self.log_panel)
        
        # Set initial sizes (content gets more space, log panel gets space for ~4 lines)
        self.splitter.setSizes([500, 150])
        
        main_layout.addWidget(self.splitter, 1)
        
        # Menu Bar
        self.menu_bar = self.menuBar()
        
        # File Menu
        t = TRANSLATIONS[self.app_state.language]
        self.menu_file = self.menu_bar.addMenu(t["menu_file"])

        self.action_ipnet = QAction(t.get("menu_ipnet", "Indirizzo IP"), self)
        self.action_ipnet.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_ipnet.text()))
        self.action_ipnet.triggered.connect(self.on_ipnet)
        self.menu_file.addAction(self.action_ipnet)

        self.action_logdata = QAction(t.get("menu_logdata", "Yaskawa logData"), self)
        self.action_logdata.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_logdata.text()))
        self.action_logdata.triggered.connect(self.on_logdata)
        self.menu_file.addAction(self.action_logdata)

        self.menu_file.addSeparator()

        self.action_compila = QAction(t.get("menu_compila", "Compila"), self)
        self.action_compila.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_compila.text()))
        self.action_compila.triggered.connect(self.on_compila)
        self.menu_file.addAction(self.action_compila)

        self.menu_file.addSeparator()

        self.action_exit = QAction(t["btn_exit"], self)
        self.action_exit.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_exit.text()))
        self.action_exit.triggered.connect(self.close)
        self.menu_file.addAction(self.action_exit)

        # Genera Menu
        self.menu_genera = self.menu_bar.addMenu(TRANSLATIONS[self.app_state.language]["menu_genera"])

        self.action_template = QAction(TRANSLATIONS[self.app_state.language]["menu_template"], self)
        self.action_template.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_template.text()))
        self.action_template.triggered.connect(self.on_template)
        self.menu_genera.addAction(self.action_template)

        self.action_nomi = QAction(TRANSLATIONS[self.app_state.language]["menu_nomi"], self)
        self.action_nomi.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_nomi.text()))
        self.action_nomi.triggered.connect(self.on_nomi)
        self.menu_genera.addAction(self.action_nomi)

        self.action_tool = QAction(TRANSLATIONS[self.app_state.language]["menu_tool"], self)
        self.action_tool.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_tool.text()))
        self.action_tool.triggered.connect(self.on_tool)
        self.menu_genera.addAction(self.action_tool)

        self.action_uf = QAction(TRANSLATIONS[self.app_state.language]["menu_uf"], self)
        self.action_uf.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_uf.text()))
        self.action_uf.triggered.connect(self.on_uf)
        self.menu_genera.addAction(self.action_uf)

        self.action_ifpanel = QAction(TRANSLATIONS[self.app_state.language].get("menu_ifpanel", "IFPanel"), self)
        self.action_ifpanel.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_ifpanel.text()))
        self.action_ifpanel.triggered.connect(self.on_ifpanel)
        self.menu_genera.addAction(self.action_ifpanel)

        # Documentation Menu
        self.menu_documentation = self.menu_bar.addMenu(TRANSLATIONS[self.app_state.language]["menu_documentation"])

        self.action_flowchart = QAction(TRANSLATIONS[self.app_state.language].get("menu_flowchart", "Flowchart"), self)
        self.action_flowchart.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_flowchart.text()))
        self.action_flowchart.triggered.connect(self.on_flowchart)
        self.menu_documentation.addAction(self.action_flowchart)

        self.menu_documentation.addSeparator()

        self.action_targhetta = QAction(TRANSLATIONS[self.app_state.language]["menu_targhetta"], self)
        self.action_targhetta.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_targhetta.text()))
        self.action_targhetta.triggered.connect(self.on_targhetta)
        self.menu_documentation.addAction(self.action_targhetta)

        self.action_panel = QAction(TRANSLATIONS[self.app_state.language]["menu_panel"], self)
        self.action_panel.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_panel.text()))
        self.action_panel.triggered.connect(self.on_panel)
        self.menu_documentation.addAction(self.action_panel)

        self.action_jobs = QAction(TRANSLATIONS[self.app_state.language]["menu_jobs"], self)
        self.action_jobs.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_jobs.text()))
        self.action_jobs.triggered.connect(self.on_jobs)
        self.menu_documentation.addAction(self.action_jobs)

        self.action_params = QAction(TRANSLATIONS[self.app_state.language]["menu_params"], self)
        self.action_params.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_params.text()))
        self.action_params.triggered.connect(self.on_params)
        self.menu_documentation.addAction(self.action_params)

        self.action_uf_tools = QAction(TRANSLATIONS[self.app_state.language]["menu_uf_tools"], self)
        self.action_uf_tools.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_uf_tools.text()))
        self.action_uf_tools.triggered.connect(self.on_uf_tools)
        self.menu_documentation.addAction(self.action_uf_tools)

        self.action_usrgrp = QAction(
            TRANSLATIONS[self.app_state.language].get("menu_usrgrp", "User Group"), self)
        self.action_usrgrp.triggered.connect(
            lambda: logger.info("log_btn_pressed", self.action_usrgrp.text()))
        self.action_usrgrp.triggered.connect(self.on_usrgrp)
        self.menu_documentation.addAction(self.action_usrgrp)

        self.action_cubeintf = QAction(
            TRANSLATIONS[self.app_state.language].get("menu_cubeintf", "Cubo interferenza"), self)
        self.action_cubeintf.triggered.connect(
            lambda: logger.info("log_btn_pressed", self.action_cubeintf.text()))
        self.action_cubeintf.triggered.connect(self.on_cubeintf)
        self.menu_documentation.addAction(self.action_cubeintf)

        self.action_formcut = QAction(
            TRANSLATIONS[self.app_state.language].get("menu_formcut", "FormCut"), self)
        self.action_formcut.triggered.connect(
            lambda: logger.info("log_btn_pressed", self.action_formcut.text()))
        self.action_formcut.triggered.connect(self.on_formcut)
        self.menu_documentation.addAction(self.action_formcut)

        self.menu_documentation.addSeparator()

        self.action_completa = QAction(TRANSLATIONS[self.app_state.language]["menu_completa"], self)
        self.action_completa.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_completa.text()))
        self.action_completa.triggered.connect(self.on_completa)
        self.menu_documentation.addAction(self.action_completa)

        # Backup Menu
        self.menu_backup = self.menu_bar.addMenu(TRANSLATIONS[self.app_state.language]["menu_backup"])

        self.action_backup_var = QAction(TRANSLATIONS[self.app_state.language]["menu_backup_var"], self)
        self.action_backup_var.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_backup_var.text()))
        self.action_backup_var.triggered.connect(self.on_backup_var)
        self.menu_backup.addAction(self.action_backup_var)

        self.action_backup_punti = QAction(TRANSLATIONS[self.app_state.language]["menu_backup_punti"], self)
        self.action_backup_punti.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_backup_punti.text()))
        self.action_backup_punti.triggered.connect(self.on_backup_punti)
        self.menu_backup.addAction(self.action_backup_punti)

        # Drive Menu
        self.menu_drive = self.menu_bar.addMenu(TRANSLATIONS[self.app_state.language]["menu_drive"])

        self.action_motion_drive = QAction(TRANSLATIONS[self.app_state.language]["menu_motion_drive"], self)
        self.action_motion_drive.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_motion_drive.text()))
        self.action_motion_drive.triggered.connect(self.on_motion_drive)
        self.menu_drive.addAction(self.action_motion_drive)

        # Help Menu
        self.menu_help = self.menu_bar.addMenu(TRANSLATIONS[self.app_state.language]["menu_help"])

        self.action_log = QAction(TRANSLATIONS[self.app_state.language]["menu_log"], self)
        self.action_log.setCheckable(True)
        self.action_log.setChecked(True)
        self.action_log.triggered.connect(self.toggle_logs)
        self.menu_help.addAction(self.action_log)

        self.action_ga500_params = QAction(TRANSLATIONS[self.app_state.language]["menu_ga500_params"], self)
        self.action_ga500_params.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_ga500_params.text()))
        self.action_ga500_params.triggered.connect(self.on_ga500_params)
        self.menu_help.addAction(self.action_ga500_params)

        self.menu_help.addSeparator()

        self.action_help_params = QAction(TRANSLATIONS[self.app_state.language]["menu_help_params"], self)
        self.action_help_params.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_help_params.text()))
        self.action_help_params.triggered.connect(self.on_help_params)
        self.menu_help.addAction(self.action_help_params)

        self.action_help_known = QAction(TRANSLATIONS[self.app_state.language]["menu_help_known"], self)
        self.action_help_known.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_help_known.text()))
        self.action_help_known.triggered.connect(self.on_help_known)
        self.menu_help.addAction(self.action_help_known)

        self.menu_help.addSeparator()

        self.action_about = QAction(TRANSLATIONS[self.app_state.language]["menu_about"], self)
        self.action_about.triggered.connect(self.show_about)
        self.menu_help.addAction(self.action_about)

        self.action_aiuto = QAction(TRANSLATIONS[self.app_state.language].get("menu_aiuto", "Aiuto"), self)
        self.action_aiuto.triggered.connect(lambda: logger.info("log_btn_pressed", self.action_aiuto.text()))
        self.action_aiuto.triggered.connect(self.on_aiuto)
        self.menu_help.addAction(self.action_aiuto)

        # Tooltip wiring: hover over a menu/button shows a quick hint.
        # Qt default behaviour: tooltip is shown after a short delay and
        # disappears either when the mouse moves away or after a few seconds.
        self._apply_tooltips()

    def _apply_tooltips(self):
        """Set descriptive tooltips on every menu, action, button and field.

        Uses ``tooltips.py`` as a localised description table — the text
        describes what each control DOES (not just its label) in the
        current language.  Falls back to the visible text when no
        descriptive tip is available.
        """
        try:
            from PySide6.QtWidgets import QMenu, QPushButton, QLineEdit
            try:
                import tooltips as _tips
            except Exception:
                _tips = None

            lang = getattr(self.app_state, "language", "IT")

            def _tip_for(name: str, fallback: str = "") -> str:
                if _tips is not None:
                    txt = _tips.get(name, lang)
                    if txt:
                        return txt
                return fallback or ""

            # ── Top-level menus ──────────────────────────────────────────────
            menus_map = {
                "menu_file":          getattr(self, "menu_file", None),
                "menu_genera":        getattr(self, "menu_genera", None),
                "menu_documentation": getattr(self, "menu_documentation", None),
                "menu_backup":        getattr(self, "menu_backup", None),
                "menu_drive":         getattr(self, "menu_drive", None),
                "menu_help":          getattr(self, "menu_help", None),
            }
            for key, menu in menus_map.items():
                if menu is None:
                    continue
                try:
                    menu.setToolTipsVisible(True)
                except Exception:
                    pass
                ma = menu.menuAction()
                if ma is not None:
                    ma.setToolTip(_tip_for(key, menu.title().replace("&", "")))

            # ── Actions stored as self.action_<name> ─────────────────────────
            for attr_name in dir(self):
                if not attr_name.startswith("action_"):
                    continue
                act = getattr(self, attr_name, None)
                if act is None or not hasattr(act, "setToolTip"):
                    continue
                try:
                    label = act.text().replace("&", "")
                except Exception:
                    label = ""
                act.setToolTip(_tip_for(attr_name, label))

            # ── PushButtons & QLineEdit — match by objectName / label ────────
            close_labels = ("Chiudi", "Close", "Fermer", "Cerrar",
                            "Schließen", "Закрити", "閉じる")
            for btn in self.findChildren(QPushButton):
                key = btn.objectName() or ""
                label = btn.text().replace("&", "")
                if not key and label in close_labels:
                    key = "preview_close"
                tip = _tip_for(key, label)
                if tip:
                    btn.setToolTip(tip)

            for le in self.findChildren(QLineEdit):
                key = le.objectName() or ""
                tip = _tip_for(key, "")
                if tip:
                    le.setToolTip(tip)
        except Exception:
            # Tooltips are a UX nicety — never let a wiring failure block the GUI.
            pass

    # ── Progress bar helpers ──────────────────────────────────────────────────

    def _progress_begin(self):
        self._progress_bar.setValue(0)
        self._progress_bar.show()
        QApplication.processEvents()

    def _progress_update(self, value):
        self._progress_bar.setValue(int(value))
        QApplication.processEvents()

    def _progress_end(self):
        self._progress_bar.setValue(100)
        QApplication.processEvents()
        self._progress_bar.hide()

    def load_logs(self):
        if os.path.exists(logger.LOG_FILE):
            with open(logger.LOG_FILE, "r", encoding="utf-8") as f:
                self.log_text.setPlainText(f.read())
            # Scroll to bottom
            scrollbar = self.log_text.verticalScrollBar()
            scrollbar.setValue(scrollbar.maximum())

    def append_log(self, text):
        self.log_text.moveCursor(QTextCursor.End)
        self.log_text.insertPlainText(text)
        self.log_text.verticalScrollBar().setValue(self.log_text.verticalScrollBar().maximum())

    def toggle_logs(self, checked):
        self.log_panel.setVisible(checked)
        t = TRANSLATIONS[self.app_state.language]
        status = t["log_visible"] if checked else t["log_hidden"]
        logger.info("log_log_toggled", status)

    def apply_theme(self):
        if self.app_state.is_dark_mode:
            self.setStyleSheet("QMainWindow { background-color: #231811; } QMenuBar { background-color: #3A2D26; color: white; } QMenu { background-color: #3A2D26; color: white; }")
            self.content_area.setStyleSheet("background-color: #231811;")
            self.log_panel.setStyleSheet("background-color: #3A2D26;")
            self.log_header.setStyleSheet("color: white; font-weight: bold; font-size: 14px;")
            self.log_text.setStyleSheet("background-color: #231811; color: white; border: 1px solid #5C4938;")
            if self._welcome_lbl is not None:
                self._welcome_lbl.setStyleSheet("font-size: 13pt; color: #aaaaaa; padding: 40px;")
        else:
            self.setStyleSheet("QMainWindow { background-color: #f0f0f0; } QMenuBar { background-color: white; color: black; } QMenu { background-color: white; color: black; }")
            self.content_area.setStyleSheet("background-color: #f0f0f0;")
            self.log_panel.setStyleSheet("background-color: #e0e0e0;")
            self.log_header.setStyleSheet("color: black; font-weight: bold; font-size: 14px;")
            self.log_text.setStyleSheet("background-color: white; color: black; border: 1px solid #ccc;")
            if self._welcome_lbl is not None:
                self._welcome_lbl.setStyleSheet("font-size: 13pt; color: #555555; padding: 40px;")

    def on_theme_toggled(self, is_dark):
        self.apply_theme()
        for panel in [self._pdf_preview, self._excel_preview, self._backup_view,
                      self._drive_view, self._ga500_view, self._uf_tools_view,
                      self._ipnet_view, self._compila_view, self._uframe_view,
                      self._tool_panel, self._help_view, self._ifpanel_view,
                      self._flowchart_view, self._usrgrp_view]:
            try:
                if panel is not None and hasattr(panel, '_apply_theme'):
                    panel._apply_theme()
            except Exception:
                pass
        t = TRANSLATIONS[self.app_state.language]
        theme_name = t["theme_dark"] if is_dark else t["theme_light"]
        logger.info("log_theme_changed", theme_name)

    def on_language_changed(self, lang):
        if lang not in TRANSLATIONS:
            return
        t = TRANSLATIONS[lang]
        logger.set_log_language(lang)
        logger.info("log_lang_changed", lang)
        self.setWindowTitle(t["title_main"])
        self.menu_file.setTitle(t["menu_file"])
        self.action_ipnet.setText(t.get("menu_ipnet", "Indirizzo IP"))
        self.action_logdata.setText(t.get("menu_logdata", "Yaskawa logData"))
        self.action_compila.setText(t.get("menu_compila", "Compila"))
        self.menu_genera.setTitle(t["menu_genera"])
        self.action_template.setText(t["menu_template"])
        self.action_nomi.setText(t["menu_nomi"])
        self.action_tool.setText(t["menu_tool"])
        self.action_uf.setText(t["menu_uf"])
        self.action_ifpanel.setText(t.get("menu_ifpanel", "IFPanel"))
        self.menu_documentation.setTitle(t["menu_documentation"])
        self.action_flowchart.setText(t.get("menu_flowchart", "Flowchart"))
        self.action_targhetta.setText(t["menu_targhetta"])
        self.action_panel.setText(t["menu_panel"])
        self.action_completa.setText(t["menu_completa"])
        self.action_jobs.setText(t["menu_jobs"])
        self.action_params.setText(t["menu_params"])
        self.action_uf_tools.setText(t["menu_uf_tools"])
        self.action_usrgrp.setText(t.get("menu_usrgrp", "User Group"))
        self.action_cubeintf.setText(t.get("menu_cubeintf", "Cubo interferenza"))
        self.action_formcut.setText(t.get("menu_formcut", "FormCut"))
        self.menu_backup.setTitle(t["menu_backup"])
        self.action_backup_var.setText(t["menu_backup_var"])
        self.action_backup_punti.setText(t["menu_backup_punti"])
        self.menu_drive.setTitle(t["menu_drive"])
        self.action_motion_drive.setText(t["menu_motion_drive"])
        self.menu_help.setTitle(t["menu_help"])
        self.action_ga500_params.setText(t["menu_ga500_params"])
        self.action_exit.setText(t["btn_exit"])
        self.action_log.setText(t["menu_log"])
        self.action_help_params.setText(t["menu_help_params"])
        self.action_help_known.setText(t["menu_help_known"])
        self.action_about.setText(t["menu_about"])
        self.action_aiuto.setText(t.get("menu_aiuto", "Aiuto"))
        self.log_header.setText(t.get("log_panel_title", "System Logs"))
        # Folder bar
        self._folder_lbl_key.setText(t.get("work_folder_label", "Cartella robot:"))
        self._folder_btn.setText(t.get("work_folder_btn", "Cartella robot"))
        if not self._work_folder:
            self._folder_lbl_path.setText(t.get("work_folder_none", "Nessuna cartella selezionata"))
        # Welcome label
        self._welcome_lbl.setText(t.get("welcome_message", ""))
        # Open panels
        if self._tool_panel is not None:
            self._tool_panel.update_language(lang)
        if self._help_view is not None:
            self._help_view.update_language(lang)
        if self._uf_tools_view is not None:
            self._uf_tools_view.update_language(lang)
        if self._backup_view is not None:
            self._backup_view.update_language(lang)
        if self._drive_view is not None:
            self._drive_view.update_language(lang)
        if self._ga500_view is not None:
            self._ga500_view.update_language(lang)
        if self._pdf_preview is not None:
            self._pdf_preview.update_language(lang)
            self._progress_begin()
            try:
                self._pdf_preview.regen()
            finally:
                self._progress_end()
        if self._excel_preview is not None:
            self._excel_preview.update_language(lang)
        if self._ipnet_view is not None:
            self._ipnet_view.update_language(lang)
        if self._uframe_view is not None:
            self._uframe_view.update_language(lang)
        if self._aiuto_view is not None:
            self._aiuto_view.update_language(lang)
        if self._compila_view is not None:
            self._compila_view.update_language(lang)
        if self._ifpanel_view is not None:
            self._ifpanel_view.update_language(lang)
        if self._flowchart_view is not None:
            self._flowchart_view.update_language(lang)
        if self._usrgrp_view is not None:
            self._usrgrp_view.update_language(lang)
        # Refresh tooltips so they match the new language labels.
        # Clear any previously set tooltips first; _apply_tooltips skips
        # widgets that already have a tooltip, so a clear-then-reapply is
        # needed when texts change.
        try:
            from PySide6.QtWidgets import QMenu, QPushButton
            for m in self.menu_bar.findChildren(QMenu):
                for act in m.actions():
                    act.setToolTip("")
            for btn in self.findChildren(QPushButton):
                btn.setToolTip("")
        except Exception:
            pass
        self._apply_tooltips()

    def on_flowchart(self):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS.get(self.app_state.language, TRANSLATIONS["IT"])
        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self,
            t.get("flowchart_select_folder", "Seleziona cartella JBI"),
            "",
            QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_flowchart", "Flowchart"))
            return
        self._close_all_panels()
        self._hide_welcome()
        try:
            from gui.flowchart_view import FlowchartView
            view = FlowchartView(folder, self.app_state,
                                 on_close_cb=self._close_flowchart_view)
            self._flowchart_view = view
            self.content_area.layout().addWidget(view)
            logger.info("log_flowchart_opened")
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def on_targhetta(self):
        import shutil
        import tempfile
        from PySide6.QtWidgets import QFileDialog

        t = TRANSLATIONS[self.app_state.language]

        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_targhetta"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_targhetta')} — {folder}")

        folder_name = os.path.basename(os.path.normpath(folder))

        if not os.path.isfile(os.path.join(folder, "SYSTEM.SYS")):
            logger.warning("log_file_not_found", "SYSTEM.SYS")
            return

        default_name = f"Targhetta_{folder_name}.pdf"
        tmp_path = secure_paths.temp_path(default_name)
        self._progress_begin()
        try:
            from docs.targhetta import generate_pdf
            generate_pdf(folder, tmp_path, lang=self.app_state.language)
            logger.info("log_pdf_generated", tmp_path)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            self._progress_end()
            return
        self._progress_end()

        def _save():
            sp, _ = QFileDialog.getSaveFileName(
                self, t["dialog_save_pdf"],
                os.path.join(os.path.expanduser("~"), "Desktop", default_name),
                "PDF Files (*.pdf)")
            if not sp:
                logger.info("log_cancelled", f"{t.get('menu_targhetta')} — PDF")
                return
            try:
                shutil.copy2(tmp_path, sp)
                logger.info("log_pdf_saved", sp)
                if self._pdf_preview is not None:
                    self._pdf_preview.on_pdf_saved(os.path.dirname(os.path.abspath(sp)))
            except Exception as exc:
                logger.error("log_error_generic", str(exc))

        def _regen():
            from docs.targhetta import generate_pdf as _gp
            _gp(folder, tmp_path, lang=self.app_state.language)

        self._show_pdf_preview(tmp_path, on_save_cb=_save, regen_fn=_regen)

    def on_panel(self):
        import shutil
        import tempfile
        from PySide6.QtWidgets import QFileDialog

        t = TRANSLATIONS[self.app_state.language]

        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_panel"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_panel')} — {folder}")

        folder_name = os.path.basename(os.path.normpath(folder))

        if not os.path.isfile(os.path.join(folder, "PANELBOX.LOG")):
            logger.warning("log_file_not_found", "PANELBOX.LOG")
            return

        default_name = f"Panel_{folder_name}.pdf"
        tmp_path = secure_paths.temp_path(default_name)
        self._progress_begin()
        try:
            from docs.panel import generate_pdf
            generate_pdf(folder, tmp_path, lang=self.app_state.language)
            logger.info("log_pdf_generated", tmp_path)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            self._progress_end()
            return
        self._progress_end()

        def _save():
            sp, _ = QFileDialog.getSaveFileName(
                self, t["dialog_save_pdf"],
                os.path.join(os.path.expanduser("~"), "Desktop", default_name),
                "PDF Files (*.pdf)")
            if not sp:
                logger.info("log_cancelled", f"{t.get('menu_panel')} — PDF")
                return
            try:
                shutil.copy2(tmp_path, sp)
                logger.info("log_pdf_saved", sp)
                if self._pdf_preview is not None:
                    self._pdf_preview.on_pdf_saved(os.path.dirname(os.path.abspath(sp)))
            except Exception as exc:
                logger.error("log_error_generic", str(exc))

        def _regen():
            from docs.panel import generate_pdf as _gp
            _gp(folder, tmp_path, lang=self.app_state.language)

        self._show_pdf_preview(tmp_path, on_save_cb=_save, regen_fn=_regen)

    def _on_compila_genera(self):
        t = TRANSLATIONS[self.app_state.language]
        logger.info("log_btn_pressed", t.get("btn_genera", "Genera"))

    def on_completa(self):
        import shutil
        import tempfile
        from PySide6.QtWidgets import QFileDialog

        t = TRANSLATIONS[self.app_state.language]

        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_completa"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_completa')} — {folder}")

        folder_name = os.path.basename(os.path.normpath(folder))

        dlg = _CompletaDialog(self, self.app_state, folder)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            logger.info("log_cancelled", t.get("menu_completa"))
            return

        attachments = dlg.attachments

        default_name = f"Documentazione_{folder_name}.pdf"
        tmp_path = secure_paths.temp_path(default_name)

        self._progress_begin()
        try:
            from docs.completa import generate_completa
            generate_completa(folder, tmp_path, attachments=attachments,
                              lang=self.app_state.language, log_fn=logger.info,
                              progress_fn=self._progress_update)
            logger.info("log_completa_generated", tmp_path)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            self._progress_end()
            return
        self._progress_end()

        def _save():
            sp, _ = QFileDialog.getSaveFileName(
                self, t["dialog_save_pdf"],
                os.path.join(os.path.expanduser("~"), "Desktop", default_name),
                "PDF Files (*.pdf)")
            if not sp:
                logger.info("log_cancelled", f"{t.get('menu_completa')} — PDF")
                return
            try:
                shutil.copy2(tmp_path, sp)
                logger.info("log_completa_saved", sp)
                if self._pdf_preview is not None:
                    self._pdf_preview.on_pdf_saved(os.path.dirname(os.path.abspath(sp)))
            except Exception as exc:
                logger.error("log_error_generic", str(exc))

        def _regen():
            from docs.completa import generate_completa as _gc
            _gc(folder, tmp_path, attachments=attachments,
                lang=self.app_state.language, log_fn=logger.info,
                progress_fn=self._progress_update)

        self._show_pdf_preview(tmp_path, on_save_cb=_save, regen_fn=_regen)

    def on_template(self):
        import shutil
        import tempfile
        from PySide6.QtWidgets import QFileDialog

        t = TRANSLATIONS[self.app_state.language]

        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "",
            QFileDialog.Option.ShowDirsOnly
        )
        if not folder:
            logger.info("log_cancelled", t.get("menu_template"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_template')} — {folder}")

        folder_name = os.path.basename(os.path.normpath(folder))
        default_name = f"Template_Nomi_{folder_name}.xlsx"

        save_path, _ = QFileDialog.getSaveFileName(
            self, t["dialog_save_pdf"],
            os.path.join(os.path.expanduser("~"), "Desktop", default_name),
            "Excel Files (*.xlsx)"
        )
        if not save_path:
            logger.info("log_cancelled", f"{t.get('menu_template')} — Excel")
            return

        tmp_path = secure_paths.temp_path(default_name)
        try:
            from docs.names import generate_template
            generate_template(folder, tmp_path, log_fn=logger.info)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        try:
            shutil.copy2(tmp_path, save_path)
            os.remove(tmp_path)
            logger.info("log_pdf_saved", save_path)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        # Anteprima Excel del template generato
        try:
            sheets_data, compatible = self._load_xl_preview(save_path)
            if compatible:
                self._show_template_excel_preview(save_path, sheets_data)
        except Exception as exc:
            logger.warning("log_error_generic", str(exc))

    # ── Excel preview helpers ─────────────────────────────────────────────────

    def _load_xl_preview(self, excel_path):
        """Open Excel, return (sheets_data, is_compatible).
        sheets_data = [(name, headers, named_rows, total), ...]
        Raises on file errors."""
        from docs.names import _VAR_TABS, _IO_TABS
        import openpyxl

        EXPECTED = set(_VAR_TABS) | {tab for tab, _, _ in _IO_TABS}
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        found = [s for s in wb.sheetnames if s in EXPECTED]
        if not found:
            wb.close()
            return [], False

        sheets_data = []
        for sheet_name in found:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            hdr = next(rows_iter, None)
            headers = [str(h) if h is not None else "" for h in (hdr or [])]
            named_rows, total = [], 0
            for row in rows_iter:
                total += 1
                if len(row) > 1 and row[1]:
                    named_rows.append(row)
            sheets_data.append((sheet_name, headers, named_rows, total))

        wb.close()
        return sheets_data, True

    def _load_xl_compila(self, excel_path):
        """Load ALL non-empty rows for the editable Compila view (not just named ones).
        Raises on file errors."""
        from docs.names import _VAR_TABS, _IO_TABS
        import openpyxl

        EXPECTED = set(_VAR_TABS) | {tab for tab, _, _ in _IO_TABS}
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        found = [s for s in wb.sheetnames if s in EXPECTED]
        if not found:
            wb.close()
            return [], False

        sheets_data = []
        for sheet_name in found:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            hdr = next(rows_iter, None)
            headers = [str(h) if h is not None else "" for h in (hdr or [])]
            all_rows = [r for r in rows_iter if any(v is not None for v in r)]
            sheets_data.append((sheet_name, headers, all_rows, len(all_rows)))

        wb.close()
        return sheets_data, True

    def _show_preview(self, excel_path, sheets_data):
        self._close_all_panels()
        self._hide_welcome()
        preview = _ExcelPreview(
            excel_path, sheets_data, self.app_state,
            on_close_cb=self._close_preview,
            on_generate_cb=lambda: self._run_dat_generation(excel_path),
        )
        self._excel_preview = preview
        self.content_area.layout().addWidget(preview)

    def _close_preview(self):
        if self._excel_preview is not None:
            self._excel_preview.setParent(None)
            self._excel_preview.deleteLater()
            self._excel_preview = None
            logger.info("log_preview_closed")

    def _run_dat_generation(self, excel_path):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        out_folder = QFileDialog.getExistingDirectory(
            self, t["select_folder"], "",
            QFileDialog.Option.ShowDirsOnly
        )
        if not out_folder:
            logger.info("log_cancelled", f"{t.get('menu_nomi')} — {t.get('dialog_export_folder', 'output')}")
            return
        logger.info("log_btn_pressed", f"Genera DAT - output: {out_folder}")
        try:
            from docs.names import generate_dat_files
            generate_dat_files(excel_path, out_folder, log_fn=logger.info)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return
        try:
            os.startfile(out_folder)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def on_nomi(self):
        import tempfile
        from PySide6.QtWidgets import QFileDialog

        t = TRANSLATIONS[self.app_state.language]
        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_nomi"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_nomi')} — {folder}")

        folder_name = os.path.basename(os.path.normpath(folder))
        tmp_path = secure_paths.temp_path(f"Nomi_{folder_name}.xlsx")

        try:
            from docs.names import generate_template
            generate_template(folder, tmp_path, log_fn=logger.info)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        try:
            sheets_data, compatible = self._load_xl_compila(tmp_path)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        if not compatible:
            logger.warning("log_incompatible_excel", os.path.basename(tmp_path))
            return

        self._close_all_panels()
        self._hide_welcome()
        preview = _ExcelPreview(
            tmp_path, sheets_data, self.app_state,
            on_close_cb=self._close_preview,
            on_generate_cb=lambda: self._run_nomi_generation(tmp_path),
            nomi_mode=True,
        )
        self._excel_preview = preview
        self.content_area.layout().addWidget(preview)
        logger.info("log_preview_opened", os.path.basename(tmp_path))

    def _run_nomi_generation(self, tmp_path):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        try:
            if self._excel_preview is not None:
                self._excel_preview.save_edits(tmp_path)
        except Exception as exc:
            logger.warning("log_error_generic", str(exc))
        out_folder = QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not out_folder:
            logger.info("log_cancelled", f"{t.get('menu_nomi')} — {t.get('dialog_export_folder', 'output')}")
            return
        logger.info("log_btn_pressed", f"Genera DAT - output: {out_folder}")
        try:
            from docs.names import generate_dat_files
            generate_dat_files(tmp_path, out_folder, log_fn=logger.info)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def _close_tool_panel(self):
        if self._tool_panel is not None:
            self._tool_panel.save_state()
            self._tool_panel.setParent(None)
            self._tool_panel.deleteLater()
            self._tool_panel = None
            logger.info("log_tool_closed")

    def _close_help_view(self):
        if self._help_view is not None:
            self._help_view.setParent(None)
            self._help_view.deleteLater()
            self._help_view = None
            logger.info("log_help_closed")

    def _close_uf_tools_view(self):
        if self._uf_tools_view is not None:
            self._uf_tools_view.setParent(None)
            self._uf_tools_view.deleteLater()
            self._uf_tools_view = None
            logger.info("log_uf_tools_closed")

    def _close_backup_view(self):
        if self._backup_view is not None:
            self._backup_view.setParent(None)
            self._backup_view.deleteLater()
            self._backup_view = None
            logger.info("log_backup_closed")

    def _close_drive_view(self):
        if self._drive_view is not None:
            self._drive_view.setParent(None)
            self._drive_view.deleteLater()
            self._drive_view = None
            logger.info("log_drive_closed")

    def _close_ga500_view(self):
        if self._ga500_view is not None:
            self._ga500_view.setParent(None)
            self._ga500_view.deleteLater()
            self._ga500_view = None
            logger.info("log_ga500_closed")

    def _close_ipnet_view(self):
        if self._ipnet_view is not None:
            self._ipnet_view.setParent(None)
            self._ipnet_view.deleteLater()
            self._ipnet_view = None
            logger.info("log_ipnet_closed")

    def _close_uframe_view(self):
        if self._uframe_view is not None:
            self._uframe_view.setParent(None)
            self._uframe_view.deleteLater()
            self._uframe_view = None
            logger.info("log_uframe_closed")

    def _close_aiuto_view(self):
        if self._aiuto_view is not None:
            self._aiuto_view.setParent(None)
            self._aiuto_view.deleteLater()
            self._aiuto_view = None

    def _close_ifpanel_view(self):
        if self._ifpanel_view is not None:
            self._ifpanel_view.setParent(None)
            self._ifpanel_view.deleteLater()
            self._ifpanel_view = None
            logger.info("log_ifpanel_closed")

    def _close_flowchart_view(self):
        if self._flowchart_view is not None:
            self._flowchart_view.setParent(None)
            self._flowchart_view.deleteLater()
            self._flowchart_view = None

    def _close_usrgrp_view(self):
        if self._usrgrp_view is not None:
            self._usrgrp_view.setParent(None)
            self._usrgrp_view.deleteLater()
            self._usrgrp_view = None
            logger.info("log_usrgrp_closed")

    def _close_compila_view(self):
        if self._compila_view is not None:
            self._compila_view.setParent(None)
            self._compila_view.deleteLater()
            self._compila_view = None

    def _show_welcome(self):
        if self._welcome_lbl is not None:
            self._welcome_lbl.setVisible(True)

    def _hide_welcome(self):
        if self._welcome_lbl is not None:
            self._welcome_lbl.setVisible(False)

    def _close_pdf_preview(self):
        if self._pdf_preview is not None:
            self._pdf_preview.setParent(None)
            self._pdf_preview.deleteLater()
            self._pdf_preview = None
            logger.info("log_pdf_preview_closed")

    def _show_pdf_preview(self, pdf_path, on_save_cb=None, on_excel_cb=None,
                          nav_items=None, regen_fn=None):
        self._close_all_panels()
        self._hide_welcome()
        # Show the progress bar for EVERY preview (uniform with "Completa").
        # Loading/rendering the PDF document is the visible work here, especially
        # for large documents. try/finally guarantees the bar is always hidden.
        self._progress_begin()
        try:
            self._progress_update(40)
            preview = _PdfPreview(
                pdf_path, self.app_state,
                on_close_cb=self._close_pdf_preview,
                on_save_cb=on_save_cb,
                on_excel_cb=on_excel_cb,
                nav_items=nav_items,
                regen_fn=regen_fn,
            )
            self._pdf_preview = preview
            self.content_area.layout().addWidget(preview)
            self._progress_update(90)
            logger.info("log_pdf_preview_opened", os.path.basename(pdf_path))
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
        finally:
            self._progress_end()

    def _show_template_excel_preview(self, excel_path, sheets_data):
        self._close_all_panels()
        self._hide_welcome()
        try:
            preview = _ExcelPreview(
                excel_path, sheets_data, self.app_state,
                on_close_cb=self._close_preview,
                on_generate_cb=None,
            )
            self._excel_preview = preview
            self.content_area.layout().addWidget(preview)
            logger.info("log_preview_opened", os.path.basename(excel_path))
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def _show_compila_preview(self, excel_path, sheets_data):
        t = TRANSLATIONS[self.app_state.language]
        valore_label = t.get("backup_col_value", "Valore")
        mapped_data = [
            (sname,
             [valore_label if h.strip().lower() == "note" else h for h in hdrs],
             rows, total)
            for sname, hdrs, rows, total in sheets_data
        ]
        self._close_all_panels()
        self._hide_welcome()
        try:
            preview = _ExcelPreview(
                excel_path, mapped_data, self.app_state,
                on_close_cb=self._close_preview,
                on_generate_cb=self._on_compila_genera,
                editable=True,
            )
            self._excel_preview = preview
            self.content_area.layout().addWidget(preview)
            logger.info("log_preview_opened", os.path.basename(excel_path))
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def _close_all_panels(self, replaced_by=""):
        """Close any open content panel and show welcome label."""
        t = TRANSLATIONS[self.app_state.language]
        if self._pdf_preview is not None:
            logger.info("log_panel_replaced", t.get("preview_pdf_info", "PDF"))
            self._close_pdf_preview()
        if self._excel_preview is not None:
            logger.info("log_panel_replaced", t.get("menu_nomi", "Nomi"))
            self._close_preview()
        if self._backup_view is not None:
            logger.info("log_panel_replaced", t.get("menu_backup", "Backup"))
            self._close_backup_view()
        if self._tool_panel is not None:
            logger.info("log_panel_replaced", t.get("menu_tool", "Tool"))
            self._close_tool_panel()
        if self._help_view is not None:
            logger.info("log_panel_replaced", t.get("menu_help", "Help"))
            self._close_help_view()
        if self._uf_tools_view is not None:
            logger.info("log_panel_replaced", t.get("menu_uf_tools", "UF#()/Tools"))
            self._close_uf_tools_view()
        if self._drive_view is not None:
            logger.info("log_panel_replaced", t.get("menu_drive", "Drive"))
            self._close_drive_view()
        if self._ga500_view is not None:
            logger.info("log_panel_replaced", t.get("menu_ga500_params", "GA500"))
            self._close_ga500_view()
        if self._ipnet_view is not None:
            logger.info("log_panel_replaced", t.get("menu_ipnet", "IP"))
            self._close_ipnet_view()
        if self._compila_view is not None:
            logger.info("log_panel_replaced", t.get("menu_compila", "Compila"))
            self._close_compila_view()
        if self._uframe_view is not None:
            logger.info("log_panel_replaced", t.get("menu_uf", "UF#()"))
            self._close_uframe_view()
        if self._aiuto_view is not None:
            self._close_aiuto_view()
        if self._ifpanel_view is not None:
            logger.info("log_panel_replaced", t.get("menu_ifpanel", "IFPanel"))
            self._close_ifpanel_view()
        if self._flowchart_view is not None:
            logger.info("log_panel_replaced", t.get("menu_flowchart", "Flowchart"))
            self._close_flowchart_view()
        if self._usrgrp_view is not None:
            logger.info("log_panel_replaced", t.get("menu_usrgrp", "User Group"))
            self._close_usrgrp_view()
        self._show_welcome()

    def on_tool(self):
        self._close_all_panels()
        self._hide_welcome()
        from gui.tool_panel import ToolPanel
        panel = ToolPanel(
            self.app_state,
            on_close_cb=self._close_tool_panel,
        )
        self._tool_panel = panel
        self.content_area.layout().addWidget(panel)
        logger.info("log_tool_opened")

    def on_uf(self):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", "UF#()")
            return
        logger.info("log_btn_pressed", f"UF#() — {folder}")
        from docs.uf_tools import UFRAME_FILE
        if not os.path.isfile(os.path.join(folder, UFRAME_FILE)):
            logger.warning("log_uframe_no_file")
            return
        self._close_all_panels()
        self._hide_welcome()
        try:
            from gui.uframe_view import UFrameView
            view = UFrameView(folder, self.app_state,
                              on_close_cb=self._close_uframe_view)
            self._uframe_view = view
            self.content_area.layout().addWidget(view)
            logger.info("log_uframe_opened")
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def on_ifpanel(self):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", "IFPanel")
            return
        import os as _os
        filepath = _os.path.join(folder, 'IFPANEL.DAT')
        if not _os.path.isfile(filepath):
            logger.warning("log_file_not_found", "IFPANEL.DAT")
            return
        self._close_all_panels()
        self._hide_welcome()
        try:
            from gui.ifpanel_view import IFPanelView
            view = IFPanelView(filepath, self.app_state,
                               on_close_cb=self._close_ifpanel_view)
            self._ifpanel_view = view
            self.content_area.layout().addWidget(view)
            logger.info("log_ifpanel_opened")
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def on_aiuto(self):
        self._close_all_panels()
        self._hide_welcome()
        logger.info("log_aiuto_opened")
        try:
            view = _AiutoView(self.app_state, on_close_cb=self._close_aiuto_view)
            self._aiuto_view = view
            self.content_area.layout().addWidget(view)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def on_usrgrp(self):
        import tempfile
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]

        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_usrgrp"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_usrgrp')} — {folder}")

        from docs.usrgrp import (parse_gpin, parse_gpot, generate_pdf,
                                  GPIN_FILE, GPOT_FILE)

        gpin_path = os.path.join(folder, GPIN_FILE)
        gpot_path = os.path.join(folder, GPOT_FILE)

        gpin_groups = []
        gpot_groups = []

        if os.path.isfile(gpin_path):
            try:
                gpin_groups = parse_gpin(folder)
                logger.info("log_usrgrp_gpin_count", len(gpin_groups))
            except Exception as exc:
                logger.warning("log_error_generic", str(exc))
        else:
            logger.warning("log_file_not_found", GPIN_FILE)

        if os.path.isfile(gpot_path):
            try:
                gpot_groups = parse_gpot(folder)
                logger.info("log_usrgrp_gpot_count", len(gpot_groups))
            except Exception as exc:
                logger.warning("log_error_generic", str(exc))
        else:
            logger.warning("log_file_not_found", GPOT_FILE)

        if not gpin_groups and not gpot_groups:
            logger.warning("log_file_not_found",
                           t.get("usrgrp_no_file", "USRGRPIN.DAT / USRGRPOT.DAT"))
            return

        folder_name = os.path.basename(os.path.normpath(folder))
        tmp_path = secure_paths.temp_path(f"UserGroup_{folder_name}.pdf")

        self._progress_begin()
        try:
            generate_pdf(gpin_groups, gpot_groups, tmp_path,
                         lang=self.app_state.language,
                         progress_fn=self._progress_update)
            logger.info("log_pdf_generated", tmp_path)
        except Exception as exc:
            logger.warning("log_error_generic", str(exc))
            self._progress_end()
            return
        self._progress_end()

        self._close_all_panels()
        self._hide_welcome()
        try:
            from gui.usrgrp_view import UsrGrpView
            view = UsrGrpView(
                folder, gpin_groups, gpot_groups, tmp_path,
                self.app_state,
                on_close_cb=self._close_usrgrp_view,
                progress_begin_fn=self._progress_begin,
                progress_end_fn=self._progress_end,
            )
            self._usrgrp_view = view
            self.content_area.layout().addWidget(view)
            logger.info("log_usrgrp_opened")
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def _on_uf_tools_generate(self, folder):
        import shutil
        import tempfile
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        folder_name = os.path.basename(os.path.normpath(folder))
        default_name = f"UF_Tools_{folder_name}.pdf"
        tmp_path = secure_paths.temp_path(default_name)
        self._progress_begin()
        try:
            from docs.uf_tools import generate_pdf
            generate_pdf(folder, tmp_path,
                         lang=self.app_state.language,
                         log_fn=logger.info)
            logger.info("log_pdf_generated", tmp_path)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            self._progress_end()
            return
        self._progress_end()

        def _save():
            sp, _ = QFileDialog.getSaveFileName(
                self, t["dialog_save_pdf"],
                os.path.join(os.path.expanduser("~"), "Desktop", default_name),
                "PDF Files (*.pdf)")
            if not sp:
                logger.info("log_cancelled", "UF#()/Tools - salva PDF")
                return
            try:
                shutil.copy2(tmp_path, sp)
                logger.info("log_pdf_saved", sp)
                if self._pdf_preview is not None:
                    self._pdf_preview.on_pdf_saved(os.path.dirname(os.path.abspath(sp)))
            except Exception as exc:
                logger.error("log_error_generic", str(exc))

        from PySide6.QtCore import QTimer
        QTimer.singleShot(0, lambda: self._show_pdf_preview(tmp_path, on_save_cb=_save))

    def on_help_params(self):
        self._close_all_panels()
        self._hide_welcome()
        try:
            from gui.help_view import HelpView
            view = HelpView(self.app_state, "params", self._close_help_view)
            self._help_view = view
            self.content_area.layout().addWidget(view)
            logger.info("log_help_params_opened")
        except Exception as exc:
            logger.warning("log_error_generic", str(exc))

    def on_help_known(self):
        self._close_all_panels()
        self._hide_welcome()
        try:
            from gui.help_view import HelpView
            view = HelpView(self.app_state, "known", self._close_help_view)
            self._help_view = view
            self.content_area.layout().addWidget(view)
            logger.info("log_help_known_opened")
        except Exception as exc:
            logger.warning("log_error_generic", str(exc))

    def on_jobs(self):
        import shutil
        import tempfile
        from PySide6.QtWidgets import QFileDialog

        t = TRANSLATIONS[self.app_state.language]

        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_jobs"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_jobs')} — {folder}")

        folder_name = os.path.basename(os.path.normpath(folder))
        default_name = f"JOBs_{folder_name}.pdf"
        tmp_path = secure_paths.temp_path(default_name)
        nav_items = None
        self._progress_begin()
        try:
            from docs.jobs import generate_pdf
            nav_items = generate_pdf(folder, tmp_path, log_fn=logger.info,
                                     lang=self.app_state.language)
            logger.info("log_pdf_generated", tmp_path)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            self._progress_end()
            return
        self._progress_end()

        def _save():
            sp, _ = QFileDialog.getSaveFileName(
                self, t["dialog_save_pdf"],
                os.path.join(os.path.expanduser("~"), "Desktop", default_name),
                "PDF Files (*.pdf)")
            if not sp:
                logger.info("log_cancelled", f"{t.get('menu_jobs')} — PDF")
                return
            try:
                shutil.copy2(tmp_path, sp)
                logger.info("log_pdf_saved", sp)
                if self._pdf_preview is not None:
                    self._pdf_preview.on_pdf_saved(os.path.dirname(os.path.abspath(sp)))
            except Exception as exc:
                logger.error("log_error_generic", str(exc))

        def _regen():
            from docs.jobs import generate_pdf as _gp
            return _gp(folder, tmp_path, log_fn=None, lang=self.app_state.language)

        self._show_pdf_preview(tmp_path, on_save_cb=_save, nav_items=nav_items,
                               regen_fn=_regen)

    def on_params(self):
        import shutil
        import tempfile
        from PySide6.QtWidgets import QFileDialog

        t = TRANSLATIONS[self.app_state.language]

        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_params"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_params')} — {folder}")

        folder_name = os.path.basename(os.path.normpath(folder))
        default_name = f"Parametri_{folder_name}.pdf"
        tmp_path = secure_paths.temp_path(default_name)
        self._progress_begin()
        try:
            from docs.params import generate_pdf
            generate_pdf(folder, tmp_path, log_fn=logger.info,
                         lang=self.app_state.language)
            logger.info("log_pdf_generated", tmp_path)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            self._progress_end()
            return
        self._progress_end()

        def _save():
            sp, _ = QFileDialog.getSaveFileName(
                self, t["dialog_save_pdf"],
                os.path.join(os.path.expanduser("~"), "Desktop", default_name),
                "PDF Files (*.pdf)")
            if not sp:
                logger.info("log_cancelled", f"{t.get('menu_params')} — PDF")
                return
            try:
                shutil.copy2(tmp_path, sp)
                logger.info("log_pdf_saved", sp)
                if self._pdf_preview is not None:
                    self._pdf_preview.on_pdf_saved(os.path.dirname(os.path.abspath(sp)))
            except Exception as exc:
                logger.error("log_error_generic", str(exc))

        def _regen():
            from docs.params import generate_pdf as _gp
            _gp(folder, tmp_path, log_fn=None, lang=self.app_state.language)

        def _excel():
            self._on_params_excel(folder)

        self._show_pdf_preview(tmp_path, on_save_cb=_save, on_excel_cb=_excel,
                               regen_fn=_regen)

    def _on_params_excel(self, folder):
        import tempfile
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        folder_name = os.path.basename(os.path.normpath(folder))
        default_name = f"Parametri_{folder_name}.xlsx"
        sp, _ = QFileDialog.getSaveFileName(
            self, t.get("dialog_save_excel", "Salva Excel"),
            os.path.join(os.path.expanduser("~"), "Desktop", default_name),
            "Excel Files (*.xlsx)")
        if not sp:
            logger.info("log_cancelled", f"{t.get('menu_params')} — Excel")
            return
        try:
            from docs.params import export_excel
            export_excel(folder, sp, lang=self.app_state.language)
            logger.info("log_pdf_saved", sp)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def on_uf_tools(self):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", "UF#()/Tools")
            return
        logger.info("log_btn_pressed", f"UF#()/Tools — {folder}")
        self._close_all_panels()
        self._hide_welcome()
        try:
            from gui.uf_tools_view import UfToolsView
            view = UfToolsView(
                self.app_state,
                folder,
                on_close_cb=self._close_uf_tools_view,
                on_generate_cb=lambda: self._on_uf_tools_generate(folder),
                on_excel_cb=lambda: self._on_uf_tools_excel(folder),
            )
            self._uf_tools_view = view
            self.content_area.layout().addWidget(view)
            logger.info("log_uf_tools_opened")
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def _on_uf_tools_excel(self, folder):
        import tempfile
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        folder_name = os.path.basename(os.path.normpath(folder))
        default_name = f"UF_Tools_{folder_name}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self, t.get("dialog_save_excel", "Salva Excel"),
            os.path.join(os.path.expanduser("~"), "Desktop", default_name),
            "Excel Files (*.xlsx)")
        if not save_path:
            logger.info("log_cancelled", "UF#()/Tools - salva Excel")
            return
        try:
            from docs.uf_tools import generate_uf_excel
            n_tools, n_frames = generate_uf_excel(
                folder, save_path,
                lang=self.app_state.language,
                log_fn=logger.info)
            logger.info("log_uf_excel_saved", save_path)
            try:
                os.startfile(os.path.dirname(os.path.abspath(save_path)))
            except Exception:
                pass
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def on_backup_var(self):
        from PySide6.QtWidgets import QFileDialog

        t = TRANSLATIONS[self.app_state.language]

        # 1 – Select template Excel
        template_path, _ = QFileDialog.getOpenFileName(
            self, t["backup_select_template"], "",
            "Excel Files (*.xlsx)"
        )
        if not template_path:
            logger.info("log_cancelled", f"{t.get('menu_backup_var')} — template")
            return
        logger.info("log_btn_pressed", f"Backup VAR - template: {template_path}")

        # 2 – Select robot folder
        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_backup_var"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_backup_var')} — {folder}")

        # 3 – Check VAR.DAT ([WARNING] + return if missing)
        var_dat_path = os.path.join(folder, "VAR.DAT")
        if not os.path.isfile(var_dat_path):
            logger.warning("log_backup_no_var_dat", folder)
            return

        # 4 – Load data into memory
        folder_name = os.path.basename(os.path.normpath(folder))
        try:
            from docs.backup import load_backup_data
            tab_data, sections = load_backup_data(template_path, folder, log_fn=logger.info)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        if tab_data is None:
            return  # error already logged in load_backup_data

        # 5 – Show _BackupView
        self._close_all_panels()
        self._hide_welcome()
        try:
            view = _BackupView(
                tab_data, folder_name, self.app_state,
                on_close_cb=self._close_backup_view,
                on_generate_excel_cb=lambda: self._on_backup_generate_excel(
                    tab_data, template_path, sections, folder_name),
                on_export_pdf_cb=lambda: self._on_backup_export_pdf(tab_data, folder_name),
            )
            self._backup_view = view
            self.content_area.layout().addWidget(view)
            logger.info("log_backup_opened", folder_name)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def _on_backup_generate_excel(self, tab_data, template_path, sections, folder_name):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]

        default_name = f"Backup_VAR_{folder_name}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self, t["dialog_save_excel"],
            os.path.join(os.path.expanduser("~"), "Desktop", default_name),
            "Excel Files (*.xlsx)"
        )
        if not save_path:
            logger.info("log_cancelled", f"{t.get('menu_backup_var')} — Excel")
            return

        try:
            from docs.backup import generate_backup_excel_from_template
            ok = generate_backup_excel_from_template(
                template_path, sections, save_path,
                col_value_label=t.get("backup_col_value", "Valore"),
                log_fn=logger.info,
            )
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        if not ok:
            return

        try:
            os.startfile(os.path.dirname(os.path.abspath(save_path)))
        except Exception as exc:
            logger.warning("log_error_generic", str(exc))

    def _on_backup_export_pdf(self, tab_data, folder_name):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]

        default_name = f"Backup_VAR_{folder_name}.pdf"
        save_path, _ = QFileDialog.getSaveFileName(
            self, t["dialog_save_pdf"],
            os.path.join(os.path.expanduser("~"), "Desktop", default_name),
            "PDF Files (*.pdf)"
        )
        if not save_path:
            logger.info("log_cancelled", f"{t.get('menu_backup_var')} — PDF")
            return

        try:
            from docs.backup import generate_backup_pdf
            ok = generate_backup_pdf(
                tab_data, save_path,
                folder_name=folder_name,
                col_id_label="ID",
                col_name_label=t.get("backup_col_name", "Nome"),
                col_value_label=t.get("backup_col_value", "Valore"),
                log_fn=logger.info,
                lang=self.app_state.language,
            )
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        if not ok:
            return

        try:
            os.startfile(os.path.dirname(os.path.abspath(save_path)))
        except Exception as exc:
            logger.warning("log_error_generic", str(exc))

    def on_backup_punti(self):
        import shutil
        import tempfile
        from PySide6.QtWidgets import QFileDialog

        t = TRANSLATIONS[self.app_state.language]

        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_backup_punti"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_backup_punti')} — {folder}")

        try:
            from docs.points import find_jbi_files, parse_jbi_points, generate_points_pdf
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        jbi_files = find_jbi_files(folder)
        if not jbi_files:
            logger.warning("log_punti_no_jbi", folder)
            return

        jbi_data = []
        for filepath in jbi_files:
            fname = os.path.splitext(os.path.basename(filepath))[0]
            try:
                points = parse_jbi_points(filepath)
            except RuntimeError as exc:
                logger.error("log_error_generic", str(exc))
                points = []
            if points:
                logger.info("log_punti_file", fname, len(points))
            else:
                logger.info("log_punti_no_points", fname)
            jbi_data.append((fname, points))

        # Load UFRAME.CND names if available
        uframe_names = {}
        uf_cnd_path = os.path.join(folder, "UFRAME.CND")
        if os.path.isfile(uf_cnd_path):
            try:
                from docs.uf_tools import parse_uframe_cnd
                frames = parse_uframe_cnd(uf_cnd_path)
                uframe_names = {f['num']: f['name'] for f in frames if f.get('name')}
            except Exception as exc:
                logger.warning("log_error_generic", str(exc))

        folder_name = os.path.basename(os.path.normpath(folder))
        default_name = f"Punti_{folder_name}.pdf"
        tmp_path = secure_paths.temp_path(default_name)

        self._progress_begin()
        try:
            ok = generate_points_pdf(
                jbi_data, tmp_path,
                col_point=t.get("punti_col_point", "Point"),
                col_tool=t.get("punti_col_tool", "Tool"),
                col_uf=t.get("punti_col_uf", "UF#()"),
                col_uf_name=t.get("punti_col_uf_name", "UF Name"),
                col_type=t.get("punti_col_type", "Type"),
                msg_no_pos=t.get("punti_no_positions", "No positions found"),
                uframe_names=uframe_names,
                lang=self.app_state.language,
                log_fn=logger.info,
            )
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            self._progress_end()
            return
        self._progress_end()

        if not ok:
            return

        def _save():
            sp, _ = QFileDialog.getSaveFileName(
                self, t["dialog_save_pdf"],
                os.path.join(os.path.expanduser("~"), "Desktop", default_name),
                "PDF Files (*.pdf)")
            if not sp:
                logger.info("log_cancelled", f"{t.get('menu_backup_punti')} — PDF")
                return
            try:
                shutil.copy2(tmp_path, sp)
                logger.info("log_pdf_saved", sp)
                if self._pdf_preview is not None:
                    self._pdf_preview.on_pdf_saved(os.path.dirname(os.path.abspath(sp)))
            except Exception as exc:
                logger.error("log_error_generic", str(exc))

        self._show_pdf_preview(tmp_path, on_save_cb=_save)

    def on_ga500_params(self):
        self._close_all_panels()
        self._hide_welcome()
        try:
            view = _GA500ParamsView(self.app_state, on_close_cb=self._close_ga500_view)
            self._ga500_view = view
            self.content_area.layout().addWidget(view)
            logger.info("log_ga500_opened")
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def on_motion_drive(self):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]

        file_path, _ = QFileDialog.getOpenFileName(
            self, t.get("drive_select_file", "Select .YDWIProj file"), "",
            "DriveWizard Project (*.YDWIProj *.YDWGProj);;All Files (*)"
        )
        if not file_path:
            logger.info("log_cancelled", "MotionDrive")
            return
        logger.info("log_btn_pressed", f"MotionDrive - file: {file_path}")

        try:
            from docs.drive import parse_ydwiproj
            info, params_by_cat = parse_ydwiproj(file_path)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        total = sum(len(v) for v in params_by_cat.values())
        logger.info("log_drive_total", total)

        file_name = os.path.basename(file_path)
        self._close_all_panels()

        try:
            view = _DriveView(
                file_name, info, params_by_cat, self.app_state,
                on_close_cb=self._close_drive_view,
                on_generate_excel_cb=lambda: self._on_drive_excel(
                    info, params_by_cat, file_name),
                on_export_pdf_cb=lambda: self._on_drive_pdf(
                    info, params_by_cat, file_name),
            )
            self._drive_view = view
            self.content_area.layout().addWidget(view)
            logger.info("log_drive_opened", file_name)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def _on_drive_excel(self, info, params_by_cat, file_name):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]

        stem = os.path.splitext(file_name)[0]
        default_name = f"Drive_{stem}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self, t["dialog_save_excel"],
            os.path.join(os.path.expanduser("~"), "Desktop", default_name),
            "Excel Files (*.xlsx)"
        )
        if not save_path:
            logger.info("log_cancelled", "MotionDrive - salva Excel")
            return

        try:
            from docs.drive import generate_drive_excel
            ok = generate_drive_excel(
                info, params_by_cat, save_path,
                col_param=t.get("drive_col_param", "Param"),
                col_name=t.get("drive_col_name", "Name"),
                col_value=t.get("drive_col_value", "Value"),
                col_default=t.get("drive_col_default", "Default"),
                col_min=t.get("drive_col_min", "Min"),
                col_max=t.get("drive_col_max", "Max"),
                col_prev=t.get("drive_col_prev", "Previous"),
                lang=self.app_state.language,
                log_fn=logger.info,
            )
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        if not ok:
            return

        try:
            os.startfile(os.path.dirname(os.path.abspath(save_path)))
        except Exception as exc:
            logger.warning("log_error_generic", str(exc))

    def _on_drive_pdf(self, info, params_by_cat, file_name):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]

        stem = os.path.splitext(file_name)[0]
        default_name = f"Drive_{stem}.pdf"
        save_path, _ = QFileDialog.getSaveFileName(
            self, t["dialog_save_pdf"],
            os.path.join(os.path.expanduser("~"), "Desktop", default_name),
            "PDF Files (*.pdf)"
        )
        if not save_path:
            logger.info("log_cancelled", "MotionDrive - salva PDF")
            return

        try:
            from docs.drive import generate_drive_pdf
            ok = generate_drive_pdf(
                info, params_by_cat, save_path,
                col_param=t.get("drive_col_param", "Param"),
                col_name=t.get("drive_col_name", "Name"),
                col_value=t.get("drive_col_value", "Value"),
                col_default=t.get("drive_col_default", "Default"),
                col_min=t.get("drive_col_min", "Min"),
                col_max=t.get("drive_col_max", "Max"),
                col_prev=t.get("drive_col_prev", "Previous"),
                toc_title=t.get("drive_toc_title", "Table of Contents"),
                info_section_title=t.get("drive_info_section", "Drive Information"),
                lang=self.app_state.language,
                log_fn=logger.info,
            )
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        if not ok:
            return

        try:
            os.startfile(os.path.dirname(os.path.abspath(save_path)))
        except Exception as exc:
            logger.warning("log_error_generic", str(exc))

    # ── Working folder (STEP 11) ──────────────────────────────────────────────

    @staticmethod
    def _config_path():
        appdata = os.environ.get('APPDATA', os.path.expanduser('~'))
        d = os.path.join(appdata, 'YaskawaTools')
        os.makedirs(d, exist_ok=True)
        return os.path.join(d, 'config.json')

    def _load_work_folder(self):
        try:
            import json, pathlib
            with open(self._config_path(), 'r', encoding='utf-8') as f:
                folder = json.load(f).get('work_folder', '') or ''
            if not folder:
                return ''
            folder = str(pathlib.Path(folder).resolve())
            return folder if os.path.isdir(folder) else ''
        except Exception:
            return ''

    def _save_work_folder(self, folder):
        try:
            import json
            with open(self._config_path(), 'w', encoding='utf-8') as f:
                json.dump({'work_folder': folder}, f)
        except Exception:
            pass

    def _on_choose_work_folder(self):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        folder = QFileDialog.getExistingDirectory(
            self, t.get("work_folder_btn", "Cartella robot"), "",
            QFileDialog.Option.ShowDirsOnly)
        if not folder:
            return
        self._work_folder = folder
        self._folder_lbl_path.setText(folder)
        self._save_work_folder(folder)
        logger.info("log_work_folder_changed", folder)

    # ── IP Net view (STEP 8) ──────────────────────────────────────────────────

    def on_ipnet(self):
        import tempfile
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_ipnet"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_ipnet')} — {folder}")

        try:
            from docs.ipnet import load_network_config, generate_ipnet_pdf
            rows = load_network_config(folder)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        folder_name = os.path.basename(os.path.normpath(folder))
        default_name = f"IPNet_{folder_name}.pdf"
        tmp_path = secure_paths.temp_path(default_name)
        self._progress_begin()
        try:
            ok = generate_ipnet_pdf(rows, tmp_path,
                                    folder_name=folder_name,
                                    lang=self.app_state.language,
                                    log_fn=logger.info)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            ok = False
        self._progress_end()

        if not ok:
            logger.warning("log_error_generic", f"{t.get('menu_ipnet', 'IP Net')} PDF")
            return

        import shutil

        def _save():
            sp, _ = QFileDialog.getSaveFileName(
                self, t.get("dialog_save_pdf", "Salva PDF"),
                os.path.join(os.path.expanduser("~"), "Desktop", default_name),
                "PDF Files (*.pdf)")
            if not sp:
                logger.info("log_cancelled", f"{t.get('menu_ipnet')} — PDF")
                return
            try:
                shutil.copy2(tmp_path, sp)
                logger.info("log_ipnet_pdf_saved", sp)
                if self._pdf_preview is not None:
                    self._pdf_preview.on_pdf_saved(os.path.dirname(os.path.abspath(sp)))
            except Exception as exc:
                logger.error("log_error_generic", str(exc))

        def _excel():
            xl_name = f"IPNet_{folder_name}.xlsx"
            sp, _ = QFileDialog.getSaveFileName(
                self, t.get("dialog_save_excel", "Salva Excel"),
                os.path.join(os.path.expanduser("~"), "Desktop", xl_name),
                "Excel Files (*.xlsx)")
            if not sp:
                logger.info("log_cancelled", f"{t.get('menu_ipnet')} — Excel")
                return
            try:
                from docs.ipnet import generate_ipnet_excel
                generate_ipnet_excel(rows, sp,
                                     folder_name=folder_name,
                                     lang=self.app_state.language,
                                     log_fn=logger.info)
                logger.info("log_ipnet_excel_saved", sp)
                os.startfile(os.path.dirname(os.path.abspath(sp)))
            except Exception as exc:
                logger.error("log_error_generic", str(exc))

        logger.info("log_ipnet_opened")
        self._show_pdf_preview(tmp_path, on_save_cb=_save, on_excel_cb=_excel)

    # ── LogData (STEP 10) ─────────────────────────────────────────────────────

    def on_logdata(self):
        import shutil
        import tempfile
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_logdata"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_logdata')} — {folder}")

        log_path = os.path.join(folder, "LOGDATA.DAT")
        if not os.path.isfile(log_path):
            logger.warning("log_logdata_no_data")
            return

        try:
            from docs.logdata import parse_logdata
            entries = parse_logdata(log_path)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            return

        if not entries:
            logger.warning("log_logdata_no_data")
            return

        folder_name = os.path.basename(os.path.normpath(folder))
        logger.info("log_logdata_opened", len(entries))
        default_name = f"LogData_{folder_name}.pdf"
        tmp_path = secure_paths.temp_path(default_name)

        self._progress_begin()
        try:
            from docs.logdata import generate_logdata_pdf
            ok = generate_logdata_pdf(entries, tmp_path, lang=self.app_state.language, log_fn=logger.info)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            self._progress_end()
            return
        self._progress_end()

        if not ok:
            return

        def _save():
            sp, _ = QFileDialog.getSaveFileName(
                self, t["dialog_save_pdf"],
                os.path.join(os.path.expanduser("~"), "Desktop", default_name),
                "PDF Files (*.pdf)")
            if not sp:
                logger.info("log_cancelled", f"{t.get('menu_logdata')} — PDF")
                return
            try:
                shutil.copy2(tmp_path, sp)
                logger.info("log_pdf_saved", sp)
                if self._pdf_preview is not None:
                    self._pdf_preview.on_pdf_saved(os.path.dirname(os.path.abspath(sp)))
            except Exception as exc:
                logger.error("log_error_generic", str(exc))

        self._show_pdf_preview(tmp_path, on_save_cb=_save)

    # ── Cubo interferenza (CUBEINTF.CND) ──────────────────────────────────────

    def on_cubeintf(self):
        import shutil
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_cubeintf", "Cubo interferenza"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_cubeintf', 'Cubo interferenza')} — {folder}")

        cube_path = os.path.join(folder, "CUBEINTF.CND")
        if not os.path.isfile(cube_path):
            logger.warning("log_cubeintf_no_file")
            return

        folder_name = os.path.basename(os.path.normpath(folder))
        default_name = f"CuboInterferenza_{folder_name}.pdf"
        tmp_path = secure_paths.temp_path(default_name)

        self._progress_begin()
        try:
            from docs.cubeintf import build_cubes, generate_pdf
            cubes = build_cubes(folder)
            ok = generate_pdf(folder, tmp_path,
                              lang=self.app_state.language, log_fn=logger.info)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            self._progress_end()
            return
        self._progress_end()

        if not ok:
            return

        logger.info("log_cubeintf_opened", len(cubes))

        def _save():
            sp, _ = QFileDialog.getSaveFileName(
                self, t["dialog_save_pdf"],
                os.path.join(os.path.expanduser("~"), "Desktop", default_name),
                "PDF Files (*.pdf)")
            if not sp:
                logger.info("log_cancelled", f"{t.get('menu_cubeintf', 'Cubo interferenza')} — PDF")
                return
            try:
                shutil.copy2(tmp_path, sp)
                logger.info("log_pdf_saved", sp)
                if self._pdf_preview is not None:
                    self._pdf_preview.on_pdf_saved(os.path.dirname(os.path.abspath(sp)))
            except Exception as exc:
                logger.error("log_error_generic", str(exc))

        def _regen():
            from docs.cubeintf import generate_pdf as _cg
            return _cg(folder, tmp_path,
                       lang=self.app_state.language, log_fn=logger.info)

        self._show_pdf_preview(tmp_path, on_save_cb=_save, regen_fn=_regen)

    # ── FormCut (FORMCUT.CND) ─────────────────────────────────────────────────

    def on_formcut(self):
        import shutil
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_formcut", "FormCut"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_formcut', 'FormCut')} — {folder}")

        fc_path = os.path.join(folder, "FORMCUT.CND")
        if not os.path.isfile(fc_path):
            logger.warning("log_formcut_no_file")
            return

        folder_name = os.path.basename(os.path.normpath(folder))
        default_name = f"FormCut_{folder_name}.pdf"
        tmp_path = secure_paths.temp_path(default_name)

        self._progress_begin()
        try:
            from docs.formcut import build_formcuts, generate_pdf
            formcuts = build_formcuts(folder)
            ok = generate_pdf(folder, tmp_path,
                              lang=self.app_state.language, log_fn=logger.info)
        except Exception as exc:
            logger.error("log_error_generic", str(exc))
            self._progress_end()
            return
        self._progress_end()

        if not ok:
            return

        logger.info("log_formcut_opened", len(formcuts))

        def _save():
            sp, _ = QFileDialog.getSaveFileName(
                self, t["dialog_save_pdf"],
                os.path.join(os.path.expanduser("~"), "Desktop", default_name),
                "PDF Files (*.pdf)")
            if not sp:
                logger.info("log_cancelled", f"{t.get('menu_formcut', 'FormCut')} — PDF")
                return
            try:
                shutil.copy2(tmp_path, sp)
                logger.info("log_pdf_saved", sp)
                if self._pdf_preview is not None:
                    self._pdf_preview.on_pdf_saved(os.path.dirname(os.path.abspath(sp)))
            except Exception as exc:
                logger.error("log_error_generic", str(exc))

        def _regen():
            from docs.formcut import generate_pdf as _fg
            return _fg(folder, tmp_path,
                       lang=self.app_state.language, log_fn=logger.info)

        self._show_pdf_preview(tmp_path, on_save_cb=_save, regen_fn=_regen)

    # ── Compila (STEP 5+6) ────────────────────────────────────────────────────

    def on_compila(self):
        from PySide6.QtWidgets import QFileDialog
        t = TRANSLATIONS[self.app_state.language]
        folder = self._work_folder or QFileDialog.getExistingDirectory(
            self, t["select_folder"], "", QFileDialog.Option.ShowDirsOnly)
        if not folder:
            logger.info("log_cancelled", t.get("menu_compila"))
            return
        logger.info("log_btn_pressed", f"{t.get('menu_compila')} — {folder}")

        var_dat = os.path.join(folder, "VAR.DAT")
        if not os.path.isfile(var_dat):
            logger.warning("compila_no_var_dat")
            return

        self._close_all_panels()
        self._hide_welcome()
        try:
            view = _CompilaView(folder, self.app_state, on_close_cb=self._close_compila_view)
            self._compila_view = view
            self.content_area.layout().addWidget(view)
            logger.info("log_compila_opened", os.path.basename(os.path.normpath(folder)))
        except Exception as exc:
            logger.error("log_error_generic", str(exc))

    def show_about(self):
        logger.info("log_about_opened")
        dialog = AboutDialog(self, self.app_state.is_dark_mode, self.app_state)
        dialog.exec()

    def closeEvent(self, event):
        super().closeEvent(event)

